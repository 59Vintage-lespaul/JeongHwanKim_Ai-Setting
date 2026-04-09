import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

src = r'C:\Users\bcave\Desktop\주간 회의 양식 신학기 마감.xlsx'
out = r'output\26FW\season-strategy\check_sales-analysis_신학기_2026-03-20.xlsx'

# ── 원본 데이터 로드
df26 = pd.read_excel(src, sheet_name='26신학기_누적수량', header=1)
df26.columns = ['key','매장명','품번','품명','유형','출고수량','실판매수량',
                '실판매율','TAG판가합','실판가합','매장구분','년도','신상이월','TAG가']
df26 = df26[df26['품번'].notna()].copy()
df26 = df26[df26['품번'].astype(str).str.match(r'^(WA|MG)\d')]
df26['출고수량'] = pd.to_numeric(df26['출고수량'], errors='coerce').fillna(0)
df26['실판매수량'] = pd.to_numeric(df26['실판매수량'], errors='coerce').fillna(0)
df26['TAG판가합'] = pd.to_numeric(df26['TAG판가합'], errors='coerce').fillna(0)
df26['실판가합'] = pd.to_numeric(df26['실판가합'], errors='coerce').fillna(0)
df26['TAG가'] = pd.to_numeric(df26['TAG가'], errors='coerce').fillna(0)

df25 = pd.read_excel(src, sheet_name='25신학기_누적', header=1)
df25.columns = ['key','매장명','품번','품명','출고','실판매','실판매율','매장구분','년도','신상이월']
df25 = df25[df25['품번'].notna()].copy()
df25 = df25[df25['품번'].astype(str).str.match(r'^(WA|MG)\d')]
df25['출고'] = pd.to_numeric(df25['출고'], errors='coerce').fillna(0)
df25['실판매'] = pd.to_numeric(df25['실판매'], errors='coerce').fillna(0)

유형맵 = pd.read_excel(src, sheet_name='유형구분', header=None, usecols=[2,3])
유형맵.columns = ['품명','유형']
유형맵 = 유형맵.dropna()
df25 = df25.merge(유형맵, on='품명', how='left')
df25['유형'] = df25['유형'].fillna('기타')

# ── 스타일 상수
DARK  = "1A1A2E"
BRAND = "E94560"
ACCENT= "16213E"
LIGHT = "F0F4F8"
WHITE = "FFFFFF"
GREEN_C = "27AE60"
RED_C   = "E74C3C"
GOLD  = "F39C12"

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bg=None, bold=False, fmt=None,
             halign='center', color=None, wrap=False, size=9):
    c = ws.cell(row=row, column=col, value=value)
    fc = color if color else '000000'
    c.font = Font(name='Arial', bold=bold, color=fc, size=size)
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    c.border = thin_border()
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def section_title(ws, row, merge_range, text, bg=BRAND):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(merge_range)
    c = ws[merge_range.split(':')[0]]
    c.value = f'▌ {text}'
    c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

def header_row(ws, row, headers, start_col=2, bg=ACCENT):
    ws.row_dimensions[row].height = 20
    for ci, h in enumerate(headers, start_col):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()

def chg_color(cell, val):
    if val > 0.05:
        cell.font = Font(name='Arial', color=GREEN_C, bold=True, size=9)
    elif val < -0.05:
        cell.font = Font(name='Arial', color=RED_C, size=9)
    else:
        cell.font = Font(name='Arial', size=9)

# 유형/채널 목록
유형목록 = ['메쉬스트링','스쿨백','아웃도어','우먼스','짐색','데일리','기타스트릿']
채널목록 = ['샵인샵','플래그쉽','온라인','면세점','해외']

# 집계
type_ch26 = df26.groupby(['유형','매장구분']).agg(실판=('실판가합','sum')).reset_index()
type_ch25 = df25.groupby(['유형','매장구분']).agg(실판=('실판매','sum')).reset_index()
total26 = 858756554
total25 = 1032438093

# SKU 집계
sku26 = df26.groupby(['품번','품명','유형','신상이월','TAG가']).agg(
    출고수량=('출고수량','sum'), 실판매수량=('실판매수량','sum'),
    TAG판합=('TAG판가합','sum'), 실판합=('실판가합','sum')
).reset_index()
sku26['판매율'] = np.where(sku26['출고수량']>0,
                           sku26['실판매수량']/sku26['출고수량']*100, np.nan)
sku26['할인율'] = np.where(sku26['TAG판합']>0,
                           (1-sku26['실판합']/sku26['TAG판합'])*100, np.nan)
sku26['기여율'] = sku26['실판합']/total26

def grade(row):
    pr = row['판매율']
    if pd.isna(pr) or row['출고수량'] <= 0: return 'SKIP'
    if pr >= 70: return 'MUST'
    if pr >= 40: return 'WATCH'
    return 'SKIP'

sku26['등급'] = sku26.apply(grade, axis=1)
신상 = sku26[sku26['신상이월']=='신상'].sort_values('실판합', ascending=False)
이월 = sku26[sku26['신상이월']=='이월'].sort_values('실판합', ascending=False).head(10)
grade_color = {'MUST': GREEN_C, 'WATCH': GOLD, 'SKIP': RED_C}

# ══════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── 시트1: 요약 ──────────────────
ws1 = wb.create_sheet('📊 요약')
ws1.sheet_view.showGridLines = False
for col, w in [('A',2),('B',14),('C',16),('D',16),('E',13),
               ('F',16),('G',16),('H',13),('I',13),('J',13)]:
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[1].height = 8

ws1.merge_cells('B2:J2')
c = ws1['B2']
c.value = '26년 신학기 가방 판매 분석 — 27SS 기획참고용'
c.font = Font(name='Arial', bold=True, color=WHITE, size=14)
c.fill = PatternFill('solid', fgColor=DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 32

ws1.merge_cells('B3:J3')
c = ws1['B3']
c.value = '분석 기간: 2025년 12월 1일 ~ 2026년 3월 2일  |  기준: 실판 금액  |  와키윌리 브랜드'
c.font = Font(name='Arial', color='AAAAAA', size=9)
c.fill = PatternFill('solid', fgColor=ACCENT)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[3].height = 16
ws1.row_dimensions[4].height = 8

# 신상/이월 섹션
section_title(ws1, 5, 'B5:J5', '전체 실적 요약 (신상/이월 구분)')
header_row(ws1, 6, ['구분','25년 TAG판','25년 실판','25년 할인율',
                    '26년 TAG판','26년 실판','26년 할인율','실판 증감액','실판 증감율'])
kpi = [
    ('신상',1142655000,892789933,0.2187,820531000,583618890,0.2887),
    ('이월',236604000,139539160,0.4102,382472000,275137664,0.2806),
    ('합계',1379259000,1032329093,0.2515,1203003000,858756554,0.2862),
]
for ri, (구분,t25,r25,d25,t26,r26,d26) in enumerate(kpi, 7):
    ws1.row_dimensions[ri].height = 20
    bg = 'FFF3CD' if 구분=='합계' else ('F0F4F8' if ri%2==0 else WHITE)
    bold = (구분=='합계')
    chg = (r26-r25)/r25
    for ci, (v, fmt) in enumerate(zip(
        [구분,t25,r25,d25,t26,r26,d26,r26-r25,chg],
        [None,'#,##0','#,##0','0.0%','#,##0','#,##0','0.0%','#,##0','0.0%']
    ), 2):
        set_cell(ws1, ri, ci, v, bg=bg, bold=bold, fmt=fmt)
    chg_c = ws1.cell(row=ri, column=10)
    chg_color(chg_c, chg)

ws1.row_dimensions[10].height = 8

# 채널별 섹션
section_title(ws1, 11, 'B11:J11', '채널별 실판 비교')
header_row(ws1, 12, ['채널','25년 TAG판','25년 실판','25년 할인율',
                     '26년 TAG판','26년 실판','26년 할인율','실판 증감액','실판 증감율'])
ch_kpi = [
    ('샵인샵',753206000,611567076,0.1880,772235000,590230120,0.2357),
    ('플래그쉽',23452000,19544515,0.1666,59723000,54384775,0.0894),
    ('온라인',491688000,333465982,0.3218,262992000,158174729,0.3986),
    ('면세점',107164000,65538200,0.3884,30604000,27045750,0.1163),
    ('해외',3749000,2213320,0.4096,77449000,28921180,0.6266),
    ('합계',1379259000,1032329093,0.2515,1203003000,858756554,0.2862),
]
for ri, (ch,t25,r25,d25,t26,r26,d26) in enumerate(ch_kpi, 13):
    ws1.row_dimensions[ri].height = 20
    bg = 'FFF3CD' if ch=='합계' else ('F0F4F8' if ri%2==0 else WHITE)
    bold = (ch=='합계')
    chg = (r26-r25)/r25 if r25 else 0
    for ci, (v, fmt) in enumerate(zip(
        [ch,t25,r25,d25,t26,r26,d26,r26-r25,chg],
        [None,'#,##0','#,##0','0.0%','#,##0','#,##0','0.0%','#,##0','0.0%']
    ), 2):
        set_cell(ws1, ri, ci, v, bg=bg, bold=bold, fmt=fmt)
    chg_color(ws1.cell(row=ri, column=10), chg)

ws1.row_dimensions[19].height = 8

# 유형별 비중
section_title(ws1, 20, 'B20:J20', '유형별 실판 비중 변화')
header_row(ws1, 21, ['유형','25년 실판','25년 비중','26년 실판','26년 비중',
                     '비중변화','실판 증감율','',''])
type_kpi = [
    ('메쉬스트링',343251865,0.3325,268604005,0.3128),
    ('스쿨백',500719702,0.4850,429613433,0.5003),
    ('아웃도어',9030000,0.0087,11578612,0.0135),
    ('우먼스',54915674,0.0532,81880507,0.0953),
    ('짐색',4333296,0.0042,3465428,0.0040),
    ('데일리',18529938,0.0179,29998691,0.0349),
    ('기타스트릿',101657618,0.0985,33615878,0.0392),
    ('합계',1032438093,1.0,858756554,1.0),
]
for ri, (유형,r25,b25,r26,b26) in enumerate(type_kpi, 22):
    ws1.row_dimensions[ri].height = 20
    bg = 'FFF3CD' if 유형=='합계' else ('F0F4F8' if ri%2==0 else WHITE)
    bold = (유형=='합계')
    bchg = b26-b25
    rchg = (r26-r25)/r25 if r25 else 0
    for ci, (v, fmt) in enumerate(zip(
        [유형,r25,b25,r26,b26,bchg,rchg,'',''],
        [None,'#,##0','0.0%','#,##0','0.0%','0.0%','0.0%',None,None]
    ), 2):
        set_cell(ws1, ri, ci, v, bg=bg, bold=bold, fmt=fmt)
    bc = ws1.cell(row=ri, column=7)
    if bchg > 0.005: bc.font = Font(name='Arial', color=GREEN_C, bold=True, size=9)
    elif bchg < -0.005: bc.font = Font(name='Arial', color=RED_C, size=9)

print("시트1 완료")

# ── 시트2: SKU 랭킹 ─────────────
ws2 = wb.create_sheet('🎯 SKU 랭킹')
ws2.sheet_view.showGridLines = False
for col, w in [('A',2),('B',5),('C',14),('D',20),('E',12),('F',8),
               ('G',11),('H',11),('I',13),('J',10),('K',9),('L',10)]:
    ws2.column_dimensions[col].width = w
ws2.row_dimensions[1].height = 8

ws2.merge_cells('B2:L2')
c = ws2['B2']
c.value = 'SKU별 판매 성과 랭킹'
c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
c.fill = PatternFill('solid', fgColor=DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[2].height = 28
ws2.row_dimensions[3].height = 8

sku_hdrs = ['#','품번','품명','유형','TAG가','출고수','판매수','판매율','실판(원)','할인율','기여율','등급']

# 신상
section_title(ws2, 4, 'B4:L4', '26년 신상 SKU 판매 성과')
header_row(ws2, 5, sku_hdrs)
for ri_off, (_, row) in enumerate(신상.iterrows()):
    ri = ri_off + 6
    ws2.row_dimensions[ri].height = 22
    bg = 'F0F4F8' if ri%2==0 else WHITE
    pr = row['판매율']/100 if pd.notna(row['판매율']) else None
    dr = row['할인율']/100 if pd.notna(row['할인율']) else None
    for ci, (v, fmt) in enumerate(zip(
        [ri_off+1, row['품번'], row['품명'], row['유형'],
         int(row['TAG가']), int(row['출고수량']), int(row['실판매수량']),
         pr, int(row['실판합']), dr, row['기여율'], row['등급']],
        [None,None,None,None,'#,##0','#,##0','#,##0','0.0%','#,##0','0.0%','0.0%',None]
    ), 2):
        set_cell(ws2, ri, ci, v, bg=bg, fmt=fmt)
    gc = grade_color.get(row['등급'], '95A5A6')
    g_c = ws2.cell(row=ri, column=13)
    g_c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
    g_c.fill = PatternFill('solid', fgColor=gc)
    if pd.notna(row['판매율']):
        pr_c = ws2.cell(row=ri, column=9)
        if row['판매율'] >= 70: pr_c.font = Font(name='Arial', color=GREEN_C, bold=True, size=9)
        elif row['판매율'] >= 40: pr_c.font = Font(name='Arial', color=GOLD, size=9)
        else: pr_c.font = Font(name='Arial', color=RED_C, size=9)

next_r = 6 + len(신상) + 2
ws2.row_dimensions[next_r-1].height = 8

# 이월
section_title(ws2, next_r, f'B{next_r}:L{next_r}', '26년 이월 SKU 소화 성과 (실판 상위 10개)', bg='5D6D7E')
header_row(ws2, next_r+1, sku_hdrs, bg='7F8C8D')
for ri_off, (_, row) in enumerate(이월.iterrows()):
    ri = next_r + 2 + ri_off
    ws2.row_dimensions[ri].height = 22
    bg = 'F0F4F8' if ri%2==0 else WHITE
    pr = row['판매율']/100 if pd.notna(row['판매율']) and row['출고수량'] > 0 else None
    dr = row['할인율']/100 if pd.notna(row['할인율']) else None
    for ci, (v, fmt) in enumerate(zip(
        [ri_off+1, row['품번'], row['품명'], row['유형'],
         int(row['TAG가']), int(row['출고수량']), int(row['실판매수량']),
         pr, int(row['실판합']), dr, row['기여율'], row['등급']],
        [None,None,None,None,'#,##0','#,##0','#,##0','0.0%','#,##0','0.0%','0.0%',None]
    ), 2):
        set_cell(ws2, ri, ci, v, bg=bg, fmt=fmt)
    gc = grade_color.get(row['등급'], '95A5A6')
    g_c = ws2.cell(row=ri, column=13)
    g_c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
    g_c.fill = PatternFill('solid', fgColor=gc)

print("시트2 완료")

# ── 시트3: 유형별 분석 ───────────
ws3 = wb.create_sheet('🏷️ 유형별')
ws3.sheet_view.showGridLines = False
for col, w in [('A',2),('B',14),('C',13),('D',13),('E',13),
               ('F',13),('G',13),('H',11),('I',11)]:
    ws3.column_dimensions[col].width = w
ws3.row_dimensions[1].height = 8

ws3.merge_cells('B2:I2')
c = ws3['B2']
c.value = '유형별 × 채널별 실판 분석'
c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
c.fill = PatternFill('solid', fgColor=DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[2].height = 28
ws3.row_dimensions[3].height = 8

type_ch_hdrs = ['유형'] + 채널목록 + ['합계','비중']

def write_type_ch_section(ws, start_row, year_df, total, bg_hdr=ACCENT, title_text='', title_bg=BRAND):
    section_title(ws, start_row, f'B{start_row}:I{start_row}', title_text, bg=title_bg)
    header_row(ws, start_row+1, type_ch_hdrs, bg=bg_hdr)
    for ri, 유형 in enumerate(유형목록 + ['합계'], start_row+2):
        ws.row_dimensions[ri].height = 20
        bg = 'FFF3CD' if 유형=='합계' else ('F0F4F8' if ri%2==0 else WHITE)
        bold = (유형=='합계')
        row_total = 0
        vals = [유형]
        for ch in 채널목록:
            if 유형 == '합계':
                v = year_df[year_df['매장구분']==ch]['실판'].sum()
            else:
                m = (year_df['유형']==유형)&(year_df['매장구분']==ch)
                v = year_df[m]['실판'].sum() if m.any() else 0
            vals.append(int(v))
            row_total += v
        vals.extend([int(row_total), row_total/total if total else 0])
        fmts = [None,'#,##0','#,##0','#,##0','#,##0','#,##0','#,##0','0.0%']
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 2):
            set_cell(ws, ri, ci, v, bg=bg, bold=bold, fmt=fmt)
    return start_row + 2 + len(유형목록) + 1

r = write_type_ch_section(ws3, 4, type_ch26, total26, title_text='26년 유형×채널 실판 (원)')
ws3.row_dimensions[r].height = 8
r += 1
r = write_type_ch_section(ws3, r, type_ch25, total25,
                           bg_hdr='7F8C8D', title_text='25년 유형×채널 실판 (원)', title_bg='5D6D7E')
ws3.row_dimensions[r].height = 8
r += 1

# 증감율 섹션
section_title(ws3, r, f'B{r}:I{r}', '전년 대비 증감율 (초록=증가 / 빨강=감소)')
header_row(ws3, r+1, type_ch_hdrs)
for ri, 유형 in enumerate(유형목록 + ['합계'], r+2):
    ws3.row_dimensions[ri].height = 20
    bg = 'FFF3CD' if 유형=='합계' else ('F0F4F8' if ri%2==0 else WHITE)
    set_cell(ws3, ri, 2, 유형, bg=bg, bold=(유형=='합계'))
    row26t = 0; row25t = 0
    for ci_off, ch in enumerate(채널목록):
        ci = ci_off + 3
        if 유형 == '합계':
            v26 = type_ch26[type_ch26['매장구분']==ch]['실판'].sum()
            v25 = type_ch25[type_ch25['매장구분']==ch]['실판'].sum()
        else:
            m26 = (type_ch26['유형']==유형)&(type_ch26['매장구분']==ch)
            m25 = (type_ch25['유형']==유형)&(type_ch25['매장구분']==ch)
            v26 = type_ch26[m26]['실판'].sum() if m26.any() else 0
            v25 = type_ch25[m25]['실판'].sum() if m25.any() else 0
        row26t += v26; row25t += v25
        chg = (v26-v25)/v25 if v25 > 0 else (1.0 if v26 > 0 else 0.0)
        c2 = set_cell(ws3, ri, ci, chg, bg=bg, fmt='0.0%')
        chg_color(c2, chg)
    tot_chg = (row26t-row25t)/row25t if row25t > 0 else 0
    c2 = set_cell(ws3, ri, 8, tot_chg, bg=bg, fmt='0.0%')
    chg_color(c2, tot_chg)
    b26v = row26t/total26 if total26 else 0
    b25v = row25t/total25 if total25 else 0
    c2 = set_cell(ws3, ri, 9, b26v-b25v, bg=bg, fmt='0.0%')
    chg_color(c2, b26v-b25v)

print("시트3 완료")

# ── 시트4: 채널별 ───────────────
ws4 = wb.create_sheet('📡 채널별')
ws4.sheet_view.showGridLines = False
for col, w in [('A',2),('B',12),('C',14),('D',14),('E',11),
               ('F',14),('G',14),('H',11),('I',11)]:
    ws4.column_dimensions[col].width = w
ws4.row_dimensions[1].height = 8

ws4.merge_cells('B2:I2')
c = ws4['B2']
c.value = '채널별 × 유형별 상세 분석'
c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
c.fill = PatternFill('solid', fgColor=DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[2].height = 28

row_cur = 4
for ch in 채널목록:
    v25t = type_ch25[type_ch25['매장구분']==ch]['실판'].sum()
    v26t = type_ch26[type_ch26['매장구분']==ch]['실판'].sum()
    chg_pct = (v26t-v25t)/v25t*100 if v25t > 0 else 0
    sign = '+' if chg_pct >= 0 else ''
    ws4.row_dimensions[row_cur].height = 22
    ws4.merge_cells(f'B{row_cur}:I{row_cur}')
    c = ws4[f'B{row_cur}']
    c.value = f'▌ {ch}  |  25년 {v25t:,.0f}원  →  26년 {v26t:,.0f}원  ({sign}{chg_pct:.1f}%)'
    c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill = PatternFill('solid', fgColor=BRAND)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row_cur += 1

    header_row(ws4, row_cur, ['유형','25년 실판','25년 비중','26년 실판',
                               '26년 비중','증감액','증감율','비중변화'])
    row_cur += 1

    for 유형 in 유형목록:
        ws4.row_dimensions[row_cur].height = 20
        bg = 'F0F4F8' if row_cur%2==0 else WHITE
        m26 = (type_ch26['유형']==유형)&(type_ch26['매장구분']==ch)
        m25 = (type_ch25['유형']==유형)&(type_ch25['매장구분']==ch)
        v26 = type_ch26[m26]['실판'].sum() if m26.any() else 0
        v25 = type_ch25[m25]['실판'].sum() if m25.any() else 0
        b26 = v26/v26t if v26t else 0
        b25 = v25/v25t if v25t else 0
        chg_v = (v26-v25)/v25 if v25 > 0 else (1.0 if v26 > 0 else 0.0)
        for ci, (v, fmt) in enumerate(zip(
            [유형, int(v25), b25, int(v26), b26, int(v26-v25), chg_v, b26-b25],
            [None,'#,##0','0.0%','#,##0','0.0%','#,##0','0.0%','0.0%']
        ), 2):
            set_cell(ws4, row_cur, ci, v, bg=bg, fmt=fmt)
        chg_color(ws4.cell(row=row_cur, column=8), chg_v)
        bchg_c = ws4.cell(row=row_cur, column=9)
        if b26-b25 > 0.005: bchg_c.font = Font(name='Arial', color=GREEN_C, bold=True, size=9)
        elif b26-b25 < -0.005: bchg_c.font = Font(name='Arial', color=RED_C, size=9)
        row_cur += 1

    # 합계
    ws4.row_dimensions[row_cur].height = 20
    tot_chg = (v26t-v25t)/v25t if v25t > 0 else 0
    for ci, (v, fmt) in enumerate(zip(
        ['합계', int(v25t), 1.0, int(v26t), 1.0, int(v26t-v25t), tot_chg, 0.0],
        [None,'#,##0','0.0%','#,##0','0.0%','#,##0','0.0%','0.0%']
    ), 2):
        set_cell(ws4, row_cur, ci, v, bg='FFF3CD', bold=True, fmt=fmt)
    row_cur += 2

print("시트4 완료")

# ── 시트5: 27SS 인사이트 ─────────
ws5 = wb.create_sheet('💡 27SS 인사이트')
ws5.sheet_view.showGridLines = False
for col, w in [('A',2),('B',18),('C',52),('D',24)]:
    ws5.column_dimensions[col].width = w
ws5.row_dimensions[1].height = 8

ws5.merge_cells('B2:D2')
c = ws5['B2']
c.value = '27SS 기획 인사이트 — 26년 신학기 판매 기반'
c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
c.fill = PatternFill('solid', fgColor=DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[2].height = 30

ws5.merge_cells('B3:D3')
c = ws5['B3']
c.value = '기간: 2025.12.01 ~ 2026.03.02 마감  |  전년(25년) 동기 대비  |  ✅ 기회  |  ⚠️ 리스크  |  📦 SKU 제언'
c.font = Font(name='Arial', color='AAAAAA', size=9)
c.fill = PatternFill('solid', fgColor=ACCENT)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[3].height = 16

insights = [
    ('✅ TOP 3 기회', '① 우먼스 카테고리 급성장',
     '실판 5,491만 → 8,188만원 (+49.1%), 비중 5.3% → 9.5% (+4.2%p)\n릴리와펜 코듀로이 이월 소화율 221%, 데일리 이월 소화율 610%\n→ 수요 대비 공급이 극도로 부족했던 카테고리',
     '27SS 우먼스 신상 SKU 2~3개로 확대\n릴리 시리즈 후속 기획 우선순위 최상위\n발주 보수적 + 시즌 중반 QR 전략 병행'),
    ('✅ TOP 3 기회', '② 플래그쉽 채널 폭발 성장',
     '실판 1,954만 → 5,438만원 (+178.3%)\n명동 플래그쉽 오픈 효과, 스쿨백+메쉬스트링 중심\n방문 고객 구매 전환율이 높음을 확인',
     '27SS 플래그쉽 전용 컬러/한정 SKU 검토\n입점 SKU 수를 줄이되 단가 높은 제품 중심\nVM 및 스토리텔링 강화'),
    ('✅ TOP 3 기회', '③ 해외 채널 신규 개척',
     '실판 221만 → 2,892만원 (+1,207%)\n대만 B.CAVE TAIWAN: 케이브 백팩·플라이트 V2 알파 검증\n전량 판매(판매율 100%) 확인',
     '27SS 해외 전용 물량 별도 계획 수립\n케이브 백팩 우선 확보, 플라이트 V2 알파 동반\n해외 확장 채널 추가 검토'),
    ('⚠️ 리스크', '온라인 채널 급감',
     '실판 3.33억 → 1.58억원 (-52.6%)\n25년 히트: 메쉬 스트링 백팩(WA2501BP03) → 26년 후속 미확보\n에이블리 판매 미미, 무신사 집중 구조',
     '27SS 온라인 히트 기대 SKU 사전 물량 확보 필수\n무신사/에이블리 전용 컬러 or 번들 기획 검토\n케이브 백팩 온라인 판매율 제고 방안 필요'),
    ('⚠️ 리스크', '신상 전체 실판 -34.6% 급감',
     'TAG판 1,142억 → 820억 (-28.2%), 실판 893억 → 584억 (-34.6%)\n26년 신상 5개 SKU 중 판매율 70% 이상 없음\n플라이트 V2 33L 41.5% 부진 — 과잉 발주 가능성',
     '27SS 신상 발주 보수적으로 접근\n플라이트 V2 사이즈 통합(33L 드랍 검토)\nSKU 수 줄이고 히트 아이템 집중'),
    ('⚠️ 리스크', '면세점·기타스트릿 급락',
     '면세점 -58.7%, 기타스트릿 비중 9.8% → 3.9% (-5.9%p)\n코듀라 웨빙 백팩 이월 소화 완료, 신상 공백 상태',
     '면세점 발주 축소 (리스크 헤지)\n기타스트릿 27SS 신상 최소화 (1개 이내)\n코듀라 소재 수요는 메쉬스트링/스쿨백에 흡수'),
    ('📦 SKU 제언', '케이브 백팩 — MUST → QR 최우선',
     '26년 신상 실판 1위 (2.23억원), 전 채널 고른 판매\n판매율 61.8% — 추가 물량 소화 가능성 높음\n해외 반응 확인(대만 100% 소화)',
     '27SS 후속 컬러 or IP 콜라보 QR 최우선 진행\n해외용 수량 별도 확보\n목표 판매율 70%+ 설정'),
    ('📦 SKU 제언', '플라이트 V2 라인 — 구조 재검토',
     '28L 60.4% / 알파 53.7% / 33L 41.5%\n33L 부진 — 발주 과잉 또는 수요 미확인 사이즈\n플래그쉽·해외에서는 알파가 강세',
     '27SS V3 기획 시 사이즈 2개(28L·알파)로 통합 검토\n33L 드랍 또는 물량 대폭 축소\n가격/컬러/IP 강화로 판매율 70%+ 목표'),
    ('📦 SKU 제언', '우먼스 신상 — 릴리 계승 기획',
     '26년 신상 우먼스(카고·릴리) 판매율 5~12% 극저\n반면 이월 우먼스 소화율 200%+ — 신상 발굴 기회\n릴리와펜 시리즈 팬층 형성 확인',
     '릴리 시리즈 후속 신상 2개 기획 (디자인 계승)\n발주 소량 + 시즌 중반 QR 전략\n신상 우먼스 가격대 검토 (89,000 → 99,000 테스트)'),
]

row_cur = 5
prev_section = None
sec_colors = {
    '✅ TOP 3 기회': '27AE60',
    '⚠️ 리스크': 'E74C3C',
    '📦 SKU 제언': '2980B9',
}

for (section, item, content, direction) in insights:
    if section != prev_section:
        ws5.row_dimensions[row_cur].height = 10
        row_cur += 1
        ws5.row_dimensions[row_cur].height = 22
        ws5.merge_cells(f'B{row_cur}:D{row_cur}')
        c = ws5[f'B{row_cur}']
        c.value = f'▌ {section}'
        c.font = Font(name='Arial', bold=True, color=WHITE, size=11)
        c.fill = PatternFill('solid', fgColor=sec_colors.get(section, DARK))
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        row_cur += 1
        ws5.row_dimensions[row_cur].height = 18
        for ci, h in enumerate(['항목','인사이트 (수치 근거)','27SS 기획 방향'], 2):
            cc = ws5.cell(row=row_cur, column=ci, value=h)
            cc.font = Font(name='Arial', bold=True, color=WHITE, size=9)
            cc.fill = PatternFill('solid', fgColor=ACCENT)
            cc.alignment = Alignment(horizontal='center', vertical='center')
            cc.border = thin_border()
        row_cur += 1
        prev_section = section

    lines = max(content.count('\n')+1, direction.count('\n')+1)
    h = max(28, lines*16+8)
    ws5.row_dimensions[row_cur].height = h
    bg = 'F0F4F8' if row_cur%2==0 else WHITE

    c_item = ws5.cell(row=row_cur, column=2, value=item)
    c_item.font = Font(name='Arial', bold=True, size=9)
    c_item.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    c_item.border = thin_border()
    c_item.fill = PatternFill('solid', fgColor=bg)

    c_cont = ws5.cell(row=row_cur, column=3, value=content)
    c_cont.font = Font(name='Arial', size=9)
    c_cont.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    c_cont.border = thin_border()
    c_cont.fill = PatternFill('solid', fgColor=bg)

    c_dir = ws5.cell(row=row_cur, column=4, value=direction)
    c_dir.font = Font(name='Arial', color='155724', size=9)
    c_dir.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    c_dir.border = thin_border()
    c_dir.fill = PatternFill('solid', fgColor='D4EDDA')

    row_cur += 1

print("시트5 완료")

wb.save(out)
print(f"\n✅ 저장 완료: {out}")
