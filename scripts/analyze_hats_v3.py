#!/usr/bin/env python3
"""
25SS 모자 판매 분석 v3 — 심층 속성 분석 + 판매 예측
입력: 판매자료 + 구조화설계 + 월별판매추이(신규파일)
출력: check_sales-analysis_v3_2026-03-19.xlsx (14시트) + check_sales-analysis_v3_2026-03-19.html
"""

import pandas as pd
import numpy as np
import json
import warnings
import sys
import os
from datetime import datetime, timedelta

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = open(sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1)

warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 경로 ────────────────────────────────────────────────────────────────────
BASE        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SALES_PATH  = os.path.join(BASE, "output/26FW/season-strategy/품번종합집계표_모자 판매자료.xlsx")
STRUCT_PATH = os.path.join(BASE, "output/26FW/season-strategy/모자 구조화 설계 최종.xlsx")
MONTHLY_PATH= os.path.join(BASE, "output/26FW/season-strategy/모자 월별 판매추이분석.xlsx")
OUT_XLSX    = os.path.join(BASE, "output/26FW/season-strategy/check_sales-analysis_v3_2026-03-19.xlsx")
OUT_HTML    = os.path.join(BASE, "output/26FW/season-strategy/check_sales-analysis_v3_2026-03-19.html")

ANALYSIS_DATE = datetime(2026, 3, 19)

MONTHS = [
    "2025-01","2025-02","2025-03","2025-04","2025-05","2025-06",
    "2025-07","2025-08","2025-09","2025-10","2025-11","2025-12",
    "2026-01","2026-02","2026-03"
]
MONTHLY_QTY_IDX = {MONTHS[i]: 13 + i * 3 for i in range(15)}

# ─── 스타일 팔레트 ────────────────────────────────────────────────────────────
C_HEADER = "1A1A2E"; C_ACCENT = "E94560"; C_SUB = "16213E"; C_LIGHT = "F0F4FF"
C_MUST   = "D4EDDA"; C_SKIP   = "F8D7DA"; C_WATCH = "FFF3CD"
C_WHITE  = "FFFFFF"; C_DARK   = "1A1A2E"; C_BORDER = "CCCCCC"

def fill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, color=C_DARK, size=10):
    return Font(name="맑은 고딕", bold=bold, color=color, size=size)
def brd():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def set_widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr_row(ws, r, vals, bg=C_HEADER, fg=C_WHITE):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = fill(bg); c.font = fnt(True, fg); c.alignment = ctr(); c.border = brd()

def data_row(ws, r, vals, bg=None):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        if bg: c.fill = fill(bg)
        c.font = fnt(); c.alignment = ctr(); c.border = brd()

def row_bg(판별):
    if "MUST" in str(판별): return C_MUST
    if "SKIP" in str(판별): return C_SKIP
    if "WATCH" in str(판별): return C_WATCH
    return None

def heat_color(val, mx):
    if mx == 0 or pd.isna(val) or val == 0: return "FAFBFF"
    r = val / mx
    R = int(250 + (233-250)*r); G = int(251 + (69-251)*r); B = int(255 + (96-255)*r)
    return f"{R:02X}{G:02X}{B:02X}"

def sheet_title(ws, text, ncols, bg=C_HEADER):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = text
    c.fill = fill(bg); c.font = fnt(True, C_WHITE, 13); c.alignment = ctr()
    ws.row_dimensions[1].height = 30
    ws.sheet_view.showGridLines = False

# ════════════════════════════════════════════════════════════════════════════
# 1. 데이터 로드 & 정제
# ════════════════════════════════════════════════════════════════════════════
print("[1/10] 데이터 로드...")

sales_raw  = pd.read_excel(SALES_PATH, header=1)
struct_raw = pd.read_excel(STRUCT_PATH, sheet_name="품번별 세부구분", header=0)
monthly_raw= pd.read_excel(MONTHLY_PATH, header=1)

# ── 판매자료 컬럼 표준화
SCOLS = {
    0:"품번", 1:"품명", 2:"색상", 6:"최초판매가", 7:"현판매가",
    10:"시즌", 14:"최초입고일", 15:"최종입고일",
    16:"최초출고일", 17:"최종출고일", 20:"발주_수량",
    23:"최초출고월", 26:"누계판매_수량", 27:"누계판매_금액",
    31:"누계판매_실판매금액", 34:"완사입비_단가", 35:"판매이익_금액",
    37:"판매이익율_사전V",
}
sales = sales_raw.copy()
sales.columns = [SCOLS.get(i, f"_c{i}") for i in range(len(sales.columns))]
sales = sales[sales["품번"].astype(str).str.startswith("WA")].copy()

# ── 구조화 컬럼 표준화
KCOLS = {0:"품번", 1:"품명_구조", 2:"대분류", 3:"활용",
         4:"로고사이즈", 5:"표현방식", 6:"스타일분류", 7:"볼캡유형",
         8:"브림", 9:"크라운", 10:"패널수", 11:"소재", 12:"소재코드"}
struct = struct_raw.copy()
struct.columns = [KCOLS.get(i, f"_k{i}") for i in range(len(struct.columns))]

# ── 월별파일 컬럼 표준화 (신규 파일 구조)
mf = monthly_raw.copy()
MCOLS_BASE = {
    0:"품번", 1:"색상", 2:"초기판매수량", 3:"기간판매수량",
    4:"발주완료일자", 5:"판매시작일자", 6:"기초", 7:"입고",
    8:"반품", 9:"판매", 10:"판매율_raw", 11:"재고", 12:"누적판매"
}
for idx_pos, col_name in MCOLS_BASE.items():
    if idx_pos < len(mf.columns):
        mf.rename(columns={mf.columns[idx_pos]: col_name}, inplace=True, errors="ignore")

for i, m in enumerate(MONTHS):
    qi = 13 + i * 3
    if qi < len(mf.columns):
        mf.rename(columns={mf.columns[qi]: f"qty_{m}"}, inplace=True, errors="ignore")

# 숫자 변환
for m in MONTHS:
    col = f"qty_{m}"
    if col in mf.columns:
        mf[col] = pd.to_numeric(mf[col], errors="coerce").fillna(0)
for col in ["기초","입고","반품","판매","재고","누적판매"]:
    if col in mf.columns:
        mf[col] = pd.to_numeric(mf[col], errors="coerce").fillna(0)

# ── 25SS 모자 필터 (WA25 + CA/HT/BN)
hat_mask = (mf["품번"].astype(str).str.startswith("WA25") &
            (mf["품번"].astype(str).str.contains("CA") |
             mf["품번"].astype(str).str.contains("HT") |
             mf["품번"].astype(str).str.contains("BN")))
monthly = mf[hat_mask].copy()
monthly = monthly[monthly["품번"].notna()].copy()

# ── 26SS 필터 (WA26 + CA/HT/BN)
wa26_mask = (mf["품번"].astype(str).str.startswith("WA26") &
             (mf["품번"].astype(str).str.contains("CA") |
              mf["품번"].astype(str).str.contains("HT") |
              mf["품번"].astype(str).str.contains("BN")))
monthly_26 = mf[wa26_mask].copy()
monthly_26 = monthly_26[monthly_26["품번"].notna()].copy()

print(f"   판매: {len(sales)}행 / 구조화: {len(struct)}행 / 월별(25SS): {len(monthly)}행 / 월별(26SS): {len(monthly_26)}행")

# ════════════════════════════════════════════════════════════════════════════
# 2. JOIN — 판매 + 구조화
# ════════════════════════════════════════════════════════════════════════════
print("[2/10] JOIN 처리...")

SCOLS_SEL = ["품번","대분류","활용","로고사이즈","표현방식","스타일분류","볼캡유형","브림","크라운","패널수","소재","소재코드"]
df = sales.merge(struct[SCOLS_SEL], on="품번", how="left")

for col in ["최초입고일","최종입고일","최초출고일","최종출고일"]:
    df[col] = pd.to_datetime(df[col], errors="coerce")

df["판매경과일"] = (df["최종출고일"] - df["최초출고일"]).dt.days
df["할인여부"]   = df["최초판매가"] != df["현판매가"]

COLOR_BASIC  = {"BK","NA","CH","GY","WH"}
COLOR_NEUTRL = {"BE","CR","IB","KH","OL","MT"}
COLOR_EARTH  = {"BR","TN","CA"}
def color_grp(c):
    c = str(c).upper().strip()
    if c in COLOR_BASIC:  return "베이직(BK/NA/GY)"
    if c in COLOR_NEUTRL: return "뉴트럴(BE/CR)"
    if c in COLOR_EARTH:  return "어스(BR/TN)"
    return "컬러풀"

df["색상계열"] = df["색상"].apply(color_grp)
df["출고월"]   = df["최초출고월"].astype(str).str[:7]

unmatched = df[df["대분류"].isna()]["품번"].unique()
if len(unmatched): print(f"   ⚠ 미매칭 {len(unmatched)}개: {list(unmatched)[:5]}")
else: print("   ✓ 전체 매칭 완료")

# ════════════════════════════════════════════════════════════════════════════
# 3. 월별 데이터 집계 (25SS)
# ════════════════════════════════════════════════════════════════════════════
print("[3/10] 월별 데이터 집계...")

qty_cols = [f"qty_{m}" for m in MONTHS]
for col in qty_cols:
    if col not in monthly.columns:
        monthly[col] = 0

monthly_sku = monthly.groupby("품번")[qty_cols].sum().reset_index()
monthly_sku.columns = ["품번"] + MONTHS

def peak_month(row):
    series = row[MONTHS]
    if series.sum() == 0: return "N/A"
    return series.idxmax()

def ss_ratio(row):
    ss_months = [m for m in MONTHS if m.startswith("2025-0")]
    total = row[MONTHS].sum()
    if total == 0: return 0
    return round(row[ss_months].sum() / total * 100, 1)

def sales_concentration(row):
    series = row[MONTHS].values.astype(float)
    total = series.sum()
    if total == 0: return 0
    shares = series / total
    return round(float((shares ** 2).sum()), 3)

def early_ratio(row):
    early = ["2025-01","2025-02","2025-03"]
    total = row[MONTHS].sum()
    if total == 0: return 0
    return round(row[early].sum() / total * 100, 1)

def sales_pattern(row):
    er = early_ratio(row); ssr = ss_ratio(row)
    if er >= 50: return "초반집중"
    if ssr >= 80: return "SS집중"
    if ssr >= 50: return "SS중심"
    return "연중분산"

def fmt_month(m):
    if m == "N/A": return m
    yy = m[2:4]; mm = m[5:7]
    labels = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
               "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}
    return f"{labels.get(mm,'?')}'{yy}"

monthly_sku["피크월"]      = monthly_sku.apply(peak_month, axis=1)
monthly_sku["SS비율pct"]   = monthly_sku.apply(ss_ratio, axis=1)
monthly_sku["초기비율pct"] = monthly_sku.apply(early_ratio, axis=1)
monthly_sku["집중도HHI"]   = monthly_sku.apply(sales_concentration, axis=1)
monthly_sku["판매패턴"]    = monthly_sku.apply(sales_pattern, axis=1)
monthly_sku["월별총계"]    = monthly_sku[MONTHS].sum(axis=1)
monthly_sku["피크월_레이블"] = monthly_sku["피크월"].apply(fmt_month)

print(f"   월별분석 완료: {len(monthly_sku)}개 품번")

# ════════════════════════════════════════════════════════════════════════════
# 4. SKU 집계 — v3 지표 (판매율, 실현이익율, 효율지수, 회전속도)
# ════════════════════════════════════════════════════════════════════════════
print("[4/10] SKU 집계 (v3 지표)...")

sku = df.groupby("품번").agg(
    품명=("품명","first"), 대분류=("대분류","first"), 활용=("활용","first"),
    로고사이즈=("로고사이즈","first"), 표현방식=("표현방식","first"),
    스타일=("스타일분류","first"), 볼캡유형=("볼캡유형","first"),
    브림=("브림","first"), 크라운=("크라운","first"), 패널수=("패널수","first"),
    소재=("소재","first"), 소재코드=("소재코드","first"),
    발주수량=("발주_수량","sum"), 판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"), 실판매금액=("누계판매_실판매금액","sum"),
    판매이익=("판매이익_금액","sum"), 최초판매가=("최초판매가","first"),
    현판매가=("현판매가","first"), 완사입비=("완사입비_단가","first"),
    색상수=("색상","nunique"),
    최초출고일=("최초출고일","min"), 최종출고일=("최종출고일","max"),
).reset_index()

# v3 핵심 지표
sku["판매율_pct"]     = (sku["판매수량"] / sku["발주수량"] * 100).round(1)
sku["실현이익율_pct"] = (sku["판매이익"] / sku["실판매금액"].replace(0, np.nan) * 100).round(1)
sku["효율지수"]       = (sku["판매율_pct"] * sku["실현이익율_pct"] / 100).round(2)
sku["판매경과일"]     = (sku["최종출고일"] - sku["최초출고일"]).dt.days
sku["회전속도"]       = (sku["판매수량"] / sku["판매경과일"].clip(lower=1) * 30).round(2)
sku["잔여재고"]       = (sku["발주수량"] - sku["판매수량"]).clip(lower=0)
sku["잔재고율_pct"]   = (sku["잔여재고"] / sku["발주수량"].replace(0, np.nan) * 100).round(1)

# v3 MUST/WATCH/SKIP (Kano: 판매율70% × 실현이익율40%)
def classify_v3(row):
    hr = row["판매율_pct"] >= 70
    hi = pd.notna(row["실현이익율_pct"]) and row["실현이익율_pct"] >= 40
    if hr and hi:   return "🟢 MUST"
    if hr and not hi: return "🟡 WATCH(VE)"
    if not hr and hi: return "🟡 WATCH(QR)"
    return "🔴 SKIP"

sku["26FW판별"] = sku.apply(classify_v3, axis=1)

# 월별 JOIN
sku = sku.merge(
    monthly_sku[["품번","피크월","피크월_레이블","SS비율pct","초기비율pct","집중도HHI","판매패턴","월별총계"] + MONTHS],
    on="품번", how="left")

sku_sorted = sku.sort_values("판매금액", ascending=False).reset_index(drop=True)

must_cnt  = (sku["26FW판별"] == "🟢 MUST").sum()
watch_cnt = (sku["26FW판별"].str.startswith("🟡")).sum()
skip_cnt  = (sku["26FW판별"] == "🔴 SKIP").sum()
total_sku = len(sku)

# KPI
total_orders  = df["발주_수량"].sum()
total_sales_q = df["누계판매_수량"].sum()
total_sales_a = df["누계판매_금액"].sum()
total_sales_r = df["누계판매_실판매금액"].sum()
total_profit  = df["판매이익_금액"].sum()
avg_sell_thru = total_sales_q / total_orders * 100 if total_orders else 0
avg_profit_r  = total_profit / total_sales_r * 100 if total_sales_r else 0
avg_효율지수   = sku["효율지수"].mean()

print(f"   KPI — 총판매금액: {total_sales_a/1e8:.2f}억 / 판매율: {avg_sell_thru:.1f}% / 실현이익율: {avg_profit_r:.1f}%")

# ── 카테고리 집계
def grp_agg(gdf, key):
    g = gdf.groupby(key, dropna=False).agg(
        SKU수=("품번","nunique"), 판매수량=("누계판매_수량","sum"),
        판매금액=("누계판매_금액","sum"), 판매이익합=("판매이익_금액","sum"),
        실판매금액합=("누계판매_실판매금액","sum"),
        발주수량합=("발주_수량","sum"), 누계판매합=("누계판매_수량","sum"),
    ).reset_index()
    g["판매율avg"]     = (g["누계판매합"] / g["발주수량합"].replace(0,np.nan) * 100).round(1)
    g["실현이익율avg"] = (g["판매이익합"] / g["실판매금액합"].replace(0,np.nan) * 100).round(1)
    g["효율지수avg"]   = (g["판매율avg"] * g["실현이익율avg"] / 100).round(2)
    return g.sort_values("판매금액", ascending=False)

cat_main    = grp_agg(df, "대분류")
cat_style2  = grp_agg(df, "볼캡유형")
cat_material= grp_agg(df, "소재코드")

# ── IP 8분류
ip_df = grp_agg(df, "활용")
ip_df["비중pct"] = (ip_df["판매금액"] / ip_df["판매금액"].sum() * 100).round(1)

df["IP구분"] = df["활용"].apply(lambda x: "키키IP" if str(x) in ["키키","릴리","IP"] else "비IP")
ip_bin = grp_agg(df, "IP구분")
ip_bin["비중pct"] = (ip_bin["판매금액"] / ip_bin["판매금액"].sum() * 100).round(1)

# ── 색상 집계
color_grp_df = df.groupby("색상계열").agg(
    건수=("품번","count"), 판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),
    발주합=("발주_수량","sum"), 판매합=("누계판매_수량","sum"),
).reset_index()
color_grp_df["판매율avg"] = (color_grp_df["판매합"] / color_grp_df["발주합"].replace(0,np.nan) * 100).round(1)
color_grp_df = color_grp_df.sort_values("판매금액", ascending=False)

# ── 출고월 집계
timing = df.groupby("출고월").agg(
    색상건수=("색상","count"), 판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"), 품번수=("품번","nunique"),
    발주합=("발주_수량","sum"), 판매합=("누계판매_수량","sum"),
).reset_index().sort_values("출고월")
timing["판매율avg"] = (timing["판매합"] / timing["발주합"].replace(0,np.nan) * 100).round(1)

peak_dist    = monthly_sku["피크월"].value_counts().sort_index()
pattern_dist = monthly_sku["판매패턴"].value_counts()

# ════════════════════════════════════════════════════════════════════════════
# 5. 속성 집계 (로고사이즈, 표현방식, 브림, 크라운, 패널수, 소재)
# ════════════════════════════════════════════════════════════════════════════
print("[5/10] 속성 심층 집계...")

attr_logo_sz  = grp_agg(df, "로고사이즈")
attr_expr     = grp_agg(df, "표현방식")
attr_brim     = grp_agg(df, "브림")
attr_crown    = grp_agg(df, "크라운")
attr_panel    = grp_agg(df, "패널수")
attr_material = grp_agg(df, "소재")

# 대분류 × 표현방식 크로스탭
cross_tab = df.groupby(["대분류","표현방식"], dropna=False).agg(
    SKU수=("품번","nunique"),
    발주합=("발주_수량","sum"),
    판매합=("누계판매_수량","sum"),
).reset_index()
cross_tab["판매율"] = (cross_tab["판매합"] / cross_tab["발주합"].replace(0,np.nan) * 100).round(1)

# 볼캡유형 × 브림 크로스탭
cross_cap_brim = df.groupby(["볼캡유형","브림"], dropna=False).agg(
    SKU수=("품번","nunique"),
    발주합=("발주_수량","sum"),
    판매합=("누계판매_수량","sum"),
).reset_index()
cross_cap_brim["판매율"] = (cross_cap_brim["판매합"] / cross_cap_brim["발주합"].replace(0,np.nan) * 100).round(1)

print("   속성 집계 완료")

# ════════════════════════════════════════════════════════════════════════════
# 6. 판매 예측 (잔여재고 + 소진 예상일)
# ════════════════════════════════════════════════════════════════════════════
print("[6/10] 판매 예측 계산...")

# 월별파일에서 잔여재고 집계 (품번별 재고 합산)
pred_cols = ["품번","재고","판매"] + [f"qty_{m}" for m in ["2026-01","2026-02","2026-03"]]
pred_cols_avail = [c for c in pred_cols if c in monthly.columns]
monthly_pred = monthly[pred_cols_avail].copy()

for col in ["재고","판매","qty_2026-01","qty_2026-02","qty_2026-03"]:
    if col in monthly_pred.columns:
        monthly_pred[col] = pd.to_numeric(monthly_pred[col], errors="coerce").fillna(0)

pred_grp = monthly_pred.groupby("품번").agg(
    잔여재고=("재고","sum"),
    누계판매=("판매","sum"),
    qty_2601=("qty_2026-01","sum") if "qty_2026-01" in monthly_pred.columns else ("재고","sum"),
    qty_2602=("qty_2026-02","sum") if "qty_2026-02" in monthly_pred.columns else ("재고","sum"),
    qty_2603=("qty_2026-03","sum") if "qty_2026-03" in monthly_pred.columns else ("재고","sum"),
).reset_index()

def safe_agg(group, col):
    if col in monthly_pred.columns:
        return monthly_pred.groupby("품번")[col].sum()
    return pd.Series(0, index=group.index)

# 재계산
g = monthly_pred.groupby("품번")
pred_grp = pd.DataFrame({"품번": monthly_pred["품번"].unique()})
pred_grp = pred_grp.set_index("품번")
pred_grp["잔여재고"] = g["재고"].sum() if "재고" in monthly_pred.columns else 0
for col in ["qty_2026-01","qty_2026-02","qty_2026-03"]:
    if col in monthly_pred.columns:
        pred_grp[col] = g[col].sum()
    else:
        pred_grp[col] = 0
pred_grp = pred_grp.reset_index()

# 월평균 판매속도
recent_cols = [c for c in ["qty_2026-01","qty_2026-02","qty_2026-03"] if c in pred_grp.columns]
if recent_cols:
    pred_grp["월평균속도"] = pred_grp[recent_cols].mean(axis=1).round(1)
else:
    pred_grp["월평균속도"] = 0

# 트렌드 기울기 보정
def calc_slope(row):
    vals = [row.get("qty_2026-01",0), row.get("qty_2026-02",0), row.get("qty_2026-03",0)]
    if sum(vals) == 0: return 0
    x = np.array([0,1,2]); y = np.array(vals, dtype=float)
    if len(x) > 1:
        slope = np.polyfit(x, y, 1)[0]
        return round(slope, 2)
    return 0

pred_grp["트렌드기울기"] = pred_grp.apply(calc_slope, axis=1)
# 보수적 속도 = 기존속도 + 기울기×0.7 (단, 음수면 0 처리)
pred_grp["보수속도"] = (pred_grp["월평균속도"] + pred_grp["트렌드기울기"] * 0.7).clip(lower=0.1).round(2)

# 예상 소진 계산
pred_grp["예상소진월수"] = (pred_grp["잔여재고"] / pred_grp["보수속도"]).round(1)
def expected_date(months):
    if pd.isna(months) or months <= 0: return "이미소진"
    d = ANALYSIS_DATE + timedelta(days=int(months * 30))
    return d.strftime("%Y-%m")
pred_grp["예상소진일"] = pred_grp["예상소진월수"].apply(expected_date)

# 위험 등급
def risk_grade(row):
    if row["잔여재고"] <= 0: return "✅ 재고없음"
    if row["월평균속도"] < 10: return "🔴 판매정체"
    if row["예상소진월수"] <= 3: return "🟢 빠른소진(3M↓)"
    if row["예상소진월수"] <= 6: return "🟡 정상소진(3~6M)"
    return "🔴 장기재고(6M+)"
pred_grp["위험등급"] = pred_grp.apply(risk_grade, axis=1)

# 완사입비 JOIN for 금액 리스크
완사입비_map = sku[["품번","완사입비"]].set_index("품번")["완사입비"].to_dict()
발주단가_map = sku[["품번","최초판매가"]].set_index("품번")["최초판매가"].to_dict()
pred_grp["완사입비"] = pred_grp["품번"].map(완사입비_map).fillna(0)
pred_grp["재고금액리스크"] = (pred_grp["잔여재고"] * pred_grp["완사입비"]).round(0)

# 품명 JOIN
품명_map = sku[["품번","품명"]].set_index("품번")["품명"].to_dict()
pred_grp["품명"] = pred_grp["품번"].map(품명_map).fillna("")

# 프로모션 우선순위: 장기재고 × 재고금액 상위
pred_grp["프로모션순위"] = 0
long_risk = pred_grp["위험등급"].isin(["🔴 장기재고(6M+)","🔴 판매정체"])
pred_grp.loc[long_risk, "프로모션순위"] = pred_grp.loc[long_risk, "재고금액리스크"].rank(ascending=False).astype(int)

pred_sorted = pred_grp.sort_values("재고금액리스크", ascending=False).reset_index(drop=True)

print(f"   예측 완료: {len(pred_sorted)}품번 / 장기재고: {(pred_grp['위험등급']=='🔴 장기재고(6M+)').sum()}개")

# ════════════════════════════════════════════════════════════════════════════
# 7. 26SS 초기 분석
# ════════════════════════════════════════════════════════════════════════════
print("[7/10] 26SS 분석...")

qty_cols_26 = [f"qty_{m}" for m in MONTHS]
for col in qty_cols_26:
    if col not in monthly_26.columns:
        monthly_26[col] = 0

monthly_26_sku = monthly_26.groupby("품번")[qty_cols_26 + (["재고","판매"] if "판매" in monthly_26.columns else [])].sum().reset_index() if len(monthly_26) else pd.DataFrame(columns=["품번"])

if len(monthly_26_sku) > 0:
    for col in qty_cols_26:
        if col not in monthly_26_sku.columns:
            monthly_26_sku[col] = 0

    # 26SS 기준: 2026-01, 2026-02, 2026-03 판매
    m26_cols = ["qty_2026-01","qty_2026-02","qty_2026-03"]
    for c in m26_cols:
        if c not in monthly_26_sku.columns: monthly_26_sku[c] = 0
    monthly_26_sku["초기3M판매"] = monthly_26_sku[m26_cols].sum(axis=1)

    # 발주수량 JOIN from monthly_26
    if "판매" in monthly_26.columns:
        발주26_grp = monthly_26.groupby("품번").agg(
            발주합=("기초","sum") if "기초" in monthly_26.columns else ("재고","sum"),
            판매합=("판매","sum"),
            잔재고합=("재고","sum") if "재고" in monthly_26.columns else ("판매","sum"),
        ).reset_index()
        monthly_26_sku = monthly_26_sku.merge(발주26_grp, on="품번", how="left")
    else:
        monthly_26_sku["발주합"] = 0
        monthly_26_sku["판매합"] = 0

    # 25SS 동기 대비 (론칭후 3개월: WA25는 2025-01~03)
    m25_early = ["qty_2025-01","qty_2025-02","qty_2025-03"]
    wa25_early = monthly_sku.copy()
    for c in m25_early:
        if c not in wa25_early.columns: wa25_early[c] = 0
    wa25_early["초기3M판매_25SS"] = wa25_early[m25_early].sum(axis=1)
    wa25_avg_3m = wa25_early["초기3M판매_25SS"].mean()

    wa26_avg_3m = monthly_26_sku["초기3M판매"].mean() if len(monthly_26_sku) else 0
    wa26_sku_cnt = len(monthly_26_sku)
else:
    wa26_avg_3m = 0
    wa26_sku_cnt = 0
    wa25_avg_3m = 0

print(f"   26SS: {wa26_sku_cnt}품번 / 초기3M평균: {wa26_avg_3m:.0f}개")

# ════════════════════════════════════════════════════════════════════════════
# 8. Excel 14시트 작성
# ════════════════════════════════════════════════════════════════════════════
print("[8/10] Excel 14시트 작성...")

wb = Workbook()
wb.remove(wb.active)

# ── S1: 종합요약
ws1 = wb.create_sheet("📊 종합요약")
sheet_title(ws1, "🧢 와키윌리 25SS 모자 판매 종합 분석 v3 | 2026-03-19", 8)
ws1.merge_cells("A2:H2")
c=ws1["A2"]; c.value="판매자료×구조화 59SKU×월별추이 15개월 | 지표: 판매율·실현이익율·효율지수·회전속도 | 기준: Kano 판매율70%×실현이익율40%"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,9); c.alignment=ctr()

kpis = [("총 발주수량",f"{int(total_orders):,} 개"),("총 판매금액",f"{int(total_sales_a):,} 원"),
        ("평균 판매율",f"{avg_sell_thru:.1f} %"),("평균 실현이익율",f"{avg_profit_r:.1f} %")]
hdr_row(ws1,4,[k for k,v in kpis], bg=C_ACCENT)
data_row(ws1,5,[v for k,v in kpis])
ws1.row_dimensions[5].height = 26

kpis2 = [("실판매금액",f"{int(total_sales_r):,} 원"),("판매이익 합계",f"{int(total_profit):,} 원"),
         ("평균 효율지수",f"{avg_효율지수:.1f}"),("분석 SKU",f"{total_sku} 개")]
hdr_row(ws1,7,[k for k,v in kpis2],bg=C_SUB)
data_row(ws1,8,[v for k,v in kpis2])

hdr_row(ws1,10,["26FW판별","SKU수","비중","기준"],bg=C_ACCENT)
for ri,(lb,cnt,기준,bg_) in enumerate([
    ("🟢 MUST", must_cnt,  "판매율≥70%×실현이익율≥40%", C_MUST),
    ("🟡 WATCH", watch_cnt, "조건부 충족", C_WATCH),
    ("🔴 SKIP",  skip_cnt,  "판매율<70%×실현이익율<40%", C_SKIP)
], 11):
    data_row(ws1,ri,[lb, cnt, f"{cnt/total_sku*100:.0f}%", 기준], bg=bg_)

hdr_row(ws1,15,["판매패턴","SKU수"],bg=C_SUB)
for ri,(pat,cnt) in enumerate(pattern_dist.items(),16):
    data_row(ws1,ri,[pat,int(cnt)])

hdr_row(ws1,21,["피크 출고월 TOP5","SKU수"],bg=C_ACCENT)
for ri,(m,cnt) in enumerate(peak_dist.head(5).items(),22):
    data_row(ws1,ri,[m,int(cnt)])

set_widths(ws1,[30,18,14,30,14,14,14,14])

# ── S2: SKU 랭킹
ws2 = wb.create_sheet("🏆 SKU랭킹")
sheet_title(ws2,"🏆 품번별 판매 랭킹 — 판매금액·판매율·실현이익율·효율지수·피크월",15)
h2=["순위","품번","품명","대분류","활용","볼캡","소재",
    "발주수량","판매수량","판매율%","판매금액(원)","실현이익율%","효율지수","피크월","26FW판별"]
hdr_row(ws2,2,h2)
for ri,(_, r) in enumerate(sku_sorted.iterrows(),3):
    data_row(ws2,ri,[ri-2, r["품번"], r["품명"], r["대분류"], r["활용"], r["볼캡유형"], r["소재"],
                     int(r["발주수량"]), int(r["판매수량"]),
                     f"{r['판매율_pct']:.1f}%", int(r["판매금액"]),
                     f"{r['실현이익율_pct']:.1f}%" if pd.notna(r["실현이익율_pct"]) else "-",
                     f"{r['효율지수']:.1f}" if pd.notna(r["효율지수"]) else "-",
                     r.get("피크월_레이블","N/A"), r["26FW판별"]], bg=row_bg(r["26FW판별"]))
set_widths(ws2,[5,14,26,10,8,12,8,10,10,9,16,12,10,10,14])
ws2.freeze_panes="A3"

# ── S3: 판매타이밍
ws3 = wb.create_sheet("📅 판매타이밍")
sheet_title(ws3,"📅 최초출고월별 판매 타이밍 분석",7)
hdr_row(ws3,2,["최초출고월","색상건수","품번수","판매수량","판매금액(원)","판매율avg%","구간"])
구간MAP = {
    "2024-11":"초기선입고","2024-12":"초기선입고",
    "2025-01":"SS초반","2025-02":"SS초반","2025-03":"SS초반",
    "2025-04":"SS피크","2025-05":"SS피크",
    "2025-06":"재고소진","2025-07":"재고소진",
}
for ri,r in enumerate(timing.itertuples(),3):
    구간 = 구간MAP.get(str(r.출고월),"기타")
    bg_ = {"초기선입고":C_LIGHT,"SS초반":"E8F5E9","SS피크":"D4EDDA","재고소진":C_WATCH}.get(구간,None)
    data_row(ws3,ri,[r.출고월, r.색상건수, r.품번수, int(r.판매수량), int(r.판매금액),
                     f"{r.판매율avg:.1f}%", 구간], bg=bg_)
set_widths(ws3,[14,10,10,12,18,12,12])

# ── S4: 색상분석
ws4 = wb.create_sheet("🎨 색상분석")
sheet_title(ws4,"🎨 색상계열별 판매율 분석",5)
hdr_row(ws4,2,["색상계열","건수","판매수량","판매금액(원)","판매율avg%"],bg=C_ACCENT)
for ri,r in enumerate(color_grp_df.itertuples(),3):
    data_row(ws4,ri,[r.색상계열, r.건수, int(r.판매수량), int(r.판매금액), f"{r.판매율avg:.1f}%"])
set_widths(ws4,[20,10,14,18,14])

# ── S5: 카테고리분석
ws5 = wb.create_sheet("🗂️ 카테고리분석")
sheet_title(ws5,"🗂️ 대분류 × 볼캡유형 × 소재별 분석",7)
hdr_row(ws5,2,["대분류","SKU수","판매수량","판매금액(원)","판매율%","실현이익율%","효율지수avg"],bg=C_ACCENT)
for ri,r in enumerate(cat_main.itertuples(),3):
    data_row(ws5,ri,[str(r.대분류) if pd.notna(r.대분류) else "-", int(r.SKU수), int(r.판매수량),
                     int(r.판매금액), f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%", f"{r.효율지수avg:.1f}"])
r0=len(cat_main)+5
hdr_row(ws5,r0,["볼캡유형","SKU수","판매수량","판매금액(원)","판매율%","실현이익율%","효율지수avg"],bg=C_SUB)
for ri,r in enumerate(cat_style2.itertuples(),r0+1):
    data_row(ws5,ri,[str(r.볼캡유형) if pd.notna(r.볼캡유형) else "-", int(r.SKU수), int(r.판매수량),
                     int(r.판매금액), f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%", f"{r.효율지수avg:.1f}"])
r1=r0+len(cat_style2)+3
hdr_row(ws5,r1,["소재코드","SKU수","판매수량","판매금액(원)","판매율%","실현이익율%","효율지수avg"],bg=C_ACCENT)
for ri,r in enumerate(cat_material.itertuples(),r1+1):
    data_row(ws5,ri,[str(r.소재코드) if pd.notna(r.소재코드) else "-", int(r.SKU수), int(r.판매수량),
                     int(r.판매금액), f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%", f"{r.효율지수avg:.1f}"])
set_widths(ws5,[18,8,12,18,10,12,12])

# ── S6: IP분석 (8분류)
ws6 = wb.create_sheet("🐱 IP분석")
sheet_title(ws6,"🐱 IP 활용 세부 8분류 판매 성과 (타이포/시즌/두줄/릴리/키키/조합/한줄/IP)",8)
hdr_row(ws6,2,["활용유형","SKU수","판매수량","판매금액(원)","비중%","판매율%","실현이익율%","효율지수avg"],bg=C_ACCENT)
for ri,r in enumerate(ip_df.itertuples(),3):
    data_row(ws6,ri,[str(r.활용) if pd.notna(r.활용) else "미분류", int(r.SKU수),
                     int(r.판매수량), int(r.판매금액), f"{r.비중pct:.1f}%",
                     f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%", f"{r.효율지수avg:.1f}"])
r2=len(ip_df)+5
hdr_row(ws6,r2,["IP구분","SKU수","판매수량","판매금액(원)","비중%","판매율%"],bg=C_SUB)
for ri,r in enumerate(ip_bin.itertuples(),r2+1):
    data_row(ws6,ri,[r.IP구분, int(r.SKU수), int(r.판매수량), int(r.판매금액),
                     f"{r.비중pct:.1f}%", f"{r.판매율avg:.1f}%"])
set_widths(ws6,[14,8,12,18,10,10,12,12])

# ── S7: 수익구조
ws7 = wb.create_sheet("💰 수익구조")
sheet_title(ws7,"💰 수익 구조 — 실현이익율·효율지수·회전속도 분포",10)
hdr_row(ws7,2,["품번","품명","최초판매가","현판매가","할인","판매율%","실현이익율%","효율지수","회전속도","26FW판별"])
discount_map = df.groupby("품번")["할인여부"].any().to_dict()
for ri,(_, r) in enumerate(sku_sorted.iterrows(),3):
    이익율 = f"{r['실현이익율_pct']:.1f}%" if pd.notna(r["실현이익율_pct"]) else "-"
    효율 = f"{r['효율지수']:.1f}" if pd.notna(r["효율지수"]) else "-"
    회전 = f"{r['회전속도']:.1f}" if pd.notna(r["회전속도"]) else "-"
    bg_ = C_MUST if (pd.notna(r["실현이익율_pct"]) and r["실현이익율_pct"] >= 40) else (
          C_SKIP if (pd.notna(r["실현이익율_pct"]) and r["실현이익율_pct"] < 20) else None)
    data_row(ws7,ri,[r["품번"], r["품명"], f"{int(r['최초판매가']):,}", f"{int(r['현판매가']):,}",
                     "Y" if discount_map.get(r["품번"],False) else "-",
                     f"{r['판매율_pct']:.1f}%", 이익율, 효율, 회전, r["26FW판별"]], bg=bg_)
set_widths(ws7,[14,28,12,12,6,10,12,10,10,14])
ws7.freeze_panes="A3"

# ── S8: 26FW시사점
ws8 = wb.create_sheet("🎯 26FW시사점")
sheet_title(ws8,"🎯 26FW 기획 시사점 — MUST/WATCH/SKIP (Kano 기준)",10)
ws8.merge_cells("A2:J2")
c=ws8["A2"]; c.value="기준: 판매율 ≥70% × 실현이익율 ≥40% (Kano 모델) | 피크월·판매패턴·효율지수 포함"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,9); c.alignment=ctr()
hdr_row(ws8,3,["26FW","품번","품명","대분류","활용","판매율%","실현이익율%","효율지수","피크월","판매패턴"])
for ri,(_, r) in enumerate(sku_sorted.iterrows(),4):
    판별 = r["26FW판별"]
    data_row(ws8,ri,[판별, r["품번"], r["품명"], r["대분류"], r["활용"],
                     f"{r['판매율_pct']:.1f}%",
                     f"{r['실현이익율_pct']:.1f}%" if pd.notna(r["실현이익율_pct"]) else "-",
                     f"{r['효율지수']:.1f}" if pd.notna(r["효율지수"]) else "-",
                     r.get("피크월_레이블","N/A"), r.get("판매패턴","N/A")], bg=row_bg(판별))
set_widths(ws8,[16,14,24,10,8,10,12,10,10,12])
ws8.freeze_panes="A4"

# ── S9: 월별판매추이 히트맵
ws9 = wb.create_sheet("📈 월별판매추이")
ws9.sheet_view.showGridLines = False
ws9.merge_cells(f"A1:{get_column_letter(len(MONTHS)+4)}1")
c=ws9["A1"]; c.value="📈 품번별 월별 판매수량 추이 히트맵 (2025-01 ~ 2026-03)"
c.fill=fill(C_HEADER); c.font=fnt(True,C_WHITE,13); c.alignment=ctr(); ws9.row_dimensions[1].height=30
hdr_vals = ["품번","품명","판별","판매패턴"] + MONTHS + ["합계"]
hdr_row(ws9,2,hdr_vals)
MONTH_LABELS = [fmt_month(m) for m in MONTHS]
for ci,lb in enumerate(["","","",""] + MONTH_LABELS + [""],1):
    c=ws9.cell(row=3,column=ci,value=lb)
    c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,8); c.alignment=ctr(); c.border=brd()
ws9.row_dimensions[3].height=14

display_skus = sku_sorted[["품번","품명","26FW판별"]].copy()
mx_val = monthly_sku[MONTHS].values.max() if len(monthly_sku) else 1
for ri,(_, sr) in enumerate(display_skus.iterrows(), 4):
    품번 = sr["품번"]; 품명 = str(sr["품명"])[:18]; 판별 = str(sr["26FW판별"])
    mrow = monthly_sku[monthly_sku["품번"]==품번]
    패턴 = mrow["판매패턴"].values[0] if len(mrow) else "N/A"
    ws9.cell(row=ri,column=1,value=품번).border=brd()
    ws9.cell(row=ri,column=2,value=품명).border=brd()
    ws9.cell(row=ri,column=3,value=판별).border=brd()
    ws9.cell(row=ri,column=4,value=패턴).border=brd()
    for ci,m in enumerate(MONTHS,5):
        val = int(mrow[m].values[0]) if len(mrow) and m in mrow.columns else 0
        hc = heat_color(val, mx_val)
        c = ws9.cell(row=ri,column=ci,value=val if val>0 else "")
        c.fill=fill(hc); c.font=fnt(False,C_DARK if val<mx_val*0.6 else C_WHITE,9)
        c.alignment=ctr(); c.border=brd()
    total = int(mrow["월별총계"].values[0]) if len(mrow) else 0
    ws9.cell(row=ri,column=len(MONTHS)+5,value=total).border=brd()
ws9.freeze_panes="E4"
set_widths(ws9, [16,20,14,10] + [8]*len(MONTHS) + [8])

# ── S10: 주판매시기
ws10 = wb.create_sheet("⏱️ 주판매시기")
sheet_title(ws10,"⏱️ 품번별 주 판매시기 & 판매집중도 분석",9)
hdr_row(ws10,2,["품번","품명","대분류","활용","피크월","SS비율%","초기3M비율%","집중도HHI","판매패턴"])
ws10.merge_cells("A3:I3")
c=ws10["A3"]; c.value="피크월=최다판매월 | SS비율=2025-01~06 판매비중 | HHI: 1.0=1개월집중"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,8); c.alignment=ctr()
sku_timing = sku_sorted.merge(monthly_sku[["품번","피크월","피크월_레이블","SS비율pct","초기비율pct","집중도HHI","판매패턴"]],
                               on="품번",how="left",suffixes=("","_m"))
for ri,(_, r) in enumerate(sku_timing.iterrows(),4):
    패턴 = str(r.get("판매패턴_m", r.get("판매패턴","N/A")))
    bg_ = {"초반집중":"E3F2FD","SS집중":C_MUST,"SS중심":"E8F5E9","연중분산":C_WATCH}.get(패턴, None)
    data_row(ws10,ri,[r["품번"],r["품명"],r["대분류"],r["활용"],
                      r.get("피크월_레이블_m", r.get("피크월_레이블","N/A")),
                      f"{r.get('SS비율pct_m', r.get('SS비율pct',0)):.1f}%",
                      f"{r.get('초기비율pct_m', r.get('초기비율pct',0)):.1f}%",
                      f"{r.get('집중도HHI_m', r.get('집중도HHI',0)):.3f}", 패턴], bg=bg_)
set_widths(ws10,[14,26,10,8,10,10,12,12,12])
ws10.freeze_panes="A4"

# ── S11: 로고속성
ws11 = wb.create_sheet("🔍 로고속성")
sheet_title(ws11,"🔍 로고 속성 심층 분석 — 로고사이즈 × 표현방식",8)
hdr_row(ws11,2,["로고사이즈","SKU수","판매수량","판매금액(원)","판매율%","실현이익율%","효율지수avg","비중%"],bg=C_ACCENT)
total_pf = attr_logo_sz["판매금액"].sum()
for ri,r in enumerate(attr_logo_sz.itertuples(),3):
    비중 = r.판매금액/total_pf*100 if total_pf else 0
    data_row(ws11,ri,[str(r.로고사이즈) if pd.notna(r.로고사이즈) else "미분류",
                      int(r.SKU수), int(r.판매수량), int(r.판매금액),
                      f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%", f"{r.효율지수avg:.1f}", f"{비중:.1f}%"])
r_e = len(attr_logo_sz)+5
hdr_row(ws11,r_e,["표현방식","SKU수","판매수량","판매금액(원)","판매율%","실현이익율%","효율지수avg","비중%"],bg=C_SUB)
for ri,r in enumerate(attr_expr.itertuples(),r_e+1):
    비중 = r.판매금액/total_pf*100 if total_pf else 0
    data_row(ws11,ri,[str(r.표현방식) if pd.notna(r.표현방식) else "미분류",
                      int(r.SKU수), int(r.판매수량), int(r.판매금액),
                      f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%", f"{r.효율지수avg:.1f}", f"{비중:.1f}%"])
r_c = r_e+len(attr_expr)+3
hdr_row(ws11,r_c,["대분류","표현방식","SKU수","판매수량","판매율%","(크로스탭)","",""],bg=C_ACCENT)
for ri,r in enumerate(cross_tab.itertuples(),r_c+1):
    data_row(ws11,ri,[str(r.대분류) if pd.notna(r.대분류) else "-",
                      str(r.표현방식) if pd.notna(r.표현방식) else "-",
                      int(r.SKU수), int(r.판매합), f"{r.판매율:.1f}%","","",""])
set_widths(ws11,[14,8,12,18,10,12,12,10])

# ── S12: 구조속성
ws12 = wb.create_sheet("🏗️ 구조속성")
sheet_title(ws12,"🏗️ 구조 속성 심층 분석 — 브림·크라운·패널수·소재",8)
sections = [
    (attr_brim,"브림","브림 유형"),
    (attr_crown,"크라운","크라운 유형"),
    (attr_panel,"패널수","패널수"),
    (attr_material,"소재","소재"),
]
cur_row = 2
for adf, key_col, label in sections:
    hdr_row(ws12,cur_row,[label,"SKU수","판매수량","판매금액(원)","판매율%","실현이익율%","효율지수avg","비중%"],bg=C_ACCENT)
    cur_row += 1
    t_pf = adf["판매금액"].sum()
    for r in adf.itertuples():
        비중 = r.판매금액/t_pf*100 if t_pf else 0
        v = getattr(r, key_col)
        data_row(ws12,cur_row,[str(v) if pd.notna(v) else "미분류", int(r.SKU수),
                               int(r.판매수량), int(r.판매금액),
                               f"{r.판매율avg:.1f}%", f"{r.실현이익율avg:.1f}%",
                               f"{r.효율지수avg:.1f}", f"{비중:.1f}%"])
        cur_row += 1
    cur_row += 2
# 볼캡유형×브림 크로스탭
hdr_row(ws12,cur_row,["볼캡유형","브림","SKU수","판매수량","판매율%","","",""],bg=C_SUB)
cur_row += 1
for r in cross_cap_brim.itertuples():
    data_row(ws12,cur_row,[str(r.볼캡유형) if pd.notna(r.볼캡유형) else "-",
                           str(r.브림) if pd.notna(r.브림) else "-",
                           int(r.SKU수), int(r.판매합), f"{r.판매율:.1f}%","","",""])
    cur_row += 1
set_widths(ws12,[16,8,12,18,10,12,12,10])

# ── S13: 판매예측
ws13 = wb.create_sheet("📦 판매예측")
sheet_title(ws13,"📦 25SS 잔여재고 판매 예측 — 소진 예상일 + 위험등급",10)
ws13.merge_cells("A2:J2")
c=ws13["A2"]; c.value=f"기준일: 2026-03-19 | 속도보정: 최근3M 트렌드기울기×0.7 | 빠른소진<3M / 정상소진3~6M / 장기재고>6M / 판매정체<10개/월"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,9); c.alignment=ctr()
hdr_row(ws13,3,["품번","품명","잔여재고","月평균속도","트렌드기울기","보수속도","예상소진月","예상소진일","위험등급","재고금액리스크(원)"],bg=C_ACCENT)
for ri,r in enumerate(pred_sorted.itertuples(),4):
    bg_ = {"🟢 빠른소진(3M↓)":C_MUST,"🟡 정상소진(3~6M)":C_WATCH,
           "🔴 장기재고(6M+)":C_SKIP,"🔴 판매정체":C_SKIP}.get(r.위험등급, None)
    data_row(ws13,ri,[r.품번, str(r.품명)[:24],
                      int(r.잔여재고), f"{r.월평균속도:.1f}",
                      f"{r.트렌드기울기:+.1f}", f"{r.보수속도:.1f}",
                      f"{r.예상소진월수:.1f}" if r.잔여재고>0 else "-",
                      r.예상소진일, r.위험등급, f"{int(r.재고금액리스크):,}"], bg=bg_)
set_widths(ws13,[14,24,10,10,12,10,10,12,18,18])
ws13.freeze_panes="A4"

# ── S14: 26SS분석
ws14 = wb.create_sheet("🆕 26SS분석")
sheet_title(ws14,"🆕 26SS(WA26) 초기 판매 패턴 — 25SS 동기 대비",10)
ws14.merge_cells("A2:J2")
c=ws14["A2"]; c.value=f"26SS: WA26 {wa26_sku_cnt}품번 | 초기3M평균(26SS): {wa26_avg_3m:.0f}개 vs (25SS): {wa25_avg_3m:.0f}개"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,9); c.alignment=ctr()

if len(monthly_26_sku) > 0:
    recent_m26 = ["qty_2026-01","qty_2026-02","qty_2026-03"]
    avail_m26 = [c for c in recent_m26 if c in monthly_26_sku.columns]
    hdr_vals14 = ["품번","초기3M판매"] + avail_m26 + ["잔재고","판매합"]
    hdr_row(ws14,3,hdr_vals14[:8])
    for ri,(_, r) in enumerate(monthly_26_sku.sort_values("초기3M판매",ascending=False).iterrows(),4):
        vals = [r["품번"], int(r["초기3M판매"])]
        for c in avail_m26:
            vals.append(int(r.get(c,0)))
        vals.append(int(r.get("잔재고합",0)) if "잔재고합" in r else "-")
        vals.append(int(r.get("판매합",0)) if "판매합" in r else "-")
        data_row(ws14,ri,vals[:8])

    # 25SS 동기 대비 요약
    r_s = len(monthly_26_sku)+6
    hdr_row(ws14,r_s,["구분","초기3M합계","초기3M평균","품번수"],bg=C_SUB)
    m25_early = [c for c in ["qty_2025-01","qty_2025-02","qty_2025-03"] if c in monthly_sku.columns]
    wa25_3m_sum = monthly_sku[m25_early].sum().sum() if m25_early else 0
    wa26_3m_sum = monthly_26_sku["초기3M판매"].sum()
    data_row(ws14,r_s+1,["25SS 론칭 초기 3M", int(wa25_3m_sum), f"{wa25_avg_3m:.0f}", len(monthly_sku)])
    data_row(ws14,r_s+2,["26SS 론칭 초기 3M", int(wa26_3m_sum), f"{wa26_avg_3m:.0f}", wa26_sku_cnt])
    yoy = (wa26_avg_3m/wa25_avg_3m*100-100) if wa25_avg_3m > 0 else 0
    data_row(ws14,r_s+3,[f"YoY 초기 판매 증감", "-", f"{yoy:+.1f}%", "-"],
             bg=C_MUST if yoy >= 0 else C_SKIP)
else:
    hdr_row(ws14,3,["안내"],bg=C_SUB)
    data_row(ws14,4,["26SS(WA26) 모자 데이터가 없거나 필터 조건에 해당 없음"])

set_widths(ws14,[14,12,12,12,12,12,12,12,12,12])
ws14.freeze_panes="A4"

wb.save(OUT_XLSX)
print(f"   ✓ Excel 저장: {os.path.basename(OUT_XLSX)}")

# ════════════════════════════════════════════════════════════════════════════
# 9. HTML 리포트 (v3 — 11섹션)
# ════════════════════════════════════════════════════════════════════════════
print("[9/10] HTML 리포트 생성...")

top15 = sku_sorted.head(15)
chart_sku_labels = json.dumps([str(r) for r in top15["품번"]], ensure_ascii=False)
chart_sku_sales  = json.dumps([int(v) for v in top15["판매금액"]])
chart_sku_sell   = json.dumps([float(v) for v in top15["판매율_pct"]])
chart_sku_colors = json.dumps([
    ("#4CAF50" if "MUST" in str(r) else ("#FF9800" if "WATCH" in str(r) else "#F44336"))
    for r in top15["26FW판별"]
])

top5_codes = list(sku_sorted.head(5)["품번"])
monthly_chart = {}
for code in top5_codes:
    mrow = monthly_sku[monthly_sku["품번"]==code]
    monthly_chart[code] = [int(mrow[m].values[0]) for m in MONTHS] if len(mrow) else [0]*len(MONTHS)

chart_monthly_labels   = json.dumps(MONTHS)
palette = ["#E94560","#0F3460","#533483","#E94560","#16213E"]
chart_monthly_datasets = []
for i,(code,vals) in enumerate(monthly_chart.items()):
    short_name = str(sku_sorted[sku_sorted["품번"]==code]["품명"].values[0])[:16] if len(sku_sorted[sku_sorted["품번"]==code]) else code
    chart_monthly_datasets.append({"label":short_name,"data":vals,"borderColor":palette[i%len(palette)],
        "backgroundColor":palette[i%len(palette)]+"33","tension":0.4,"fill":False})
chart_monthly_data = json.dumps(chart_monthly_datasets, ensure_ascii=False)

ip_clean = ip_df[ip_df["활용"].notna()].copy()
chart_ip_labels = json.dumps([str(v) for v in ip_clean["활용"]], ensure_ascii=False)
chart_ip_data   = json.dumps([int(v) for v in ip_clean["판매금액"]])

cat_clean = cat_main[cat_main["대분류"].notna()].copy()
chart_cat_labels = json.dumps([str(v) for v in cat_clean["대분류"]], ensure_ascii=False)
chart_cat_data   = json.dumps([int(v) for v in cat_clean["판매금액"]])

chart_color_labels = json.dumps([str(v) for v in color_grp_df["색상계열"]], ensure_ascii=False)
chart_color_sell   = json.dumps([float(v) for v in color_grp_df["판매율avg"]])
chart_color_sales  = json.dumps([int(v) for v in color_grp_df["판매금액"]])

# 버블 (판매율 × 실현이익율 × 판매금액)
bubble_data = []
for _, r in sku_sorted.iterrows():
    판별 = r["26FW판별"]
    color = "#4CAF50" if "MUST" in str(판별) else ("#FF9800" if "WATCH" in str(판별) else "#F44336")
    x = round(float(r["판매율_pct"]),1) if pd.notna(r["판매율_pct"]) else 0
    y = round(float(r["실현이익율_pct"]),1) if pd.notna(r["실현이익율_pct"]) else 0
    bubble_data.append({"x":x,"y":y,"r":max(4,min(30,int(r["판매금액"])//5000000)),
                         "label":str(r["품번"]),"bg":color,"판별":str(판별)})
chart_bubble = json.dumps(bubble_data, ensure_ascii=False)

peak_months_sorted = sorted(peak_dist.index.tolist())
chart_peak_labels = json.dumps(peak_months_sorted)
chart_peak_data   = json.dumps([int(peak_dist.get(m,0)) for m in peak_months_sorted])

chart_판별_labels = json.dumps(["MUST","WATCH","SKIP"], ensure_ascii=False)
chart_판별_data   = json.dumps([int(must_cnt),int(watch_cnt),int(skip_cnt)])

# 로고속성 차트 (표현방식별 판매율+실현이익율)
expr_sorted = attr_expr.sort_values("판매율avg", ascending=False)
chart_expr_labels  = json.dumps([str(v) for v in expr_sorted["표현방식"].fillna("미분류")], ensure_ascii=False)
chart_expr_sell    = json.dumps([float(v) for v in expr_sorted["판매율avg"]])
chart_expr_profit  = json.dumps([float(v) for v in expr_sorted["실현이익율avg"]])

# 구조속성 차트 (브림별 효율지수)
brim_sorted = attr_brim.sort_values("효율지수avg", ascending=False)
chart_brim_labels = json.dumps([str(v) for v in brim_sorted["브림"].fillna("미분류")], ensure_ascii=False)
chart_brim_eff    = json.dumps([float(v) for v in brim_sorted["효율지수avg"]])

# 소재별 효율지수
mat_sorted = attr_material.sort_values("효율지수avg", ascending=False)
chart_mat_labels = json.dumps([str(v) for v in mat_sorted["소재"].fillna("미분류")], ensure_ascii=False)
chart_mat_eff    = json.dumps([float(v) for v in mat_sorted["효율지수avg"]])

# 판매예측 차트 (위험등급별 재고금액)
risk_grp = pred_sorted.groupby("위험등급")["재고금액리스크"].sum().reset_index()
chart_risk_labels = json.dumps([str(v) for v in risk_grp["위험등급"]], ensure_ascii=False)
chart_risk_data   = json.dumps([int(v) for v in risk_grp["재고금액리스크"]])
chart_risk_colors = json.dumps([
    "#4CAF50" if "빠른" in str(v) else ("#FF9800" if "정상" in str(v) else "#F44336")
    for v in risk_grp["위험등급"]
])

# 26SS 차트 (WA26 월별 판매)
if len(monthly_26_sku) > 0:
    top5_26 = monthly_26_sku.sort_values("초기3M판매", ascending=False).head(5)
    m26_recent = ["qty_2026-01","qty_2026-02","qty_2026-03"]
    chart_26ss_labels = json.dumps(["2026-01","2026-02","2026-03"])
    chart_26ss_datasets = []
    for i,(_, r) in enumerate(top5_26.iterrows()):
        vals = [int(r.get(c,0)) for c in m26_recent]
        chart_26ss_datasets.append({"label":str(r["품번"]),"data":vals,
            "borderColor":palette[i%len(palette)],"backgroundColor":palette[i%len(palette)]+"33","tension":0.4,"fill":False})
    chart_26ss_data = json.dumps(chart_26ss_datasets, ensure_ascii=False)
else:
    chart_26ss_labels = json.dumps(["2026-01","2026-02","2026-03"])
    chart_26ss_data   = json.dumps([])

# 히트맵 HTML
heatmap_rows_html = []
mx_heat = monthly_sku[MONTHS].values.max() if len(monthly_sku) else 1
for _, sr in display_skus.iterrows():
    품번 = sr["품번"]; 품명 = str(sr["품명"])[:20]; 판별_str = str(sr["26FW판별"])
    mrow = monthly_sku[monthly_sku["품번"]==품번]
    패턴 = mrow["판매패턴"].values[0] if len(mrow) else "N/A"
    판별_cls = "must" if "MUST" in 판별_str else ("skip" if "SKIP" in 판별_str else "watch")
    cells = f'<td class="sku-code {판별_cls}">{품번}</td><td class="sku-name">{품명}</td><td class="pattern">{패턴}</td>'
    total = 0
    for m in MONTHS:
        val = int(mrow[m].values[0]) if len(mrow) and m in mrow.columns else 0
        total += val
        if val == 0:
            cells += '<td class="heat-cell heat-zero">-</td>'
        else:
            ratio = val/mx_heat
            r_c=int(250+(233-250)*ratio); g_c=int(251+(69-251)*ratio); b_c=int(255+(96-255)*ratio)
            bg_hex=f"#{r_c:02x}{g_c:02x}{b_c:02x}"; txt_color="#fff" if ratio>0.6 else "#1a1a2e"
            cells += f'<td class="heat-cell" style="background:{bg_hex};color:{txt_color}">{val:,}</td>'
    cells += f'<td class="heat-total">{total:,}</td>'
    heatmap_rows_html.append(f"<tr>{cells}</tr>")
heatmap_html = "\n".join(heatmap_rows_html)
heatmap_month_headers = "".join(f'<th>{fmt_month(m)}</th>' for m in MONTHS)

# 상위 SKU 테이블
top_table_rows = []
for _, r in sku_sorted.head(20).iterrows():
    판별 = r["26FW판별"]; cls = "must" if "MUST" in str(판별) else ("skip" if "SKIP" in str(판별) else "watch")
    top_table_rows.append(
        f'<tr><td class="sku-code {cls}">{r["품번"]}</td>'
        f'<td class="sku-name">{str(r["품명"])[:22]}</td>'
        f'<td>{str(r["대분류"]) if pd.notna(r["대분류"]) else "-"}</td>'
        f'<td>{str(r["활용"]) if pd.notna(r["활용"]) else "-"}</td>'
        f'<td>{str(r["볼캡유형"]) if pd.notna(r["볼캡유형"]) else "-"}</td>'
        f'<td class="num">{int(r["판매금액"]):,}</td>'
        f'<td class="num">{r["판매율_pct"]:.1f}%</td>'
        f'<td class="num">{r["실현이익율_pct"]:.1f}%</td>'
        f'<td class="num">{r["효율지수"]:.1f}</td>'
        f'<td class="badge {cls}">{판별}</td></tr>'
    )
top_table_html = "\n".join(top_table_rows)

# 예측 테이블 HTML (상위 10)
pred_table_rows = []
for _, r in pred_sorted[pred_sorted["잔여재고"]>0].head(15).iterrows():
    cls = "skip" if "장기" in str(r["위험등급"]) or "정체" in str(r["위험등급"]) else (
          "must" if "빠른" in str(r["위험등급"]) else "watch")
    pred_table_rows.append(
        f'<tr><td class="sku-code">{r["품번"]}</td>'
        f'<td class="sku-name">{str(r["품명"])[:22]}</td>'
        f'<td class="num">{int(r["잔여재고"]):,}</td>'
        f'<td class="num">{r["월평균속도"]:.1f}</td>'
        f'<td class="num">{r["예상소진일"]}</td>'
        f'<td class="num">{int(r["재고금액리스크"]):,}</td>'
        f'<td class="badge {cls}">{r["위험등급"]}</td></tr>'
    )
pred_table_html = "\n".join(pred_table_rows)

kiki_pct_val = ip_bin[ip_bin["IP구분"]=="키키IP"]["비중pct"].values
kiki_pct_str = f"{kiki_pct_val[0]:.1f}" if len(kiki_pct_val) else "0"

html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>와키윌리 25SS 모자 판매 분석 v3 리포트</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --navy:#1A1A2E; --accent:#E94560; --sub:#16213E;
    --light:#F0F4FF; --must:#4CAF50; --skip:#F44336;
    --watch:#FF9800; --white:#fff; --grey:#f5f6fa;
    --border:#e0e4f0; --text:#2c2c3e;
  }}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:"맑은 고딕","Malgun Gothic",sans-serif;background:var(--grey);color:var(--text);font-size:13px;}}
  .hero{{background:linear-gradient(135deg,var(--navy) 0%,var(--sub) 60%,var(--accent) 100%);padding:36px 40px 28px;color:var(--white);}}
  .hero h1{{font-size:26px;font-weight:800;letter-spacing:-0.5px;margin-bottom:6px;}}
  .hero .sub{{font-size:13px;opacity:0.75;}}
  .hero .tags{{margin-top:12px;display:flex;gap:8px;flex-wrap:wrap;}}
  .tag{{background:rgba(255,255,255,0.15);border-radius:20px;padding:3px 12px;font-size:11px;}}
  .container{{max-width:1400px;margin:0 auto;padding:24px;}}
  .section{{background:var(--white);border-radius:12px;padding:24px;margin-bottom:24px;box-shadow:0 2px 12px rgba(26,26,46,0.06);}}
  .section-title{{font-size:16px;font-weight:700;color:var(--navy);margin-bottom:16px;padding-bottom:10px;border-bottom:2px solid var(--accent);display:flex;align-items:center;gap:8px;}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;}}
  .kpi-card{{background:linear-gradient(135deg,var(--navy),var(--sub));border-radius:10px;padding:20px;color:var(--white);text-align:center;}}
  .kpi-card .val{{font-size:24px;font-weight:800;margin:8px 0 4px;}}
  .kpi-card .lbl{{font-size:11px;opacity:0.7;}}
  .kpi-card.accent{{background:linear-gradient(135deg,var(--accent),#c0392b);}}
  .kpi-card.green{{background:linear-gradient(135deg,#27ae60,#1e8449);}}
  .badge{{font-size:11px;border-radius:4px;padding:2px 6px;font-weight:600;white-space:nowrap;}}
  .badge.must{{background:#e8f5e9;color:#2e7d32;}}
  .badge.watch{{background:#fff3e0;color:#e65100;}}
  .badge.skip{{background:#ffebee;color:#c62828;}}
  .data-table{{width:100%;border-collapse:collapse;font-size:12px;}}
  .data-table th{{background:var(--navy);color:var(--white);padding:9px 10px;text-align:center;font-weight:600;white-space:nowrap;}}
  .data-table td{{padding:7px 10px;text-align:center;border-bottom:1px solid var(--border);}}
  .data-table tr:hover{{background:var(--light);}}
  .data-table .sku-code{{font-family:monospace;font-size:11px;font-weight:600;}}
  .data-table .sku-name{{text-align:left;}}
  .data-table .num{{text-align:right;font-variant-numeric:tabular-nums;}}
  .must{{color:var(--must);}} .skip{{color:var(--skip);}} .watch{{color:var(--watch);}}
  .heatmap-wrap{{overflow-x:auto;}}
  .heatmap-table{{border-collapse:collapse;font-size:11px;min-width:900px;}}
  .heatmap-table th{{background:var(--navy);color:var(--white);padding:7px 6px;text-align:center;white-space:nowrap;font-size:10px;}}
  .heatmap-table .sku-code{{font-family:monospace;font-weight:700;font-size:10px;text-align:left;padding:5px 8px;}}
  .heatmap-table .sku-name{{text-align:left;padding:5px 6px;white-space:nowrap;}}
  .heatmap-table .pattern{{font-size:10px;white-space:nowrap;}}
  .heatmap-table .heat-cell{{text-align:center;padding:4px 3px;width:48px;font-size:10px;border:1px solid #eee;}}
  .heatmap-table .heat-zero{{background:#fafbff;color:#ccc;text-align:center;padding:4px 3px;border:1px solid #eee;}}
  .heatmap-table .heat-total{{font-weight:700;background:var(--light);text-align:right;padding:4px 8px;}}
  .chart-grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:24px;}}
  .chart-grid-3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;}}
  .chart-box{{position:relative;height:260px;}}
  .chart-box-tall{{position:relative;height:340px;}}
  .chart-box-wide{{position:relative;height:300px;}}
  .판별-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:20px;}}
  .판별-card{{border-radius:10px;padding:16px;text-align:center;}}
  .판별-card.must{{background:#e8f5e9;border:2px solid var(--must);}}
  .판별-card.watch{{background:#fff3e0;border:2px solid var(--watch);}}
  .판별-card.skip{{background:#ffebee;border:2px solid var(--skip);}}
  .판별-card .big{{font-size:32px;font-weight:800;}}
  .판별-card .lbl{{font-size:12px;font-weight:600;margin-top:4px;}}
  .판별-card .desc{{font-size:10px;color:#666;margin-top:4px;}}
  .footer{{text-align:center;padding:24px;color:#999;font-size:11px;}}
  @media(max-width:900px){{.kpi-grid,.chart-grid-2,.chart-grid-3,.판별-grid{{grid-template-columns:1fr;}}}}
</style>
</head>
<body>

<div class="hero">
  <h1>🧢 와키윌리 25SS 모자 판매 분석 v3 리포트</h1>
  <div class="sub">담당: 데이터 인텔리전스 — 트렌드 애널리스트 + 인사이트 아키텍트</div>
  <div class="tags">
    <span class="tag">📅 분석기준: 2026-03-19</span>
    <span class="tag">📊 판매자료 162행</span>
    <span class="tag">🏗 구조화 59SKU</span>
    <span class="tag">📈 월별추이 15개월</span>
    <span class="tag">✨ v3: 속성심층분석+판매예측</span>
  </div>
</div>

<div class="container">

<div class="section">
  <div class="section-title">📌 핵심 KPI (v3 지표)</div>
  <div class="kpi-grid">
    <div class="kpi-card accent"><div class="lbl">총 누계판매금액</div><div class="val">{total_sales_a/1e8:.1f}억</div><div class="lbl">원</div></div>
    <div class="kpi-card"><div class="lbl">평균 판매율(수량)</div><div class="val">{avg_sell_thru:.1f}%</div><div class="lbl">발주 대비 판매</div></div>
    <div class="kpi-card"><div class="lbl">평균 실현이익율</div><div class="val">{avg_profit_r:.1f}%</div><div class="lbl">판매이익/실판매금액</div></div>
    <div class="kpi-card green"><div class="lbl">평균 효율지수</div><div class="val">{avg_효율지수:.1f}</div><div class="lbl">판매율×실현이익율/100</div></div>
  </div>
</div>

<div class="section">
  <div class="section-title">🎯 26FW 기획 판별 — MUST / WATCH / SKIP (Kano)</div>
  <div class="판별-grid">
    <div class="판별-card must">
      <div class="big" style="color:var(--must)">{must_cnt}</div>
      <div class="lbl">🟢 MUST — 재편성 확정</div>
      <div class="desc">판매율 ≥70% × 실현이익율 ≥40%</div>
    </div>
    <div class="판별-card watch">
      <div class="big" style="color:var(--watch)">{watch_cnt}</div>
      <div class="lbl">🟡 WATCH — 조건부 유지</div>
      <div class="desc">VE 검토 or QR 전환 필요</div>
    </div>
    <div class="판별-card skip">
      <div class="big" style="color:var(--skip)">{skip_cnt}</div>
      <div class="lbl">🔴 SKIP — 26FW 제외</div>
      <div class="desc">재고 소진 후 단종</div>
    </div>
  </div>
  <div class="chart-grid-3">
    <div class="chart-box"><canvas id="chart판별"></canvas></div>
    <div class="chart-box"><canvas id="chartIP"></canvas></div>
    <div class="chart-box"><canvas id="chartCat"></canvas></div>
  </div>
</div>

<div class="section">
  <div class="section-title">🏆 SKU 판매 랭킹 (Top 20)</div>
  <div class="chart-box-tall"><canvas id="chartSKU"></canvas></div>
  <div style="overflow-x:auto;margin-top:20px;">
    <table class="data-table">
      <thead><tr><th>품번</th><th>품명</th><th>대분류</th><th>활용</th><th>유형</th>
        <th>판매금액(원)</th><th>판매율%</th><th>실현이익율%</th><th>효율지수</th><th>26FW</th></tr></thead>
      <tbody>{top_table_html}</tbody>
    </table>
  </div>
</div>

<div class="section">
  <div class="section-title">📈 월별 판매수량 추이 — Top5 SKU</div>
  <div class="chart-box-wide"><canvas id="chartMonthly"></canvas></div>
</div>

<div class="section">
  <div class="section-title">🔵 판매율 × 실현이익율 매트릭스 (버블=판매금액)</div>
  <div style="font-size:11px;color:#888;margin-bottom:12px;">우상단 = MUST | X축=판매율 Y축=실현이익율 | <span style="color:var(--must)">●MUST</span> <span style="color:var(--watch)">●WATCH</span> <span style="color:var(--skip)">●SKIP</span></div>
  <div class="chart-box-wide"><canvas id="chartBubble"></canvas></div>
</div>

<div class="section">
  <div class="section-title">🎨 색상계열 판매율 & ⏱ 피크 판매월 분포</div>
  <div class="chart-grid-2">
    <div class="chart-box"><canvas id="chartColor"></canvas></div>
    <div class="chart-box"><canvas id="chartPeak"></canvas></div>
  </div>
</div>

<div class="section">
  <div class="section-title">🌡️ 품번 × 월별 판매수량 히트맵</div>
  <div class="heatmap-wrap">
    <table class="heatmap-table">
      <thead><tr><th>품번</th><th>품명</th><th>패턴</th>{heatmap_month_headers}<th>합계</th></tr></thead>
      <tbody>{heatmap_html}</tbody>
    </table>
  </div>
</div>

<div class="section">
  <div class="section-title">🔍 로고 속성 분석 — 표현방식별 판매율 & 실현이익율</div>
  <div class="chart-grid-2">
    <div class="chart-box"><canvas id="chartExpr"></canvas></div>
    <div class="chart-box"><canvas id="chartLogoSz"></canvas></div>
  </div>
</div>

<div class="section">
  <div class="section-title">🏗️ 구조 속성 분석 — 브림·소재별 효율지수</div>
  <div class="chart-grid-2">
    <div class="chart-box"><canvas id="chartBrim"></canvas></div>
    <div class="chart-box"><canvas id="chartMat"></canvas></div>
  </div>
</div>

<div class="section">
  <div class="section-title">📦 잔여재고 판매 예측 — 소진 위험도</div>
  <div class="chart-grid-2">
    <div class="chart-box"><canvas id="chartRisk"></canvas></div>
    <div style="overflow-x:auto;">
      <table class="data-table">
        <thead><tr><th>품번</th><th>품명</th><th>잔여재고</th><th>月평균속도</th><th>예상소진일</th><th>재고금액리스크(원)</th><th>위험등급</th></tr></thead>
        <tbody>{pred_table_html}</tbody>
      </table>
    </div>
  </div>
</div>

<div class="section">
  <div class="section-title">🆕 26SS 초기 판매 추이 — WA26 월별 판매 (2026-01~03)</div>
  <div style="font-size:11px;color:#888;margin-bottom:10px;">25SS 동기(2025-01~03) 대비 초기 3개월 판매 비교 | 26SS 평균: {wa26_avg_3m:.0f}개 vs 25SS: {wa25_avg_3m:.0f}개</div>
  <div class="chart-box-wide"><canvas id="chart26SS"></canvas></div>
</div>

</div>
<div class="footer">Generated by FPOF 데이터 인텔리전스 v3 | 와키윌리 패션 하우스 오케스트레이션 | 2026-03-19</div>

<script>
const COLORS_PALETTE = ['#E94560','#0F3460','#533483','#16213E','#1a6b4a'];
const MONTHS = {chart_monthly_labels};

new Chart(document.getElementById('chartSKU'),{{
  type:'bar',
  data:{{labels:{chart_sku_labels},datasets:[{{label:'판매금액(원)',data:{chart_sku_sales},backgroundColor:{chart_sku_colors},borderRadius:6}}]}},
  options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>`  ${{Number(ctx.raw).toLocaleString()}}원`}}}}}},
    scales:{{x:{{ticks:{{callback:v=>(v/1e8).toFixed(1)+'억'}},grid:{{color:'#f0f0f0'}}}},y:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});

new Chart(document.getElementById('chartMonthly'),{{
  type:'line',
  data:{{labels:MONTHS,datasets:{chart_monthly_data}}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}},
    scales:{{y:{{beginAtZero:true,grid:{{color:'#f0f0f0'}}}},x:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});

new Chart(document.getElementById('chart판별'),{{
  type:'doughnut',
  data:{{labels:{chart_판별_labels},datasets:[{{data:{chart_판별_data},backgroundColor:['#4CAF50','#FF9800','#F44336'],borderWidth:2,borderColor:'#fff'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},title:{{display:true,text:'26FW 판별 분포'}}}}}}
}});

new Chart(document.getElementById('chartIP'),{{
  type:'doughnut',
  data:{{labels:{chart_ip_labels},datasets:[{{data:{chart_ip_data},backgroundColor:COLORS_PALETTE,borderWidth:2,borderColor:'#fff'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},title:{{display:true,text:'IP 활용별 판매금액'}}}}}}
}});

new Chart(document.getElementById('chartCat'),{{
  type:'doughnut',
  data:{{labels:{chart_cat_labels},datasets:[{{data:{chart_cat_data},backgroundColor:['#1A1A2E','#E94560','#533483','#16213E'],borderWidth:2,borderColor:'#fff'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},title:{{display:true,text:'대분류별 판매금액'}}}}}}
}});

new Chart(document.getElementById('chartColor'),{{
  type:'bar',
  data:{{labels:{chart_color_labels},datasets:[
    {{label:'판매율(%)',data:{chart_color_sell},backgroundColor:'#E94560aa',yAxisID:'y',borderRadius:4}},
    {{label:'판매금액',data:{chart_color_sales},backgroundColor:'#1A1A2Eaa',yAxisID:'y1',borderRadius:4}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom'}},title:{{display:true,text:'색상계열별 판매율 & 판매금액'}}}},
    scales:{{y:{{position:'left',max:100,ticks:{{callback:v=>v+'%'}}}},y1:{{position:'right',ticks:{{callback:v=>(v/1e8).toFixed(1)+'억'}},grid:{{drawOnChartArea:false}}}}}}
  }}
}});

new Chart(document.getElementById('chartPeak'),{{
  type:'bar',
  data:{{labels:{chart_peak_labels},datasets:[{{label:'SKU 수',data:{chart_peak_data},backgroundColor:'#0F3460cc',borderRadius:6}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},title:{{display:true,text:'품번별 피크 판매월 분포'}}}},
    scales:{{y:{{beginAtZero:true,ticks:{{stepSize:1}}}},x:{{ticks:{{font:{{size:9}}}}}}}}
  }}
}});

const bubbleRaw = {chart_bubble};
new Chart(document.getElementById('chartBubble'),{{
  type:'bubble',
  data:{{datasets:[
    {{label:'MUST', data:bubbleRaw.filter(d=>d.판별.includes('MUST')).map(d=>{{return{{x:d.x,y:d.y,r:d.r,label:d.label}}}}), backgroundColor:'rgba(76,175,80,0.6)',borderColor:'rgba(76,175,80,0.9)'}},
    {{label:'WATCH',data:bubbleRaw.filter(d=>d.판별.includes('WATCH')).map(d=>{{return{{x:d.x,y:d.y,r:d.r,label:d.label}}}}),backgroundColor:'rgba(255,152,0,0.6)',borderColor:'rgba(255,152,0,0.9)'}},
    {{label:'SKIP', data:bubbleRaw.filter(d=>d.판별.includes('SKIP')).map(d=>{{return{{x:d.x,y:d.y,r:d.r,label:d.label}}}}), backgroundColor:'rgba(244,67,54,0.6)',borderColor:'rgba(244,67,54,0.9)'}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom'}},tooltip:{{callbacks:{{label:ctx=>`${{ctx.raw.label}} 판매율:${{ctx.raw.x}}% 이익율:${{ctx.raw.y}}%`}}}}}},
    scales:{{
      x:{{title:{{display:true,text:'판매율 (%)'}},min:0,max:110,ticks:{{callback:v=>v+'%'}}}},
      y:{{title:{{display:true,text:'실현이익율 (%)'}},ticks:{{callback:v=>v+'%'}}}}
    }}
  }}
}});

// ── 로고속성: 표현방식별 바차트
new Chart(document.getElementById('chartExpr'),{{
  type:'bar',
  data:{{labels:{chart_expr_labels},datasets:[
    {{label:'판매율(%)',data:{chart_expr_sell},backgroundColor:'#E94560bb',borderRadius:4,yAxisID:'y'}},
    {{label:'실현이익율(%)',data:{chart_expr_profit},backgroundColor:'#0F3460bb',borderRadius:4,yAxisID:'y'}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom'}},title:{{display:true,text:'표현방식별 판매율 & 실현이익율'}}}},
    scales:{{y:{{beginAtZero:true,ticks:{{callback:v=>v+'%'}}}},x:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});

// ── 로고사이즈별 바차트
const logoSzLabels = {json.dumps([str(v) for v in attr_logo_sz["로고사이즈"].fillna("미분류")], ensure_ascii=False)};
const logoSzSell   = {json.dumps([float(v) for v in attr_logo_sz["판매율avg"]])};
const logoSzProfit = {json.dumps([float(v) for v in attr_logo_sz["실현이익율avg"]])};
new Chart(document.getElementById('chartLogoSz'),{{
  type:'bar',
  data:{{labels:logoSzLabels,datasets:[
    {{label:'판매율(%)',data:logoSzSell,backgroundColor:'#533483bb',borderRadius:4,yAxisID:'y'}},
    {{label:'실현이익율(%)',data:logoSzProfit,backgroundColor:'#16213Ebb',borderRadius:4,yAxisID:'y'}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom'}},title:{{display:true,text:'로고사이즈별 판매율 & 실현이익율'}}}},
    scales:{{y:{{beginAtZero:true,ticks:{{callback:v=>v+'%'}}}},x:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});

// ── 브림별 효율지수
new Chart(document.getElementById('chartBrim'),{{
  type:'bar',
  data:{{labels:{chart_brim_labels},datasets:[{{label:'효율지수',data:{chart_brim_eff},backgroundColor:'#E94560cc',borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',
    plugins:{{legend:{{display:false}},title:{{display:true,text:'브림 유형별 효율지수 (판매율×실현이익율/100)'}}}},
    scales:{{x:{{beginAtZero:true}},y:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});

// ── 소재별 효율지수
new Chart(document.getElementById('chartMat'),{{
  type:'bar',
  data:{{labels:{chart_mat_labels},datasets:[{{label:'효율지수',data:{chart_mat_eff},backgroundColor:'#0F3460cc',borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',
    plugins:{{legend:{{display:false}},title:{{display:true,text:'소재별 효율지수'}}}},
    scales:{{x:{{beginAtZero:true}},y:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});

// ── 위험등급별 재고금액
new Chart(document.getElementById('chartRisk'),{{
  type:'doughnut',
  data:{{labels:{chart_risk_labels},datasets:[{{data:{chart_risk_data},backgroundColor:{chart_risk_colors},borderWidth:2,borderColor:'#fff'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},title:{{display:true,text:'위험등급별 잔재고 금액(원)'}}}}
  }}
}});

// ── 26SS 초기 추이
new Chart(document.getElementById('chart26SS'),{{
  type:'line',
  data:{{labels:{chart_26ss_labels},datasets:{chart_26ss_data}}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}},title:{{display:true,text:'26SS WA26 초기 월별 판매수량 (Top5)'}}}},
    scales:{{y:{{beginAtZero:true,ticks:{{callback:v=>v.toLocaleString()}}}},x:{{ticks:{{font:{{size:10}}}}}}}}
  }}
}});
</script>
</body>
</html>"""

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(html)
print(f"   ✓ HTML 저장: {os.path.basename(OUT_HTML)}")

# ════════════════════════════════════════════════════════════════════════════
# 10. .fpof-state.json 업데이트
# ════════════════════════════════════════════════════════════════════════════
print("[10/10] 상태 업데이트...")
import json as _json

STATE_PATH = os.path.join(BASE, ".fpof-state.json")
with open(STATE_PATH, "r", encoding="utf-8") as f:
    state = _json.load(f)

if "operational" not in state: state["operational"] = {}
if "recent_outputs" not in state["operational"]: state["operational"]["recent_outputs"] = []

state["operational"]["recent_outputs"] = [
    e for e in state["operational"]["recent_outputs"]
    if "check_sales-analysis_v3" not in str(e)
]
state["operational"]["recent_outputs"].append({
    "date": "2026-03-19",
    "type": "check",
    "version": "v3",
    "files": [
        "output/26FW/season-strategy/check_sales-analysis_v3_2026-03-19.xlsx",
        "output/26FW/season-strategy/check_sales-analysis_v3_2026-03-19.html",
    ],
    "summary": (f"25SS 모자 판매분석v3 — SKU {total_sku}개, 판매율 {avg_sell_thru:.1f}%, "
                f"실현이익율 {avg_profit_r:.1f}%, 효율지수 {avg_효율지수:.1f}, "
                f"MUST {must_cnt}/WATCH {watch_cnt}/SKIP {skip_cnt}, "
                f"속성심층+판매예측+26SS 14시트")
})

with open(STATE_PATH, "w", encoding="utf-8") as f:
    _json.dump(state, f, ensure_ascii=False, indent=2)

print("\n" + "="*60)
print("✅ 분석 v3 완료")
print("="*60)
print(f"  📊 Excel (14시트): {os.path.basename(OUT_XLSX)}")
print(f"  🌐 HTML 리포트:    {os.path.basename(OUT_HTML)}")
print(f"\n  KPI:")
print(f"  - 총 판매금액:     {total_sales_a/1e8:.2f}억원")
print(f"  - 평균 판매율:     {avg_sell_thru:.1f}%")
print(f"  - 평균 실현이익율: {avg_profit_r:.1f}%")
print(f"  - 평균 효율지수:   {avg_효율지수:.1f}")
print(f"  - MUST/WATCH/SKIP: {must_cnt}/{watch_cnt}/{skip_cnt}")
print(f"  - 26SS WA26:       {wa26_sku_cnt}품번")
print("="*60)
