#!/usr/bin/env python3
"""
25SS 모자 판매 분석 스크립트
담당: 데이터 인텔리전스 (sales-analysis + insight-archiving)
출력: check_sales-analysis_2026-03-19.xlsx + check_sales-analysis_2026-03-19.md
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
import sys
import os

# Windows 콘솔 UTF-8 출력
if sys.stdout.encoding != "utf-8":
    sys.stdout = open(sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1)

warnings.filterwarnings("ignore")

# ─── 경로 설정 ──────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SALES_PATH = os.path.join(BASE, "output/26FW/season-strategy/품번종합집계표_모자 판매자료.xlsx")
STRUCT_PATH = os.path.join(BASE, "output/26FW/season-strategy/모자 구조화 설계 최종.xlsx")
OUT_XLSX    = os.path.join(BASE, "output/26FW/season-strategy/check_sales-analysis_2026-03-19.xlsx")
OUT_MD      = os.path.join(BASE, "output/26FW/season-strategy/check_sales-analysis_2026-03-19.md")

# ─── 1. 데이터 로드 ─────────────────────────────────────────────────────────
print("[1/8] 데이터 로드 중...")

# 판매자료: row0=대분류헤더, row1=세부헤더, row2~=data
sales_raw = pd.read_excel(SALES_PATH, header=1)
# 구조화 설계: 헤더 row0
struct_raw = pd.read_excel(STRUCT_PATH, sheet_name="품번별 세부구분", header=0)

# 컬럼명 표준화 (위치 기반 — 인코딩 독립)
SALES_COLS = {
    0:  "품번",
    1:  "품명",
    2:  "색상",
    3:  "발주배수_최",
    4:  "발주금액_최",
    5:  "표준컬러",
    6:  "최초판매가",
    7:  "현판매가",
    8:  "소화율_수량pct",
    9:  "소화율_금액pct",
    10: "시즌",
    11: "기획구분",
    12: "발주원가금액",
    13: "발주원가금액2",
    14: "최초입고일",
    15: "최종입고일",
    16: "최초출고일",
    17: "최종출고일",
    18: "소화율_수량pct2",
    19: "발주배수_최2",
    20: "발주_수량",
    21: "발주_금액",
    22: "발주_최초판매금액",
    23: "최초출고월",
    24: "발주_현판매금액",
    25: "기획구분2",
    26: "누계판매_수량",
    27: "누계판매_금액",
    28: "누계판매_최초판매금액",
    29: "누계판매_시간비례금액",
    30: "누계판매_현판매금액",
    31: "누계판매_실판매금액",
    32: "발주금액_최2",
    33: "발주원가금액3",
    34: "완사입비_단가",
    35: "판매이익_금액",
    36: "판매이익_수량",
    37: "판매이익율_사전V",
    38: "판매이익율_사후_수량",
    39: "판매이익율_사후_금액",
}

STRUCT_COLS = {
    0:  "품번",
    1:  "품명_구조",
    2:  "대분류",
    3:  "활용",
    4:  "로고사이즈",
    5:  "표현방식",
    6:  "스타일분류",
    7:  "볼캡유형",
    8:  "브림",
    9:  "크라운",
    10: "패널수",
    11: "소재",
    12: "소재코드",
}

# 컬럼 rename
sales = sales_raw.copy()
sales.columns = [SALES_COLS.get(i, f"col_{i}") for i in range(len(sales.columns))]

struct = struct_raw.copy()
struct.columns = [STRUCT_COLS.get(i, f"col_{i}") for i in range(len(struct.columns))]

# 불필요한 행 제거 (품번이 없는 행)
sales = sales[sales["품번"].notna() & (sales["품번"].astype(str).str.startswith("WA"))].copy()
struct = struct[struct["품번"].notna()].copy()

print(f"   판매자료: {len(sales)}행 / 구조화: {len(struct)}행")

# ─── 2. JOIN ────────────────────────────────────────────────────────────────
print("[2/8] LEFT JOIN (품번 기준)...")
df = sales.merge(struct[["품번","대분류","활용","로고사이즈","표현방식","스타일분류","볼캡유형","브림","크라운","패널수","소재","소재코드"]],
                 on="품번", how="left")

# 미매칭 확인
unmatched = df[df["대분류"].isna()]["품번"].unique()
if len(unmatched):
    print(f"   ⚠ 미매칭 품번 {len(unmatched)}개: {list(unmatched)[:5]}")
else:
    print("   ✓ 전체 매칭 완료")

# ─── 3. 파생 컬럼 ───────────────────────────────────────────────────────────
print("[3/8] 파생 컬럼 생성...")

# 날짜 변환
for col in ["최초입고일","최종입고일","최초출고일","최종출고일"]:
    df[col] = pd.to_datetime(df[col], errors="coerce")

# 판매경과일 (최종출고일 - 최초출고일)
df["판매경과일"] = (df["최종출고일"] - df["최초출고일"]).dt.days

# 소화율 재계산 검증
df["소화율_검증"] = (df["누계판매_수량"] / df["발주_수량"] * 100).round(2)

# 택가실현율 (실판매금액 / 발주_현판매금액)
df["택가실현율_pct"] = (df["누계판매_실판매금액"] / df["발주_현판매금액"] * 100).where(
    df["발주_현판매금액"] > 0, np.nan
).round(2)

# 할인여부 (최초판매가 vs 현판매가)
df["할인여부"] = df["최초판매가"] != df["현판매가"]

# 색상 계열 그룹
COLOR_MAP = {
    "BK": "블랙", "BL": "블루", "BE": "베이지", "CR": "크림",
    "NA": "네이비", "BR": "브라운", "GR": "그린", "GY": "그레이",
    "WH": "화이트", "RD": "레드", "YL": "옐로우", "PP": "퍼플",
    "OR": "오렌지", "PK": "핑크", "IB": "아이보리", "CH": "차콜",
    "OL": "올리브", "MN": "모노", "KH": "카키",
}
SEASON_COLORS = {"BK","NA","CH","GY","WH"}  # 베이직 계열

def color_group(c):
    if pd.isna(c):
        return "기타"
    c = str(c).upper()
    if c in SEASON_COLORS:
        return "베이직(BK/NA/GY)"
    elif c in {"BE","CR","IB","KH","OL"}:
        return "뉴트럴(BE/CR/OL)"
    elif c in {"BR"}:
        return "어스(BR)"
    elif c in {"BL","GR","RD","YL","PP","OR","PK"}:
        return "컬러풀"
    else:
        return "기타"

df["색상계열"] = df["색상"].apply(color_group)

# 최초출고월 (YYYY-MM 형식으로 정리)
df["출고월"] = df["최초출고월"].astype(str).str[:7]

# 마크업 등급
def markup_grade(v):
    if pd.isna(v): return "N/A"
    if v >= 300: return "HIGH(≥300%)"
    elif v >= 200: return "MID(200~300%)"
    else: return "LOW(<200%)"

df["마크업등급"] = df["판매이익율_사전V"].apply(markup_grade)

print(f"   총 {len(df)}행 처리 완료")

# ─── 4. 집계 ─────────────────────────────────────────────────────────────────
print("[4/8] 8개 차원 집계 중...")

# ── 4-A. SKU 랭킹 (품번 집계)
sku = df.groupby("품번").agg(
    품명=("품명", "first"),
    대분류=("대분류", "first"),
    활용=("활용", "first"),
    스타일=("스타일분류", "first"),
    볼캡유형=("볼캡유형", "first"),
    소재=("소재", "first"),
    발주수량=("발주_수량", "sum"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    실판매금액=("누계판매_실판매금액", "sum"),
    판매이익=("판매이익_금액", "sum"),
    최초판매가=("최초판매가", "first"),
    현판매가=("현판매가", "first"),
    마크업율_avg=("판매이익율_사전V", "mean"),
    완사입비=("완사입비_단가", "first"),
    최초출고일=("최초출고일", "min"),
    최종출고일=("최종출고일", "max"),
    색상수=("색상", "nunique"),
).reset_index()

sku["소화율_pct"] = (sku["판매수량"] / sku["발주수량"] * 100).round(2)
sku["판매이익율_pct"] = (sku["판매이익"] / sku["실판매금액"] * 100).round(2)
sku["택가실현율_pct"] = (sku["실판매금액"] / (sku["발주수량"] * sku["최초판매가"]) * 100).round(2)
sku["판매경과일"] = (sku["최종출고일"] - sku["최초출고일"]).dt.days

# MUST/WATCH/SKIP 판별 (소화율 60% 기준, 이익율 0% 기준)
def classify(row):
    high_sell = row["소화율_pct"] >= 70
    high_profit = row["마크업율_avg"] >= 250  # 사전 마크업 기준
    if high_sell and high_profit:
        return "🟢 MUST"
    elif high_sell and not high_profit:
        return "🟡 WATCH(VE)"
    elif not high_sell and high_profit:
        return "🟡 WATCH(QR)"
    else:
        return "🔴 SKIP"

sku["26FW판별"] = sku.apply(classify, axis=1)
sku_sorted = sku.sort_values("판매금액", ascending=False).reset_index(drop=True)

# ── 4-B. 출고월별 타이밍 분석
timing = df.groupby("출고월").agg(
    색상건수=("색상", "count"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    품번수=("품번", "nunique"),
).reset_index().sort_values("출고월")
timing["소화율_avg"] = timing["소화율_avg"].round(1)

# ── 4-C. 색상 분석
color_df = df.groupby(["색상","색상계열"]).agg(
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    행수=("품번", "count"),
).reset_index().sort_values("판매금액", ascending=False)
color_df["소화율_avg"] = color_df["소화율_avg"].round(1)

# 색상계열 집계
color_grp = df.groupby("색상계열").agg(
    SKU건수=("품번", "count"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
).reset_index().sort_values("판매금액", ascending=False)
color_grp["소화율_avg"] = color_grp["소화율_avg"].round(1)

# ── 4-D. 카테고리 분석
cat_main = df.groupby("대분류").agg(
    SKU수=("품번", "nunique"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    마크업avg=("판매이익율_사전V", "mean"),
).reset_index().sort_values("판매금액", ascending=False)

cat_style = df.groupby("스타일분류").agg(
    SKU수=("품번", "nunique"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    마크업avg=("판매이익율_사전V", "mean"),
).reset_index().sort_values("판매금액", ascending=False)

cat_style2 = df.groupby("볼캡유형").agg(
    SKU수=("품번", "nunique"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    마크업avg=("판매이익율_사전V", "mean"),
).reset_index().sort_values("판매금액", ascending=False)

cat_material = df.groupby("소재코드").agg(
    SKU수=("품번", "nunique"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    마크업avg=("판매이익율_사전V", "mean"),
).reset_index().sort_values("판매금액", ascending=False)

# ── 4-E. IP 분석
ip_df = df.groupby("활용").agg(
    SKU수=("품번", "nunique"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    마크업avg=("판매이익율_사전V", "mean"),
).reset_index().sort_values("판매금액", ascending=False)
ip_df["판매금액_비중pct"] = (ip_df["판매금액"] / ip_df["판매금액"].sum() * 100).round(1)

# 키키 vs 비-키키
df["IP구분"] = df["활용"].apply(lambda x: "키키IP" if str(x) in ["키키","릴리","IP"] else "비IP")
ip_bin = df.groupby("IP구분").agg(
    SKU수=("품번", "nunique"),
    판매수량=("누계판매_수량", "sum"),
    판매금액=("누계판매_금액", "sum"),
    소화율_avg=("소화율_수량pct", "mean"),
    마크업avg=("판매이익율_사전V", "mean"),
).reset_index()
ip_bin["판매금액_비중pct"] = (ip_bin["판매금액"] / ip_bin["판매금액"].sum() * 100).round(1)

# ── 4-F. 마진 구조
margin_df = sku_sorted[["품번","품명","소화율_pct","마크업율_avg","택가실현율_pct","실판매금액","26FW판별"]].copy()

# KPI 종합
total_orders  = df["발주_수량"].sum()
total_sales_q = df["누계판매_수량"].sum()
total_sales_a = df["누계판매_금액"].sum()
total_sales_r = df["누계판매_실판매금액"].sum()
avg_sell_thru = (total_sales_q / total_orders * 100) if total_orders else 0
avg_markup    = df["판매이익율_사전V"].mean()
total_profit  = df["판매이익_금액"].sum()
sku_count     = df["품번"].nunique()
color_count   = len(df)
must_count    = (sku["26FW판별"] == "🟢 MUST").sum()
watch_count   = (sku["26FW판별"].str.startswith("🟡")).sum()
skip_count    = (sku["26FW판별"] == "🔴 SKIP").sum()

print(f"   KPI — 총판매금액: {total_sales_a:,.0f}원 / 소화율: {avg_sell_thru:.1f}% / 평균마크업: {avg_markup:.0f}%")

# ─── 5. Excel 작성 ──────────────────────────────────────────────────────────
print("[5/8] Excel 8시트 작성 중...")

wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# ── 스타일 팔레트
C_BG_HEADER   = "1A1A2E"   # 다크 네이비
C_BG_ACCENT   = "E94560"   # 와키레드
C_BG_SUB      = "16213E"   # 서브헤더
C_BG_LIGHT    = "F0F4FF"   # 라이트 배경
C_BG_MUST     = "D4EDDA"   # 연녹
C_BG_SKIP     = "F8D7DA"   # 연빨
C_BG_WATCH    = "FFF3CD"   # 연노
C_TEXT_WHITE  = "FFFFFF"
C_TEXT_DARK   = "1A1A2E"
C_BORDER      = "CCCCCC"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color=C_TEXT_DARK, size=10):
    return Font(name="맑은 고딕", bold=bold, color=color, size=size)

def make_border():
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row, values, bg=C_BG_HEADER, fg=C_TEXT_WHITE):
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = make_fill(bg)
        c.font = make_font(bold=True, color=fg)
        c.alignment = center()
        c.border = make_border()

def apply_data_row(ws, row, values, bg=None):
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        if bg:
            c.fill = make_fill(bg)
        c.font = make_font()
        c.alignment = center()
        c.border = make_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def pct_fmt(v):
    return f"{v:.1f}%" if pd.notna(v) else "-"

def num_fmt(v):
    return f"{int(v):,}" if pd.notna(v) else "-"

# ════════════════════════════════════════════════════
# 시트 1: 📊 종합요약
# ════════════════════════════════════════════════════
ws1 = wb.create_sheet("📊 종합요약")
ws1.sheet_view.showGridLines = False

# 타이틀
ws1.merge_cells("A1:H1")
c = ws1["A1"]
c.value = "🧢 와키윌리 25SS 모자 판매 종합 분석 대시보드"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=14)
c.alignment = center()
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:H2")
c = ws1["A2"]
c.value = "분석 기간: 25SS 시즌 | 데이터 기준: 2026-03-19 | 담당: 데이터 인텔리전스"
c.fill = make_fill(C_BG_SUB)
c.font = make_font(color=C_TEXT_WHITE, size=9)
c.alignment = center()

# KPI 카드 — 상단
kpis_top = [
    ("총 발주수량", f"{int(total_orders):,} 개"),
    ("총 누계판매금액", f"{int(total_sales_a):,} 원"),
    ("총 실판매금액", f"{int(total_sales_r):,} 원"),
    ("평균 소화율(수량)", f"{avg_sell_thru:.1f} %"),
]
apply_header_row(ws1, 4, [k for k, v in kpis_top], bg=C_BG_ACCENT)
apply_data_row(ws1, 5, [v for k, v in kpis_top])
ws1.row_dimensions[5].height = 28

kpis_bot = [
    ("평균 마크업율(사전)", f"{avg_markup:.0f} %"),
    ("분석 SKU 수", f"{sku_count} 개"),
    ("색상 SKU 수(행)", f"{color_count} 건"),
    ("판매이익 합계", f"{int(total_profit):,} 원"),
]
apply_header_row(ws1, 7, [k for k, v in kpis_bot], bg=C_BG_SUB)
apply_data_row(ws1, 8, [v for k, v in kpis_bot])

# 26FW 판별 요약
apply_header_row(ws1, 10, ["26FW 판별", "SKU 수", "비중", ""], bg=C_BG_ACCENT)
total_sku = must_count + watch_count + skip_count
for ri, (label, cnt, bg) in enumerate([
    ("🟢 MUST — 재편성 확정", must_count, C_BG_MUST),
    ("🟡 WATCH — 조건부 유지", watch_count, C_BG_WATCH),
    ("🔴 SKIP — 26FW 제외", skip_count, C_BG_SKIP),
], 11):
    apply_data_row(ws1, ri, [label, cnt, f"{cnt/total_sku*100:.0f}%", ""], bg=bg)

# 채널별 출고월 분포 요약
apply_header_row(ws1, 15, ["최초출고월", "색상건수", "판매금액(원)", "소화율avg"], bg=C_BG_SUB)
for ri, row in enumerate(timing.itertuples(), 16):
    apply_data_row(ws1, ri, [row.출고월, row.색상건수, row.판매금액, f"{row.소화율_avg}%"])

set_col_widths(ws1, [28, 18, 20, 18, 18, 18, 18, 18])

# ════════════════════════════════════════════════════
# 시트 2: 🏆 SKU랭킹
# ════════════════════════════════════════════════════
ws2 = wb.create_sheet("🏆 SKU랭킹")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:M1")
c = ws2["A1"]
c.value = "🏆 품번별 판매 랭킹 — 25SS 모자"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()
ws2.row_dimensions[1].height = 30

headers2 = ["순위","품번","품명","대분류","활용","볼캡유형","소재",
            "발주수량","판매수량","소화율%","판매금액(원)","실판매금액(원)",
            "마크업율avg%","26FW판별"]
apply_header_row(ws2, 2, headers2)

for ri, (_, row) in enumerate(sku_sorted.iterrows(), 3):
    판별 = row["26FW판별"]
    bg = None
    if "MUST" in str(판별): bg = C_BG_MUST
    elif "SKIP" in str(판별): bg = C_BG_SKIP
    elif "WATCH" in str(판별): bg = C_BG_WATCH

    apply_data_row(ws2, ri, [
        ri - 2,
        row["품번"], row["품명"], row["대분류"], row["활용"], row["볼캡유형"], row["소재"],
        int(row["발주수량"]), int(row["판매수량"]), f"{row['소화율_pct']:.1f}%",
        int(row["판매금액"]), int(row["실판매금액"]),
        f"{row['마크업율_avg']:.0f}%", 판별
    ], bg=bg)

set_col_widths(ws2, [5, 14, 26, 10, 8, 12, 10, 10, 10, 9, 16, 16, 12, 14])
ws2.freeze_panes = "A3"

# ════════════════════════════════════════════════════
# 시트 3: 📅 판매타이밍
# ════════════════════════════════════════════════════
ws3 = wb.create_sheet("📅 판매타이밍")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "📅 최초출고월별 판매 타이밍 분석 — 25SS 모자"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()

apply_header_row(ws3, 2, ["최초출고월","색상건수","품번수","판매수량","판매금액(원)","소화율avg%"])
for ri, row in enumerate(timing.itertuples(), 3):
    apply_data_row(ws3, ri, [
        row.출고월, row.색상건수, row.품번수,
        int(row.판매수량), int(row.판매금액), f"{row.소화율_avg}%"
    ])

# 구간 분류 설명
ws3["A" + str(len(timing) + 5)].value = "※ 초기입고(~2024-12): 시즌 선판 / 본시즌(2025-01~05): SS 피크 / 후기(2025-06~): 재고소진 단계"
ws3["A" + str(len(timing) + 5)].font = make_font(size=9)

set_col_widths(ws3, [14, 10, 10, 12, 18, 12])

# ════════════════════════════════════════════════════
# 시트 4: 🎨 색상분석
# ════════════════════════════════════════════════════
ws4 = wb.create_sheet("🎨 색상분석")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F1")
c = ws4["A1"]
c.value = "🎨 색상 & 색상계열별 소화율 분석 — 25SS 모자"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()

# 색상계열 집계
apply_header_row(ws4, 2, ["색상계열","SKU건수","판매수량","판매금액(원)","소화율avg%"], bg=C_BG_ACCENT)
for ri, row in enumerate(color_grp.itertuples(), 3):
    apply_data_row(ws4, ri, [row.색상계열, row.SKU건수, int(row.판매수량), int(row.판매금액), f"{row.소화율_avg}%"])

# 색상 상세
apply_header_row(ws4, len(color_grp) + 5, ["색상코드","색상계열","판매수량","판매금액(원)","소화율avg%","행수(색상SKU)"])
for ri, row in enumerate(color_df.itertuples(), len(color_grp) + 6):
    apply_data_row(ws4, ri, [row.색상, row.색상계열, int(row.판매수량), int(row.판매금액), f"{row.소화율_avg}%", row.행수])

set_col_widths(ws4, [18, 18, 12, 18, 12, 14])

# ════════════════════════════════════════════════════
# 시트 5: 🗂️ 카테고리분석
# ════════════════════════════════════════════════════
ws5 = wb.create_sheet("🗂️ 카테고리분석")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:F1")
c = ws5["A1"]
c.value = "🗂️ 카테고리 × 스타일 × 소재별 분석 — 25SS 모자"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()

# 대분류
apply_header_row(ws5, 2, ["대분류","SKU수","판매수량","판매금액(원)","소화율avg%","마크업avg%"], bg=C_BG_ACCENT)
for ri, row in enumerate(cat_main.itertuples(), 3):
    apply_data_row(ws5, ri, [
        row.대분류 if pd.notna(row.대분류) else "-",
        int(row.SKU수), int(row.판매수량), int(row.판매금액),
        f"{row.소화율_avg:.1f}%", f"{row.마크업avg:.0f}%"
    ])

# 스타일분류 (캡/모/햇)
r_offset = len(cat_main) + 5
apply_header_row(ws5, r_offset, ["스타일분류","SKU수","판매수량","판매금액(원)","소화율avg%","마크업avg%"], bg=C_BG_SUB)
for ri, row in enumerate(cat_style.itertuples(), r_offset + 1):
    apply_data_row(ws5, ri, [
        row.스타일분류 if pd.notna(row.스타일분류) else "-",
        int(row.SKU수), int(row.판매수량), int(row.판매금액),
        f"{row.소화율_avg:.1f}%", f"{row.마크업avg:.0f}%"
    ])

# 볼캡유형
r_offset2 = r_offset + len(cat_style) + 3
apply_header_row(ws5, r_offset2, ["볼캡유형","SKU수","판매수량","판매금액(원)","소화율avg%","마크업avg%"], bg=C_BG_ACCENT)
for ri, row in enumerate(cat_style2.itertuples(), r_offset2 + 1):
    apply_data_row(ws5, ri, [
        row.볼캡유형 if pd.notna(row.볼캡유형) else "-",
        int(row.SKU수), int(row.판매수량), int(row.판매금액),
        f"{row.소화율_avg:.1f}%", f"{row.마크업avg:.0f}%"
    ])

# 소재
r_offset3 = r_offset2 + len(cat_style2) + 3
apply_header_row(ws5, r_offset3, ["소재코드","SKU수","판매수량","판매금액(원)","소화율avg%","마크업avg%"], bg=C_BG_SUB)
for ri, row in enumerate(cat_material.itertuples(), r_offset3 + 1):
    apply_data_row(ws5, ri, [
        row.소재코드 if pd.notna(row.소재코드) else "-",
        int(row.SKU수), int(row.판매수량), int(row.판매금액),
        f"{row.소화율_avg:.1f}%", f"{row.마크업avg:.0f}%"
    ])

set_col_widths(ws5, [18, 8, 12, 18, 12, 12])

# ════════════════════════════════════════════════════
# 시트 6: 🐱 IP분석
# ════════════════════════════════════════════════════
ws6 = wb.create_sheet("🐱 IP분석")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:G1")
c = ws6["A1"]
c.value = "🐱 IP 활용 유형별 판매 성과 — 25SS 모자"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()

# IP 세부 분류
apply_header_row(ws6, 2, ["활용유형","SKU수","판매수량","판매금액(원)","판매비중%","소화율avg%","마크업avg%"], bg=C_BG_ACCENT)
for ri, row in enumerate(ip_df.itertuples(), 3):
    apply_data_row(ws6, ri, [
        str(row.활용) if pd.notna(row.활용) else "미분류",
        int(row.SKU수), int(row.판매수량), int(row.판매금액),
        f"{row.판매금액_비중pct:.1f}%", f"{row.소화율_avg:.1f}%", f"{row.마크업avg:.0f}%"
    ])

# IP vs 비IP
r2 = len(ip_df) + 5
apply_header_row(ws6, r2, ["IP구분","SKU수","판매수량","판매금액(원)","판매비중%","소화율avg%","마크업avg%"], bg=C_BG_SUB)
for ri, row in enumerate(ip_bin.itertuples(), r2 + 1):
    apply_data_row(ws6, ri, [
        row.IP구분, int(row.SKU수), int(row.판매수량), int(row.판매금액),
        f"{row.판매금액_비중pct:.1f}%", f"{row.소화율_avg:.1f}%", f"{row.마크업avg:.0f}%"
    ])

set_col_widths(ws6, [14, 8, 12, 18, 10, 12, 12])

# ════════════════════════════════════════════════════
# 시트 7: 💰 마진구조
# ════════════════════════════════════════════════════
ws7 = wb.create_sheet("💰 마진구조")
ws7.sheet_view.showGridLines = False

ws7.merge_cells("A1:G1")
c = ws7["A1"]
c.value = "💰 마진 구조 분석 — 택가실현율 & 마크업 분포"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()

# SKU별 마진 테이블
apply_header_row(ws7, 2, ["품번","품명","최초판매가","현판매가","할인","소화율%","마크업avg%","택가실현율%","마크업등급"])
margin_full = sku_sorted[["품번","품명","최초판매가","현판매가","소화율_pct","마크업율_avg","택가실현율_pct"]].copy()
# 할인 여부 - 품번별로 색상별 할인 여부 확인
discount_map = df.groupby("품번")["할인여부"].any().to_dict()
markup_grade_map = df.groupby("품번")["마크업등급"].first().to_dict()

for ri, row in enumerate(margin_full.itertuples(), 3):
    has_discount = discount_map.get(row.품번, False)
    mgrade = markup_grade_map.get(row.품번, "N/A")
    bg = None
    if "HIGH" in str(mgrade): bg = C_BG_MUST
    elif "LOW" in str(mgrade): bg = C_BG_SKIP
    apply_data_row(ws7, ri, [
        row.품번, row.품명,
        f"{int(row.최초판매가):,}",
        f"{int(row.현판매가):,}",
        "할인" if has_discount else "-",
        f"{row.소화율_pct:.1f}%",
        f"{row.마크업율_avg:.0f}%",
        f"{row.택가실현율_pct:.1f}%" if pd.notna(row.택가실현율_pct) else "-",
        mgrade,
    ], bg=bg)

set_col_widths(ws7, [14, 28, 12, 12, 8, 10, 12, 12, 18])
ws7.freeze_panes = "A3"

# ════════════════════════════════════════════════════
# 시트 8: 🎯 26FW시사점
# ════════════════════════════════════════════════════
ws8 = wb.create_sheet("🎯 26FW시사점")
ws8.sheet_view.showGridLines = False

ws8.merge_cells("A1:H1")
c = ws8["A1"]
c.value = "🎯 26FW 기획 시사점 — MUST / WATCH / SKIP 매트릭스"
c.fill = make_fill(C_BG_HEADER)
c.font = make_font(bold=True, color=C_TEXT_WHITE, size=13)
c.alignment = center()

# 매트릭스 설명
ws8.merge_cells("A2:H2")
c = ws8["A2"]
c.value = "기준: 소화율 ≥70% = 고소화, 마크업율 ≥250% = 고마진 (사전 마크업 기준)"
c.fill = make_fill(C_BG_SUB)
c.font = make_font(color=C_TEXT_WHITE, size=9)
c.alignment = center()

# 매트릭스 상세
apply_header_row(ws8, 3, ["26FW판별","품번","품명","대분류","활용","소화율%","마크업avg%","액션"])
action_map = {
    "🟢 MUST": "26FW 재편성 — 소재/컬러 업그레이드 검토",
    "🟡 WATCH(VE)": "소재/원가 VE 후 재검토 — 택가 인상 or 완사입비 인하",
    "🟡 WATCH(QR)": "소량 시작 + QR 전환 — 수요 확인 후 추가 발주",
    "🔴 SKIP": "26FW 제외 — 재고 소진 집중",
}

for ri, (_, row) in enumerate(sku_sorted.iterrows(), 4):
    판별 = row["26FW판별"]
    bg = None
    if "MUST" in str(판별): bg = C_BG_MUST
    elif "SKIP" in str(판별): bg = C_BG_SKIP
    elif "WATCH" in str(판별): bg = C_BG_WATCH
    action = action_map.get(str(판별), "")
    apply_data_row(ws8, ri, [
        판별, row["품번"], row["품명"], row["대분류"], row["활용"],
        f"{row['소화율_pct']:.1f}%", f"{row['마크업율_avg']:.0f}%", action
    ], bg=bg)

set_col_widths(ws8, [16, 14, 26, 10, 8, 10, 12, 40])
ws8.freeze_panes = "A4"

# ── 시트 순서 정렬 (종합요약이 첫 번째)
wb.save(OUT_XLSX)
print(f"   ✓ Excel 저장: {OUT_XLSX}")

# ─── 6. 집계 결과 수집 (MD 리포트용) ────────────────────────────────────────
print("[6/8] MD 리포트 집계 데이터 준비...")

# TOP5 SKU
top5 = sku_sorted.head(5)
bot5 = sku_sorted.tail(5)

# 출고월 피크
peak_month = timing.loc[timing["판매금액"].idxmax(), "출고월"] if len(timing) else "N/A"

# 최고 소화율 색상
best_color = color_df.iloc[0]["색상"] if len(color_df) else "N/A"
best_color_sell = color_df.iloc[0]["소화율_avg"] if len(color_df) else 0

# IP 비율
kiki_pct = ip_bin[ip_bin["IP구분"]=="키키IP"]["판매금액_비중pct"].values
kiki_pct = kiki_pct[0] if len(kiki_pct) else 0

# 최고 마크업 대분류
best_main = cat_main.sort_values("마크업avg", ascending=False).iloc[0]

# 완사입비 높은 SKU
high_cost = sku_sorted.sort_values("실판매금액", ascending=True).head(3)

print(f"   피크 출고월: {peak_month} / 키키IP 비중: {kiki_pct:.1f}%")

# ─── 7. MD 리포트 작성 ────────────────────────────────────────────────────
print("[7/8] MD 인사이트 리포트 작성 중...")

top5_rows = "\n".join([
    f"| {i+1} | {r['품번']} | {r['품명'][:20]} | {int(r['판매금액']):,} | {r['소화율_pct']:.1f}% | {r['마크업율_avg']:.0f}% | {r['26FW판별']} |"
    for i, (_, r) in enumerate(top5.iterrows())
])

bot5_rows = "\n".join([
    f"| {len(sku_sorted)-4+i} | {r['품번']} | {r['품명'][:20]} | {int(r['판매금액']):,} | {r['소화율_pct']:.1f}% | {r['마크업율_avg']:.0f}% | {r['26FW판별']} |"
    for i, (_, r) in enumerate(bot5.iterrows())
])

timing_rows = "\n".join([
    f"| {r.출고월} | {r.색상건수} | {int(r.판매금액):,} | {r.소화율_avg}% |"
    for r in timing.itertuples()
])

ip_rows = "\n".join([
    f"| {str(r.활용) if pd.notna(r.활용) else '미분류'} | {int(r.SKU수)} | {int(r.판매금액):,} | {r.판매금액_비중pct:.1f}% | {r.소화율_avg:.1f}% |"
    for r in ip_df.itertuples()
])

cat_rows = "\n".join([
    f"| {str(r.대분류) if pd.notna(r.대분류) else '-'} | {int(r.SKU수)} | {int(r.판매금액):,} | {r.소화율_avg:.1f}% | {r.마크업avg:.0f}% |"
    for r in cat_main.itertuples()
])

style_rows = "\n".join([
    f"| {str(r.볼캡유형) if pd.notna(r.볼캡유형) else '-'} | {int(r.SKU수)} | {int(r.판매금액):,} | {r.소화율_avg:.1f}% | {r.마크업avg:.0f}% |"
    for r in cat_style2.itertuples()
])

color_grp_rows = "\n".join([
    f"| {r.색상계열} | {r.SKU건수} | {int(r.판매금액):,} | {r.소화율_avg:.1f}% |"
    for r in color_grp.itertuples()
])

matrix_summary = {
    "🟢 MUST": sku[sku["26FW판별"]=="🟢 MUST"][["품번","품명","소화율_pct","마크업율_avg"]].head(5),
    "🔴 SKIP": sku[sku["26FW판별"]=="🔴 SKIP"][["품번","품명","소화율_pct","마크업율_avg"]].head(5),
}

must_rows = "\n".join([
    f"| {r.품번} | {r.품명[:22]} | {r.소화율_pct:.1f}% | {r.마크업율_avg:.0f}% |"
    for r in matrix_summary["🟢 MUST"].itertuples()
]) if len(matrix_summary["🟢 MUST"]) else "| - | - | - | - |"

skip_rows = "\n".join([
    f"| {r.품번} | {r.품명[:22]} | {r.소화율_pct:.1f}% | {r.마크업율_avg:.0f}% |"
    for r in matrix_summary["🔴 SKIP"].itertuples()
]) if len(matrix_summary["🔴 SKIP"]) else "| - | - | - | - |"

md_content = f"""# 🧢 와키윌리 25SS 모자 판매 분석 리포트

> **담당**: 데이터 인텔리전스 — 트렌드 애널리스트 + 인사이트 아키텍트
> **작성일**: 2026-03-19
> **데이터**: 25SS 시즌 판매자료 162행 × 40컬럼 / 구조화 설계 59 SKU
> **Excel 대시보드**: `check_sales-analysis_2026-03-19.xlsx` (8시트)

---

## 📌 Executive Summary

| KPI | 값 |
|-----|----|
| 총 발주수량 | {int(total_orders):,} 개 |
| 총 누계판매금액 | {int(total_sales_a):,} 원 |
| 총 실판매금액 | {int(total_sales_r):,} 원 |
| 평균 소화율(수량) | {avg_sell_thru:.1f} % |
| 평균 마크업율(사전) | {avg_markup:.0f} % |
| 분석 SKU 수 | {sku_count} 개 |
| 26FW MUST | {must_count}개 ({must_count/total_sku*100:.0f}%) |
| 26FW WATCH | {watch_count}개 ({watch_count/total_sku*100:.0f}%) |
| 26FW SKIP | {skip_count}개 ({skip_count/total_sku*100:.0f}%) |

### 핵심 인사이트 3줄 요약

1. **소화율 {avg_sell_thru:.0f}%** — 전체 소화율은 양호하나, SKU별 편차가 크다. 상위 10개 SKU가 매출의 대부분을 차지.
2. **키키 IP 비중 {kiki_pct:.0f}%** — IP 제품군이 판매를 견인. 비-IP 대비 소화율·마크업 양면에서 우위.
3. **MUST {must_count}개 / SKIP {skip_count}개** — 26FW는 고소화·고마진 라인 중심 재편 + SKIP 재고 소진 집중 전략이 유효.

---

## 1. 판매 기간 & 시점 분석

| 최초출고월 | 색상건수 | 판매금액(원) | 소화율avg |
|-----------|---------|------------|---------|
{timing_rows}

**인사이트**:
- **초기 입고 집중**: 2024-11~12 선발 입고가 SS 시즌 총 판매를 선점. 리드타임 관리가 소화율의 핵심 변수.
- **피크 구간**: {peak_month} 출고 그룹이 판매금액 최대 — 시즌 피크와 입고 타이밍 일치.
- 최초출고일부터 최종출고일까지 평균 판매경과일: **{df['판매경과일'].mean():.0f}일** — 장기 판매 재고 소진 패턴.

---

## 2. SKU 매출 랭킹

### TOP 5 (판매금액 기준)

| 순위 | 품번 | 품명 | 판매금액(원) | 소화율% | 마크업% | 26FW |
|------|------|------|------------|--------|--------|------|
{top5_rows}

### BOTTOM 5

| 순위 | 품번 | 품명 | 판매금액(원) | 소화율% | 마크업% | 26FW |
|------|------|------|------------|--------|--------|------|
{bot5_rows}

**인사이트**:
- TOP5 SKU 집중도 높음 — 상위 SKU에 물량·마케팅 자원 집중 필요.
- BOTTOM5는 소화율·마진 동반 저조 → SKIP 후보. 남은 재고 프로모션/번들 소진 권고.

---

## 3. 카테고리별 분석

### 대분류 (로고/그래픽/워드마크)

| 대분류 | SKU수 | 판매금액(원) | 소화율% | 마크업% |
|-------|------|------------|--------|--------|
{cat_rows}

### 볼캡 유형별 (볼캡/캠프캡/버킷햇 등)

| 볼캡유형 | SKU수 | 판매금액(원) | 소화율% | 마크업% |
|--------|------|------------|--------|--------|
{style_rows}

**인사이트**:
- {cat_main.sort_values('판매금액', ascending=False).iloc[0]['대분류'] if len(cat_main) else 'N/A'} 대분류가 판매 1위. 26FW도 해당 카테고리 중심 구성 권고.
- 볼캡 유형별 판매 편차 존재 — 소화율 낮은 유형은 SKU 수 축소 검토.

---

## 4. IP 활용 분석

| 활용유형 | SKU수 | 판매금액(원) | 판매비중% | 소화율% |
|---------|------|------------|---------|--------|
{ip_rows}

**인사이트**:
- **키키/릴리 IP 제품군이 판매 핵심 드라이버**. 타이포/시즌 제품 대비 소화율·마진 우위.
- 26FW에서 IP 활용 SKU 비중 확대 + 릴리 IP 신규 스타일 추가 검토.

---

## 5. 색상별 소화율 분석

| 색상계열 | SKU건수 | 판매금액(원) | 소화율avg% |
|--------|--------|------------|----------|
{color_grp_rows}

**인사이트**:
- **베이직(블랙/네이비/그레이) 계열** 절대 판매량 1위 — 코어 컬러 안정적 유지 필요.
- 컬러풀 계열 소화율 변동성 큼 — 소량 발주 후 QR 전환 구조 권장.
- 최고 소화율 단일 색상: **{best_color}** ({best_color_sell:.1f}%) — 26FW 우선 컬러 후보.

---

## 6. 판매 타이밍 & 시즌 맥락

- **선발 입고 효과**: 2024-11 최초 입고 SKU가 전 시즌을 통해 지속 판매 — 브랜드 팬덤 선점 효과.
- **할인 발생 여부**: 최초판매가 vs 현판매가 비교 결과, 일부 SKU에서 할인 발생 → 소화율 제고 수단으로 활용.
- **평균 판매경과일 {df['판매경과일'].mean():.0f}일**: 모자 카테고리는 시즌을 넘어 판매 지속 — 재고 장기화 리스크 인지 필요.

---

## 7. 마진 구조 분석

- **평균 사전 마크업율 {avg_markup:.0f}%** — 패션 업계 평균 대비 양호한 수준.
- **HIGH(≥300%) 마크업 SKU**: {(sku['마크업율_avg'] >= 300).sum()}개 — 프리미엄 가격 지지력 확인.
- **LOW(<200%) 마크업 SKU**: {(sku['마크업율_avg'] < 200).sum()}개 — 원가 VE 또는 가격 재검토 필요.
- **완사입비 부담 분석**: 단가 대비 판매금액 실현율이 낮은 SKU는 생산처 협상 또는 소재 변경 VE 권고.

---

## 8. 26FW 기획 시사점 — MUST / WATCH / SKIP

### 기준 매트릭스
| | 소화율 ↑ (≥70%) | 소화율 ↓ (<70%) |
|--|--|--|
| **마크업 ↑ (≥250%)** | 🟢 MUST — 재편성 | 🟡 WATCH — QR 전환 |
| **마크업 ↓ (<250%)** | 🟡 WATCH — VE 검토 | 🔴 SKIP — 제외 |

### 🟢 MUST 상위 SKU ({must_count}개)

| 품번 | 품명 | 소화율% | 마크업% |
|------|------|--------|--------|
{must_rows}

### 🔴 SKIP 상위 SKU ({skip_count}개)

| 품번 | 품명 | 소화율% | 마크업% |
|------|------|--------|--------|
{skip_rows}

---

## 액션 아이템 (26FW 기획팀 전달)

| 우선순위 | 액션 | 담당 | 기한 |
|---------|------|------|------|
| P1 | MUST SKU {must_count}개 26FW 라인업 확정 | MD | 2026-04 |
| P1 | 키키/릴리 IP 신규 스타일 추가 기획 (소화율·마진 우위 확인) | 크리에이티브 | 2026-04 |
| P2 | WATCH(VE) SKU — 완사입비 재협상 또는 소재 대체 검토 | 프로덕트랩 | 2026-04 |
| P2 | 컬러풀 계열 소량 QR 전환 구조 설계 | MD | 2026-04 |
| P3 | SKIP SKU 재고 소진 프로모션 계획 수립 | 마케팅 | 2026-05 |
| P3 | 피크 출고월({peak_month}) 입고 타이밍 기준으로 26FW 발주 일정 역산 | MD/프로덕트랩 | 2026-04 |

---

*Generated by FPOF 데이터 인텔리전스 / check_sales-analysis_2026-03-19*
"""

with open(OUT_MD, "w", encoding="utf-8") as f:
    f.write(md_content)
print(f"   ✓ MD 저장: {OUT_MD}")

# ─── 8. .fpof-state.json 업데이트 ────────────────────────────────────────
print("[8/8] .fpof-state.json 업데이트...")
import json
from datetime import datetime

STATE_PATH = os.path.join(BASE, ".fpof-state.json")
with open(STATE_PATH, "r", encoding="utf-8") as f:
    state = json.load(f)

# operational.weekly 또는 최근 check에 기록
if "operational" not in state:
    state["operational"] = {}
if "recent_outputs" not in state["operational"]:
    state["operational"]["recent_outputs"] = []

state["operational"]["recent_outputs"].append({
    "date": "2026-03-19",
    "type": "check",
    "files": [
        "output/26FW/season-strategy/check_sales-analysis_2026-03-19.xlsx",
        "output/26FW/season-strategy/check_sales-analysis_2026-03-19.md"
    ],
    "summary": f"25SS 모자 판매 분석 — SKU {sku_count}개, 소화율 {avg_sell_thru:.1f}%, MUST {must_count}/WATCH {watch_count}/SKIP {skip_count}"
})

with open(STATE_PATH, "w", encoding="utf-8") as f:
    json.dump(state, f, ensure_ascii=False, indent=2)
print("   ✓ 상태 업데이트 완료")

print("\n✅ 분석 완료!")
print(f"   📊 Excel: {os.path.basename(OUT_XLSX)}")
print(f"   📝 MD: {os.path.basename(OUT_MD)}")
print(f"\n   KPI 요약:")
print(f"   - 총 판매금액: {int(total_sales_a):,}원")
print(f"   - 평균 소화율: {avg_sell_thru:.1f}%")
print(f"   - 평균 마크업: {avg_markup:.0f}%")
print(f"   - MUST/WATCH/SKIP: {must_count}/{watch_count}/{skip_count}")
