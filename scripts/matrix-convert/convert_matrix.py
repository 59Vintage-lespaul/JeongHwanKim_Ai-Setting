"""
ACC 26SS MATRIX 양식 변환기
기준 파일(커버낫)의 컬럼 구조에 맞춰 LEE, WW 데이터를 변환하여
3개 시트로 구성된 새 Excel 파일을 생성합니다.

실행:
  py -3.12 scripts/matrix-convert/convert_matrix.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy
import os
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# ─── 파일 경로 ───────────────────────────────────────────────────────────────
BASE_DIR  = r"C:\Users\bcave\Desktop\MD플랜 제작 중"
REF_FILE  = os.path.join(BASE_DIR, "(ACC) 26SS MATRIX.xlsx")
LEE_FILE  = os.path.join(BASE_DIR, "LEE_26SS 기획현황_0329.xlsx")
WW_FILE   = os.path.join(BASE_DIR, "와키윌리_26SS 발주리스트_데이터 APP+ACC.xlsx")

today = datetime.now().strftime("%Y-%m-%d")
OUT_FILE  = os.path.join(BASE_DIR, f"26SS ACC MATRIX_통합_{today}.xlsx")

# ─── 기준 컬럼 인덱스 (0-based, row4 헤더 기준) ─────────────────────────────
R = dict(
    co_item_no   = 0,   # C/O품번
    spw          = 1,   # SPW
    stock_qty    = 2,   # 재고수량
    hold_week    = 3,   # 보유주
    season_type  = 4,   # 시즌구분
    category     = 5,   # 대분류
    new_carry    = 6,   # 신규/캐리오버
    main_type    = 7,   # 메인/리핏/스팟
    gender       = 8,   # 성별
    concept      = 9,   # 컨셉
    sty          = 10,  # STY
    sku          = 11,  # SKU
    sty_bta      = 12,  # STY BTA
    col_bta      = 13,  # COL BTA
    item_no      = 14,  # 품번
    style_no     = 15,  # 스타일넘버
    img          = 16,  # 제품이미지
    color_code   = 17,  # 컬러 코드
    color        = 18,  # 컬러
    brand        = 19,  # 브랜드
    year         = 20,  # 연도
    season       = 21,  # 시즌
    item_code    = 22,  # 아이템
    number       = 23,  # 번호
    product_name = 24,  # 제품명
    dom_total    = 25,  # 국내 TOTAL
    main_qty     = 26,  # 메인
    main_global  = 27,  # 메인(글로벌)
    order2       = 28,  # 2차오더
    order3       = 29,  # 3차오더
    order4       = 30,  # 4차오더
    order5       = 31,  # 5차오더
    intl_total   = 32,  # 해외 TOTAL
    taiwan       = 33,  # 대만
    china        = 34,  # 중국
    global_      = 35,  # 글로벌
    total_qty    = 36,  # 국내+해외수량
    sale_amt     = 37,  # 판매 금액
    cost_amt     = 38,  # 원가 금액
    target_cost  = 39,  # 타겟원가(V-)
    target_rate  = 40,  # 타겟원가율
    final_cost_v = 41,  # 확정원가(V+)
    final_price  = 42,  # 최종 판매가
    confirm_cost = 43,  # 확정원가(V-)
    confirm_costv= 44,  # 확정원가(V+)
    final_price2 = 45,  # 최종 판매가
    cost_rate    = 46,  # 원가율
    vendor       = 47,  # 생산업체
    country      = 48,  # 생산국
    order_rank   = 49,  # 차수(투입순위)
    order_date   = 50,  # 발주일자
    target_date  = 51,  # 목표납기
    sourcing_date= 52,  # 소싱 납기
    inbound_date = 53,  # 입고일(차이)
    delivery     = 54,  # 발매딜리버리
    online_rank  = 55,  # 온라인 차수
    online_drop  = 56,  # 온라인 DROP
    reorder_date = 58,  # 리오더 목표납기
    inbound2     = 59,  # 입고
    size_fit     = 60,  # 사이즈/핏
    material     = 61,  # 소재
    note         = 62,  # 비고
    dsn_status   = 63,  # DSN 진행현황
    key_point    = 64,  # KEY POINT
)
N_COLS = 65  # 총 컬럼 수


def make_row(n=N_COLS):
    return [None] * n


def parse_item_no(item_no):
    """품번(10자리)에서 브랜드·연도·시즌·아이템·번호 추출"""
    s = str(item_no).strip() if item_no else ""
    if len(s) < 10:
        return {}, s
    return {
        "brand":      s[0:2],   # CO / WA / LE
        "year":       s[2:4],   # 26
        "season":     s[4:6],   # 01 = SS, 02 = FW 등
        "item_code":  s[6:8],   # BP, BG, SC, WT ...
        "number":     s[8:10],  # 01, 02 ...
    }, s[0:2]  # (dict, 브랜드코드)


def si(v):
    """safe int"""
    try:
        return int(float(str(v))) if v is not None and v != '' else 0
    except (ValueError, TypeError):
        return 0


def sf(v):
    """safe float → None if error"""
    try:
        return float(str(v)) if v is not None and v != '' else None
    except (ValueError, TypeError):
        return None


# ─── REF 시트 복사 ───────────────────────────────────────────────────────────
def copy_ref_sheet(src_ws, dst_ws):
    """기준 시트를 그대로 복사 (값 + 스타일)"""
    for row in src_ws.iter_rows():
        for cell in row:
            nc = dst_ws.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                nc.font      = copy.copy(cell.font)
                nc.fill      = copy.copy(cell.fill)
                nc.border    = copy.copy(cell.border)
                nc.alignment = copy.copy(cell.alignment)
                nc.number_format = cell.number_format

    # 열 너비 복사
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
    # 행 높이 복사
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
    # 병합 셀 복사
    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))


# ─── 헤더 스타일 적용 ────────────────────────────────────────────────────────
def write_header(ws, headers, row=1):
    """헤더 행 작성 (회색 배경 + 굵은 글씨)"""
    fill = PatternFill("solid", fgColor="D9D9D9")
    font = Font(bold=True, name="맑은 고딕", size=9)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border


def set_data_cell(ws, row_idx, col_idx, value):
    cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)  # openpyxl은 1-based
    cell.font = Font(name="맑은 고딕", size=9)
    cell.alignment = Alignment(vertical="center")


# ─── LEE 변환 ────────────────────────────────────────────────────────────────
def build_lee_rows():
    wb = openpyxl.load_workbook(LEE_FILE, read_only=True, data_only=True)
    ws = wb["기획현황"]

    # 2레벨 헤더: row5(기본) + row6(서브) — 데이터는 row8부터
    rows = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        # 카테고리[3] == 'ACC' 인 행만
        if not row[3] or str(row[3]).strip().upper() != "ACC":
            continue
        # 빈 행 스킵
        if not row[7]:  # 품번 없으면 스킵
            continue

        r = make_row()
        item_no = str(row[7]).strip() if row[7] else ""
        meta, brand_code = parse_item_no(item_no)

        r[R["item_no"]]      = item_no                       # 품번
        r[R["brand"]]        = brand_code or "LE"            # 브랜드
        r[R["year"]]         = meta.get("year", "")          # 연도
        r[R["season"]]       = meta.get("season", "")        # 시즌
        r[R["item_code"]]    = meta.get("item_code", "")     # 아이템
        r[R["number"]]       = meta.get("number", "")        # 번호
        r[R["product_name"]] = row[8]                        # 품명 → 제품명
        r[R["category"]]     = row[4]                        # 복종 → 대분류
        r[R["gender"]]       = row[64]                       # 성별
        # BTA: row[11] → STY BTA + COL BTA 동일하게 사용
        bta_val = row[11]
        r[R["sty_bta"]]      = bta_val
        r[R["col_bta"]]      = bta_val
        r[R["main_type"]]    = row[10]                       # MSR → 메인/리핏/스팟
        r[R["dom_total"]]    = si(row[40])                   # 총기획량 → 국내 TOTAL
        r[R["main_qty"]]     = si(row[95])                   # 발주수량 → 메인
        r[R["final_price"]]  = sf(row[58])                   # TAG가 → 최종 판매가
        r[R["final_price2"]] = sf(row[58])
        r[R["cost_rate"]]    = sf(row[62])                   # 원가율
        r[R["vendor"]]       = row[45]                       # 생산처 → 생산업체
        r[R["country"]]      = row[46]                       # 생산지 → 생산국
        r[R["target_date"]]  = row[12]                       # 기획납기 → 목표납기
        r[R["delivery"]]     = row[13]                       # 딜리버리OFF → 발매딜리버리
        r[R["material"]]     = row[47]                       # 원단 → 소재
        r[R["note"]]         = row[97]                       # 비고

        # 컬러 코드: 발주 컬러가 col19~28에 있음 — 첫 번째 비어있지 않은 값
        for c_idx in range(19, 29):
            if row[c_idx]:
                r[R["color_code"]] = row[c_idx]
                break

        rows.append(r)

    wb.close()
    return rows


# ─── WW 변환 ─────────────────────────────────────────────────────────────────
def build_ww_rows():
    wb = openpyxl.load_workbook(WW_FILE, read_only=True, data_only=True)
    ws = wb["상품기획안_종합"]

    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        # 복종[3]에 'ACC' 포함
        if not row[3] or 'ACC' not in str(row[3]).upper():
            continue
        # 확정품번 없으면 스킵
        if not row[13]:
            continue
        # 컬러코드 없거나 TTL=0인 빈 행 스킵 (예약/자리 행)
        if not row[27] or si(row[28]) == 0:
            continue

        r = make_row()
        item_no = str(row[13]).strip() if row[13] else ""
        meta, brand_code = parse_item_no(item_no)

        r[R["item_no"]]      = item_no                        # 확정품번 → 품번
        r[R["brand"]]        = brand_code or "WA"             # 브랜드
        r[R["year"]]         = meta.get("year", "")           # 연도
        r[R["season"]]       = meta.get("season", "")         # 시즌
        r[R["item_code"]]    = meta.get("item_code", "")      # 아이템
        r[R["number"]]       = meta.get("number", "")         # 번호
        r[R["product_name"]] = row[16]                        # 상품명 → 제품명
        # 복종에서 '/용품' 제거
        cat_raw = str(row[3]).replace("ACC/용품", "ACC/용품").strip()
        r[R["category"]]     = cat_raw                        # 복종 → 대분류
        r[R["main_type"]]    = row[4]                         # 발주구분 → 메인/리핏/스팟
        r[R["concept"]]      = row[9]                         # 라인 → 컨셉
        r[R["color_code"]]   = row[27]                        # 컬러코드
        r[R["dom_total"]]    = si(row[28])                    # TTL → 국내 TOTAL
        r[R["main_qty"]]     = si(row[29])                    # Main → 메인
        r[R["final_price"]]  = sf(row[15])                    # 확정소비자가 → 최종 판매가
        r[R["final_price2"]] = sf(row[15])
        r[R["confirm_cost"]] = sf(row[45])                    # 확정원가(소싱) → 확정원가(V-)
        r[R["cost_rate"]]    = sf(row[49])                    # 발주원가율 → 원가율
        r[R["vendor"]]       = row[25]                        # 생산업체(확정) → 생산업체
        r[R["country"]]      = row[26]                        # 생산국가(확정) → 생산국
        r[R["target_date"]]  = row[20]                        # 납기일 → 목표납기
        r[R["delivery"]]     = row[21]                        # 딜리버리 → 발매딜리버리
        r[R["note"]]         = row[94]                        # 비고

        rows.append(r)

    wb.close()
    return rows


# ─── 데이터 시트 작성 ────────────────────────────────────────────────────────
HEADER_LABELS = [
    "C/O품번", "SPW", "재고수량", "보유주", "시즌구분", "대분류",
    "신규/캐리오버", "메인/리핏/스팟", "성별", "컨셉", "STY", "SKU",
    "STY BTA", "COL BTA", "품번", "스타일넘버", "제품이미지", "컬러 코드",
    "컬러", "브랜드", "연도", "시즌", "아이템", "번호", "제품명",
    "국내 TOTAL", "메인", "메인(글로벌오더)", "2차오더", "3차오더",
    "4차오더", "5차오더", "해외 TOTAL", "대만", "중국", "글로벌",
    "국내+해외수량", "판매 금액", "원가 금액", "타겟원가(V-)", "타겟원가율",
    "확정원가(V+)", "최종 판매가", "확정원가(V-)", "확정원가(V+)", "최종 판매가",
    "원가율", "생산업체", "생산국", "차수(투입순위)", "발주일자", "목표납기",
    "소싱 납기", "입고일(차이)", "발매딜리버리", "온라인 차수", "온라인 DROP",
    "", "리오더 목표납기", "입고", "사이즈/핏", "소재", "비고",
    "DSN 진행현황", "KEY POINT",
]


def write_data_sheet(ws, brand_rows, sheet_title):
    ws.freeze_panes = "A2"
    write_header(ws, HEADER_LABELS, row=1)

    # 열 너비 기본값
    for col_i in range(1, N_COLS + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 12
    ws.column_dimensions[get_column_letter(R["product_name"] + 1)].width = 28
    ws.column_dimensions[get_column_letter(R["item_no"] + 1)].width = 16
    ws.column_dimensions[get_column_letter(R["style_no"] + 1)].width = 18
    ws.column_dimensions[get_column_letter(R["vendor"] + 1)].width = 14
    ws.row_dimensions[1].height = 30

    data_font = Font(name="맑은 고딕", size=9)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, row_data in enumerate(brand_rows, start=2):
        ws.row_dimensions[row_i].height = 18
        for col_i, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            # 날짜 포맷
            if isinstance(val, datetime) or (hasattr(val, '__class__') and 'datetime' in type(val).__name__):
                cell.number_format = "YYYY-MM-DD"
            # 숫자 포맷
            if col_i - 1 in (R["final_price"], R["final_price2"], R["confirm_cost"], R["target_cost"]):
                cell.number_format = "#,##0"
            if col_i - 1 in (R["cost_rate"], R["target_rate"]):
                cell.number_format = "0.0%"
            if col_i - 1 in (R["dom_total"], R["main_qty"], R["total_qty"], R["intl_total"]):
                cell.number_format = "#,##0"

    print(f"  → {sheet_title}: {len(brand_rows)}행 작성 완료")


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    print("=== ACC 26SS MATRIX 변환 시작 ===\n")

    # 1) 기준 시트 복사 (read only로 열어서 복사)
    print("[1/4] 기준(커버낫) 시트 복사 중...")
    ref_wb = openpyxl.load_workbook(REF_FILE, data_only=True)
    ref_ws = ref_wb["26SS ACC"]

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # 기본 시트 제거

    ws_ref = out_wb.create_sheet("기준(커버낫)")
    copy_ref_sheet(ref_ws, ws_ref)
    ref_wb.close()
    print("  → 기준(커버낫) 복사 완료")

    # 2) LEE 변환
    print("\n[2/4] LEE 데이터 변환 중...")
    lee_rows = build_lee_rows()
    ws_lee = out_wb.create_sheet("LEE")
    write_data_sheet(ws_lee, lee_rows, "LEE")

    # 3) WW 변환
    print("\n[3/4] 와키윌리 데이터 변환 중...")
    ww_rows = build_ww_rows()
    ws_ww = out_wb.create_sheet("와키윌리")
    write_data_sheet(ws_ww, ww_rows, "와키윌리")

    # 4) 저장
    print(f"\n[4/4] 파일 저장 중: {OUT_FILE}")
    out_wb.save(OUT_FILE)
    print(f"\n[완료] 저장 위치: {OUT_FILE}")
    print(f"   - 기준(커버낫): 원본 그대로")
    print(f"   - LEE:         {len(lee_rows)}행")
    print(f"   - 와키윌리:    {len(ww_rows)}행")


if __name__ == "__main__":
    main()
