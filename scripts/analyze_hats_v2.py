#!/usr/bin/env python3
"""
25SS 모자 판매 분석 v2 — 월별 판매추이 통합
입력: 판매자료(162행) + 구조화설계(59SKU) + 월별판매추이(110행)
출력: check_sales-analysis_2026-03-19.xlsx (10시트) + check_sales-analysis_2026-03-19.html
"""

import pandas as pd
import numpy as np
import json
import warnings
import sys
import os

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = open(sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1)

warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 경로 ────────────────────────────────────────────────────────────────────
BASE       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SALES_PATH = os.path.join(BASE, "output/26FW/season-strategy/품번종합집계표_모자 판매자료.xlsx")
STRUCT_PATH= os.path.join(BASE, "output/26FW/season-strategy/모자 구조화 설계 최종.xlsx")
MONTHLY_PATH=os.path.join(BASE, "output/26FW/season-strategy/월별 판매추이분석_모자월별판매.xlsx")
OUT_XLSX   = os.path.join(BASE, "output/26FW/season-strategy/check_sales-analysis_v2_2026-03-19.xlsx")
OUT_HTML   = os.path.join(BASE, "output/26FW/season-strategy/check_sales-analysis_2026-03-19.html")

MONTHS = [
    "2025-01","2025-02","2025-03","2025-04","2025-05","2025-06",
    "2025-07","2025-08","2025-09","2025-10","2025-11","2025-12",
    "2026-01","2026-02","2026-03"
]
# 월별파일: 기간수량 컬럼 인덱스 (row1 header 기준)
MONTHLY_QTY_IDX = {MONTHS[i]: 13 + i * 3 for i in range(15)}

# ─── 스타일 팔레트 ────────────────────────────────────────────────────────────
C_HEADER   = "1A1A2E"
C_ACCENT   = "E94560"
C_SUB      = "16213E"
C_LIGHT    = "F0F4FF"
C_MUST     = "D4EDDA"
C_SKIP     = "F8D7DA"
C_WATCH    = "FFF3CD"
C_WHITE    = "FFFFFF"
C_DARK     = "1A1A2E"
C_BORDER   = "CCCCCC"

def fill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, color=C_DARK, size=10):
    return Font(name="맑은 고딕", bold=bold, color=color, size=size)
def brd():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)
def ctr():   return Alignment(horizontal="center", vertical="center", wrap_text=True)
def set_widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr_row(ws, r, vals, bg=C_HEADER, fg=C_WHITE):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = fill(bg); c.font = fnt(True, fg); c.alignment = ctr(); c.border = brd()

def data_row(ws, r, vals, bg=None):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        if bg: c.fill = fill(bg)
        c.font = fnt(); c.alignment = ctr(); c.border = brd()

def row_bg(판별):
    if "MUST" in str(판별): return C_MUST
    if "SKIP" in str(판별): return C_SKIP
    if "WATCH" in str(판별): return C_WATCH
    return None

def heat_color(val, mx):
    """0→white, mx→accent(E94560)"""
    if mx == 0 or pd.isna(val) or val == 0:
        return "FAFBFF"
    r = val / mx
    R = int(250 + (233-250)*r)
    G = int(251 + (69-251)*r)
    B = int(255 + (96-255)*r)
    return f"{R:02X}{G:02X}{B:02X}"

# ════════════════════════════════════════════════════════════════════════════
# 1. 데이터 로드 & 정제
# ════════════════════════════════════════════════════════════════════════════
print("[1/9] 데이터 로드...")

sales_raw = pd.read_excel(SALES_PATH, header=1)
struct_raw = pd.read_excel(STRUCT_PATH, sheet_name="품번별 세부구분", header=0)
monthly_raw = pd.read_excel(MONTHLY_PATH, header=1)

# ── 판매자료 컬럼 표준화
SCOLS = {
    0:"품번", 1:"품명", 2:"색상", 3:"발주배수_최", 4:"발주금액",
    6:"최초판매가", 7:"현판매가", 8:"소화율_수량pct", 9:"소화율_금액pct",
    10:"시즌", 11:"기획구분", 14:"최초입고일", 15:"최종입고일",
    16:"최초출고일", 17:"최종출고일", 20:"발주_수량", 21:"발주_금액",
    22:"발주_최초판매금액", 23:"최초출고월", 24:"발주_현판매금액",
    25:"기획구분2", 26:"누계판매_수량", 27:"누계판매_금액",
    31:"누계판매_실판매금액", 34:"완사입비_단가", 35:"판매이익_금액",
    37:"판매이익율_사전V", 38:"판매이익율_사후_수량", 39:"판매이익율_사후_금액",
}
sales = sales_raw.copy()
sales.columns = [SCOLS.get(i, f"_c{i}") for i in range(len(sales.columns))]
sales = sales[sales["품번"].astype(str).str.startswith("WA")].copy()

# ── 구조화 컬럼 표준화
KCOLS = {0:"품번", 1:"품명_구조", 2:"대분류", 3:"활용",
          4:"로고사이즈", 5:"표현방식", 6:"스타일분류", 7:"볼캡유형",
          8:"브림", 9:"크라운", 10:"패널수", 11:"소재", 12:"소재코드"}
struct = struct_raw.copy()
struct.columns = [KCOLS.get(i, f"_k{i}") for i in range(len(struct.columns))]

# ── 월별파일: 25SS 모자 필터 (WA25 + CA/HT/BN)
mf = monthly_raw.copy()
# 컬럼 이름 (위치 기반)
MCOLS_BASE = {0:"품번", 1:"색상", 2:"최초판매가", 3:"현판매가",
              4:"최초입고일", 5:"최초출고일", 6:"발주", 7:"입고",
              8:"출고", 9:"판매", 10:"판매율", 11:"재고"}
for i, m in enumerate(MONTHS):
    mf.rename(columns={mf.columns[13+i*3]: f"qty_{m}",
                        mf.columns[14+i*3]: f"cum_{m}",
                        mf.columns[15+i*3]: f"rate_{m}"}, inplace=True)
for idx_pos, col_name in MCOLS_BASE.items():
    mf.rename(columns={mf.columns[idx_pos]: col_name}, inplace=True, errors="ignore")

hat_mask = (mf["품번"].astype(str).str.startswith("WA25") &
            (mf["품번"].astype(str).str.contains("CA") |
             mf["품번"].astype(str).str.contains("HT") |
             mf["품번"].astype(str).str.contains("BN")))
monthly = mf[hat_mask].copy()
monthly = monthly[monthly["품번"].notna()].copy()

print(f"   판매: {len(sales)}행 / 구조화: {len(struct)}행 / 월별: {len(monthly)}행({monthly['품번'].nunique()}품번)")

# ════════════════════════════════════════════════════════════════════════════
# 2. JOIN — 판매 + 구조화
# ════════════════════════════════════════════════════════════════════════════
print("[2/9] JOIN 처리...")

SCOLS_SEL = ["품번","대분류","활용","로고사이즈","표현방식","스타일분류","볼캡유형","브림","크라운","패널수","소재","소재코드"]
df = sales.merge(struct[SCOLS_SEL], on="품번", how="left")

# 날짜 변환
for col in ["최초입고일","최종입고일","최초출고일","최종출고일"]:
    df[col] = pd.to_datetime(df[col], errors="coerce")

# ── 파생 컬럼
df["판매경과일"] = (df["최종출고일"] - df["최초출고일"]).dt.days
df["할인여부"]   = df["최초판매가"] != df["현판매가"]
df["택가실현율"] = (df["누계판매_실판매금액"] / df["발주_현판매금액"] * 100).where(
    df["발주_현판매금액"] > 0, np.nan).round(2)

COLOR_BASIC  = {"BK","NA","CH","GY","WH"}
COLOR_NEUTRL = {"BE","CR","IB","KH","OL","MT"}
COLOR_EARTH  = {"BR","TN","CA"}
def color_grp(c):
    c = str(c).upper().strip()
    if c in COLOR_BASIC:  return "베이직(BK/NA/GY)"
    if c in COLOR_NEUTRL: return "뉴트럴(BE/CR)"
    if c in COLOR_EARTH:  return "어스(BR/TN)"
    return "컬러풀"
df["색상계열"] = df["색상"].apply(color_grp)
df["출고월"]   = df["최초출고월"].astype(str).str[:7]

def markup_grade(v):
    if pd.isna(v): return "N/A"
    return "HIGH(≥300%)" if v >= 300 else "MID(200~300%)" if v >= 200 else "LOW(<200%)"
df["마크업등급"] = df["판매이익율_사전V"].apply(markup_grade)

unmatched = df[df["대분류"].isna()]["품번"].unique()
if len(unmatched): print(f"   ⚠ 미매칭 {len(unmatched)}개: {list(unmatched)[:5]}")
else: print("   ✓ 전체 매칭 완료")

# ════════════════════════════════════════════════════════════════════════════
# 3. 월별 데이터 — SKU별 집계
# ════════════════════════════════════════════════════════════════════════════
print("[3/9] 월별 데이터 집계...")

qty_cols = [f"qty_{m}" for m in MONTHS]
for col in qty_cols:
    if col not in monthly.columns:
        monthly[col] = 0
    monthly[col] = pd.to_numeric(monthly[col], errors="coerce").fillna(0)

# 품번별 월별 판매수량 합산
monthly_sku = monthly.groupby("품번")[qty_cols].sum().reset_index()
monthly_sku.columns = ["품번"] + MONTHS

# 주 판매시기 분석 지표
def peak_month(row):
    series = row[MONTHS]
    if series.sum() == 0: return "N/A"
    return series.idxmax()

def ss_ratio(row):
    """SS 시즌(2025-01~06) 판매 비중"""
    ss_months = [m for m in MONTHS if m.startswith("2025-0")]
    total = row[MONTHS].sum()
    if total == 0: return 0
    return round(row[ss_months].sum() / total * 100, 1)

def sales_concentration(row):
    """판매 집중도 — HHI 변형 (1=1개월 집중, ~0=분산)"""
    series = row[MONTHS].values.astype(float)
    total = series.sum()
    if total == 0: return 0
    shares = series / total
    return round(float((shares ** 2).sum()), 3)

def early_ratio(row):
    """초기 3개월 판매 비중 (2025-01~03)"""
    early = ["2025-01","2025-02","2025-03"]
    total = row[MONTHS].sum()
    if total == 0: return 0
    return round(row[early].sum() / total * 100, 1)

def sales_pattern(row):
    """판매 패턴 분류"""
    er = early_ratio(row)
    ssr = ss_ratio(row)
    if er >= 50: return "초반집중"
    if ssr >= 80: return "SS집중"
    if ssr >= 50: return "SS중심"
    return "연중분산"

monthly_sku["피크월"]      = monthly_sku.apply(peak_month, axis=1)
monthly_sku["SS비율pct"]   = monthly_sku.apply(ss_ratio, axis=1)
monthly_sku["초기비율pct"] = monthly_sku.apply(early_ratio, axis=1)
monthly_sku["집중도HHI"]   = monthly_sku.apply(sales_concentration, axis=1)
monthly_sku["판매패턴"]    = monthly_sku.apply(sales_pattern, axis=1)
monthly_sku["월별총계"]    = monthly_sku[MONTHS].sum(axis=1)

# 피크월 월별 레이블 (YYYY-MM → Mon'YY)
def fmt_month(m):
    if m == "N/A": return m
    yy = m[2:4]; mm = m[5:7]
    labels = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
               "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}
    return f"{labels.get(mm,'?')}'{yy}"

monthly_sku["피크월_레이블"] = monthly_sku["피크월"].apply(fmt_month)

print(f"   월별분석 완료: {len(monthly_sku)}개 품번")

# ════════════════════════════════════════════════════════════════════════════
# 4. SKU 집계 (판매자료 기준)
# ════════════════════════════════════════════════════════════════════════════
print("[4/9] SKU 집계...")

sku = df.groupby("품번").agg(
    품명=("품명","first"), 대분류=("대분류","first"), 활용=("활용","first"),
    스타일=("스타일분류","first"), 볼캡유형=("볼캡유형","first"),
    소재=("소재","first"), 소재코드=("소재코드","first"),
    발주수량=("발주_수량","sum"), 판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"), 실판매금액=("누계판매_실판매금액","sum"),
    판매이익=("판매이익_금액","sum"), 최초판매가=("최초판매가","first"),
    현판매가=("현판매가","first"), 완사입비=("완사입비_단가","first"),
    마크업율_avg=("판매이익율_사전V","mean"), 색상수=("색상","nunique"),
    최초출고일=("최초출고일","min"), 최종출고일=("최종출고일","max"),
).reset_index()

sku["소화율_pct"]    = (sku["판매수량"] / sku["발주수량"] * 100).round(1)
sku["택가실현율_pct"]= (sku["실판매금액"] / (sku["발주수량"] * sku["최초판매가"]) * 100).round(1)
sku["판매이익율_pct"]= (sku["판매이익"] / sku["실판매금액"] * 100).round(1)
sku["판매경과일"]    = (sku["최종출고일"] - sku["최초출고일"]).dt.days

def classify(row):
    if row["소화율_pct"] >= 70 and row["마크업율_avg"] >= 250: return "🟢 MUST"
    if row["소화율_pct"] >= 70:                                 return "🟡 WATCH(VE)"
    if row["마크업율_avg"] >= 250:                               return "🟡 WATCH(QR)"
    return "🔴 SKIP"

sku["26FW판별"] = sku.apply(classify, axis=1)

# 월별 데이터 JOIN
sku = sku.merge(monthly_sku[["품번","피크월","피크월_레이블","SS비율pct","초기비율pct","집중도HHI","판매패턴","월별총계"]+MONTHS],
                on="품번", how="left")

sku_sorted = sku.sort_values("판매금액", ascending=False).reset_index(drop=True)

must_cnt  = (sku["26FW판별"] == "🟢 MUST").sum()
watch_cnt = (sku["26FW판별"].str.startswith("🟡")).sum()
skip_cnt  = (sku["26FW판별"] == "🔴 SKIP").sum()
total_sku = len(sku)

# KPI
total_orders  = df["발주_수량"].sum()
total_sales_q = df["누계판매_수량"].sum()
total_sales_a = df["누계판매_금액"].sum()
total_sales_r = df["누계판매_실판매금액"].sum()
avg_sell_thru = total_sales_q / total_orders * 100 if total_orders else 0
avg_markup    = df["판매이익율_사전V"].mean()
total_profit  = df["판매이익_금액"].sum()

print(f"   KPI — 총판매금액: {total_sales_a/1e8:.2f}억 / 소화율: {avg_sell_thru:.1f}% / 마크업: {avg_markup:.0f}%")

# ── 카테고리 집계
cat_main = df.groupby("대분류").agg(SKU수=("품번","nunique"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean"),마크업avg=("판매이익율_사전V","mean")).reset_index().sort_values("판매금액",ascending=False)

cat_style2 = df.groupby("볼캡유형").agg(SKU수=("품번","nunique"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean"),마크업avg=("판매이익율_사전V","mean")).reset_index().sort_values("판매금액",ascending=False)

cat_material = df.groupby("소재코드").agg(SKU수=("품번","nunique"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean"),마크업avg=("판매이익율_사전V","mean")).reset_index().sort_values("판매금액",ascending=False)

# ── IP 집계
ip_df = df.groupby("활용").agg(SKU수=("품번","nunique"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean"),마크업avg=("판매이익율_사전V","mean")).reset_index().sort_values("판매금액",ascending=False)
ip_df["비중pct"] = (ip_df["판매금액"]/ip_df["판매금액"].sum()*100).round(1)

df["IP구분"] = df["활용"].apply(lambda x: "키키IP" if str(x) in ["키키","릴리","IP"] else "비IP")
ip_bin = df.groupby("IP구분").agg(SKU수=("품번","nunique"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean")).reset_index()
ip_bin["비중pct"] = (ip_bin["판매금액"]/ip_bin["판매금액"].sum()*100).round(1)

# ── 색상 집계
color_grp_df = df.groupby("색상계열").agg(건수=("품번","count"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean")).reset_index().sort_values("판매금액",ascending=False)

# ── 출고월 집계
timing = df.groupby("출고월").agg(색상건수=("색상","count"),판매수량=("누계판매_수량","sum"),
    판매금액=("누계판매_금액","sum"),소화율avg=("소화율_수량pct","mean"),품번수=("품번","nunique")).reset_index().sort_values("출고월")

# ── 피크월 분포
peak_dist = monthly_sku["피크월"].value_counts().sort_index()

# ── 판매패턴 분포
pattern_dist = monthly_sku["판매패턴"].value_counts()

print("   집계 완료")

# ════════════════════════════════════════════════════════════════════════════
# 5. Excel 10시트 작성
# ════════════════════════════════════════════════════════════════════════════
print("[5/9] Excel 10시트 작성...")

wb = Workbook()
wb.remove(wb.active)

def sheet_title(ws, text, ncols, bg=C_HEADER):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = text
    c.fill = fill(bg); c.font = fnt(True, C_WHITE, 13); c.alignment = ctr()
    ws.row_dimensions[1].height = 30
    ws.sheet_view.showGridLines = False

# ── S1: 종합요약
ws1 = wb.create_sheet("📊 종합요약")
sheet_title(ws1, "🧢 와키윌리 25SS 모자 판매 종합 분석 | 2026-03-19", 8)
ws1.merge_cells("A2:H2")
c=ws1["A2"]; c.value="판매자료 162행 × 구조화 59SKU × 월별추이 15개월 통합 분석"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,9); c.alignment=ctr()

# KPI 상단
kpis = [("총 발주수량",f"{int(total_orders):,} 개"),("총 판매금액",f"{int(total_sales_a):,} 원"),
        ("평균 소화율",f"{avg_sell_thru:.1f} %"),("평균 마크업율",f"{avg_markup:.0f} %")]
hdr_row(ws1,4,[k for k,v in kpis], bg=C_ACCENT)
data_row(ws1,5,[v for k,v in kpis])
ws1.row_dimensions[5].height = 26

kpis2 = [("실판매금액",f"{int(total_sales_r):,} 원"),("판매이익 합계",f"{int(total_profit):,} 원"),
         ("분석 SKU",f"{total_sku} 개"),("키키IP 비중",f"{ip_bin[ip_bin['IP구분']=='키키IP']['비중pct'].values[0] if len(ip_bin)>0 else 0:.1f} %")]
hdr_row(ws1,7,[k for k,v in kpis2],bg=C_SUB)
data_row(ws1,8,[v for k,v in kpis2])

hdr_row(ws1,10,["26FW판별","SKU수","비중"],bg=C_ACCENT)
for ri,(lb,cnt,bg_) in enumerate([(f"🟢 MUST — 재편성",must_cnt,C_MUST),
                                    (f"🟡 WATCH — 조건부",watch_cnt,C_WATCH),
                                    (f"🔴 SKIP — 제외",skip_cnt,C_SKIP)],11):
    data_row(ws1,ri,[lb,cnt,f"{cnt/total_sku*100:.0f}%"],bg=bg_)

hdr_row(ws1,15,["판매패턴","SKU수"],bg=C_SUB)
for ri,(pat,cnt) in enumerate(pattern_dist.items(),16):
    data_row(ws1,ri,[pat,int(cnt)])

hdr_row(ws1,21,["피크 출고월 TOP5","SKU수"],bg=C_ACCENT)
for ri,(m,cnt) in enumerate(peak_dist.head(5).items(),22):
    data_row(ws1,ri,[m,int(cnt)])

set_widths(ws1,[30,18,14,14,14,14,14,14])

# ── S2: SKU 랭킹
ws2 = wb.create_sheet("🏆 SKU랭킹")
sheet_title(ws2,"🏆 품번별 판매 랭킹 — 판매금액·소화율·마크업·주판매시기",14)
h2=["순위","품번","품명","대분류","활용","볼캡","소재",
    "발주수량","판매수량","소화율%","판매금액(원)","마크업%","피크월","26FW판별"]
hdr_row(ws2,2,h2)
for ri,(_, r) in enumerate(sku_sorted.iterrows(),3):
    data_row(ws2,ri,[ri-2,r["품번"],r["품명"],r["대분류"],r["활용"],r["볼캡유형"],r["소재"],
                     int(r["발주수량"]),int(r["판매수량"]),f"{r['소화율_pct']:.1f}%",
                     int(r["판매금액"]),f"{r['마크업율_avg']:.0f}%",
                     r.get("피크월_레이블","N/A"), r["26FW판별"]], bg=row_bg(r["26FW판별"]))
set_widths(ws2,[5,14,26,10,8,12,8,10,10,9,16,10,10,14])
ws2.freeze_panes="A3"

# ── S3: 판매타이밍
ws3 = wb.create_sheet("📅 판매타이밍")
sheet_title(ws3,"📅 최초출고월별 판매 타이밍 분석",7)
hdr_row(ws3,2,["최초출고월","색상건수","품번수","판매수량","판매금액(원)","소화율avg%","구간"])
구간MAP = {
    "2024-11":"초기선입고","2024-12":"초기선입고",
    "2025-01":"SS초반","2025-02":"SS초반","2025-03":"SS초반",
    "2025-04":"SS피크","2025-05":"SS피크",
    "2025-06":"재고소진","2025-07":"재고소진",
}
for ri,r in enumerate(timing.itertuples(),3):
    구간 = 구간MAP.get(str(r.출고월),"기타")
    bg_ = {"초기선입고":C_LIGHT,"SS초반":"E8F5E9","SS피크":"D4EDDA","재고소진":C_WATCH}.get(구간,None)
    data_row(ws3,ri,[r.출고월,r.색상건수,r.품번수,int(r.판매수량),int(r.판매금액),f"{r.소화율avg:.1f}%",구간],bg=bg_)
set_widths(ws3,[14,10,10,12,18,12,12])

# ── S4: 색상분석
ws4 = wb.create_sheet("🎨 색상분석")
sheet_title(ws4,"🎨 색상계열별 소화율 분석",5)
hdr_row(ws4,2,["색상계열","건수","판매수량","판매금액(원)","소화율avg%"],bg=C_ACCENT)
for ri,r in enumerate(color_grp_df.itertuples(),3):
    data_row(ws4,ri,[r.색상계열,r.건수,int(r.판매수량),int(r.판매금액),f"{r.소화율avg:.1f}%"])
set_widths(ws4,[20,10,14,18,14])

# ── S5: 카테고리분석
ws5 = wb.create_sheet("🗂️ 카테고리분석")
sheet_title(ws5,"🗂️ 대분류 × 볼캡유형 × 소재별 분석",6)
hdr_row(ws5,2,["대분류","SKU수","판매수량","판매금액(원)","소화율%","마크업avg%"],bg=C_ACCENT)
for ri,r in enumerate(cat_main.itertuples(),3):
    data_row(ws5,ri,[str(r.대분류) if pd.notna(r.대분류) else "-",int(r.SKU수),int(r.판매수량),int(r.판매금액),f"{r.소화율avg:.1f}%",f"{r.마크업avg:.0f}%"])
r0 = len(cat_main)+5
hdr_row(ws5,r0,["볼캡유형","SKU수","판매수량","판매금액(원)","소화율%","마크업avg%"],bg=C_SUB)
for ri,r in enumerate(cat_style2.itertuples(),r0+1):
    data_row(ws5,ri,[str(r.볼캡유형) if pd.notna(r.볼캡유형) else "-",int(r.SKU수),int(r.판매수량),int(r.판매금액),f"{r.소화율avg:.1f}%",f"{r.마크업avg:.0f}%"])
r1 = r0+len(cat_style2)+3
hdr_row(ws5,r1,["소재코드","SKU수","판매수량","판매금액(원)","소화율%","마크업avg%"],bg=C_ACCENT)
for ri,r in enumerate(cat_material.itertuples(),r1+1):
    data_row(ws5,ri,[str(r.소재코드) if pd.notna(r.소재코드) else "-",int(r.SKU수),int(r.판매수량),int(r.판매금액),f"{r.소화율avg:.1f}%",f"{r.마크업avg:.0f}%"])
set_widths(ws5,[18,8,12,18,12,12])

# ── S6: IP분석
ws6 = wb.create_sheet("🐱 IP분석")
sheet_title(ws6,"🐱 IP 활용 유형별 판매 성과",7)
hdr_row(ws6,2,["활용유형","SKU수","판매수량","판매금액(원)","비중%","소화율%","마크업avg%"],bg=C_ACCENT)
for ri,r in enumerate(ip_df.itertuples(),3):
    data_row(ws6,ri,[str(r.활용) if pd.notna(r.활용) else "미분류",int(r.SKU수),int(r.판매수량),int(r.판매금액),f"{r.비중pct:.1f}%",f"{r.소화율avg:.1f}%",f"{r.마크업avg:.0f}%"])
r2=len(ip_df)+5
hdr_row(ws6,r2,["IP구분","SKU수","판매수량","판매금액(원)","비중%","소화율%"],bg=C_SUB)
for ri,r in enumerate(ip_bin.itertuples(),r2+1):
    data_row(ws6,ri,[r.IP구분,int(r.SKU수),int(r.판매수량),int(r.판매금액),f"{r.비중pct:.1f}%",f"{r.소화율avg:.1f}%"])
set_widths(ws6,[14,8,12,18,10,12,12])

# ── S7: 마진구조
ws7 = wb.create_sheet("💰 마진구조")
sheet_title(ws7,"💰 마진 구조 — 마크업 분포 & 택가실현율",9)
hdr_row(ws7,2,["품번","품명","최초판매가","현판매가","할인","소화율%","마크업avg%","택가실현율%","마크업등급"])
discount_map   = df.groupby("품번")["할인여부"].any().to_dict()
markup_gr_map  = df.groupby("품번")["마크업등급"].first().to_dict()
for ri,(_, r) in enumerate(sku_sorted.iterrows(),3):
    mg = markup_gr_map.get(r["품번"],"N/A")
    bg_ = C_MUST if "HIGH" in str(mg) else (C_SKIP if "LOW" in str(mg) else None)
    data_row(ws7,ri,[r["품번"],r["품명"],f"{int(r['최초판매가']):,}",f"{int(r['현판매가']):,}",
                     "Y" if discount_map.get(r["품번"],False) else "-",
                     f"{r['소화율_pct']:.1f}%",f"{r['마크업율_avg']:.0f}%",
                     f"{r['택가실현율_pct']:.1f}%" if pd.notna(r['택가실현율_pct']) else "-",mg],bg=bg_)
set_widths(ws7,[14,28,12,12,6,10,12,12,18])
ws7.freeze_panes="A3"

# ── S8: 26FW시사점
ws8 = wb.create_sheet("🎯 26FW시사점")
sheet_title(ws8,"🎯 26FW 기획 시사점 — MUST/WATCH/SKIP + 판매패턴",9)
ws8.merge_cells("A2:I2")
c=ws8["A2"]; c.value="기준: 소화율 ≥70% × 마크업 ≥250% (사전) | 피크월·판매패턴 포함"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,9); c.alignment=ctr()
hdr_row(ws8,3,["26FW","품번","품명","대분류","활용","소화율%","마크업%","피크월","판매패턴"])
ACT = {"🟢 MUST":"재편성 — 소재/컬러 업그레이드",
       "🟡 WATCH(VE)":"소재 VE + 택가 재검토",
       "🟡 WATCH(QR)":"소량 시작 + QR 전환",
       "🔴 SKIP":"재고소진 후 26FW 제외"}
for ri,(_, r) in enumerate(sku_sorted.iterrows(),4):
    판별=r["26FW판별"]
    data_row(ws8,ri,[판별,r["품번"],r["품명"],r["대분류"],r["활용"],
                     f"{r['소화율_pct']:.1f}%",f"{r['마크업율_avg']:.0f}%",
                     r.get("피크월_레이블","N/A"),r.get("판매패턴","N/A")],bg=row_bg(판별))
set_widths(ws8,[16,14,24,10,8,10,10,10,12])
ws8.freeze_panes="A4"

# ── S9: 월별판매추이 히트맵
ws9 = wb.create_sheet("📈 월별판매추이")
ws9.sheet_view.showGridLines = False
ws9.merge_cells(f"A1:{get_column_letter(len(MONTHS)+3)}1")
c=ws9["A1"]; c.value="📈 품번별 월별 판매수량 추이 히트맵 (2025-01 ~ 2026-03)"
c.fill=fill(C_HEADER); c.font=fnt(True,C_WHITE,13); c.alignment=ctr(); ws9.row_dimensions[1].height=30

# 헤더
hdr_vals = ["품번","품명","판매패턴"] + MONTHS + ["합계"]
hdr_row(ws9,2,hdr_vals)
# 월 레이블 alt row
MONTH_LABELS = [fmt_month(m) for m in MONTHS]
for ci,lb in enumerate(["","",""] + MONTH_LABELS + [""],1):
    c=ws9.cell(row=3,column=ci,value=lb)
    c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,8); c.alignment=ctr(); c.border=brd()
ws9.row_dimensions[3].height=14

# 데이터 — 품번 순 (판매금액 내림차순)
display_skus = sku_sorted[["품번","품명"]].copy()
mx_val = monthly_sku[MONTHS].values.max() if len(monthly_sku) else 1

for ri,(_, sr) in enumerate(display_skus.iterrows(), 4):
    품번 = sr["품번"]
    품명 = str(sr["품명"])[:18]
    mrow = monthly_sku[monthly_sku["품번"]==품번]
    패턴 = mrow["판매패턴"].values[0] if len(mrow) else "N/A"

    ws9.cell(row=ri,column=1,value=품번).border=brd()
    ws9.cell(row=ri,column=2,value=품명).border=brd()
    ws9.cell(row=ri,column=3,value=패턴).border=brd()
    for ci,m in enumerate(MONTHS,4):
        val = int(mrow[m].values[0]) if len(mrow) and m in mrow.columns else 0
        hc = heat_color(val, mx_val)
        c = ws9.cell(row=ri,column=ci,value=val if val>0 else "")
        c.fill=fill(hc); c.font=fnt(False,C_DARK if val<mx_val*0.6 else C_WHITE,9)
        c.alignment=ctr(); c.border=brd()
    total = int(mrow["월별총계"].values[0]) if len(mrow) else 0
    ws9.cell(row=ri,column=len(MONTHS)+4,value=total).border=brd()

ws9.freeze_panes="D4"
col_ws = [16,20,10] + [8]*len(MONTHS) + [8]
set_widths(ws9, col_ws)

# ── S10: 주판매시기 상세
ws10 = wb.create_sheet("⏱️ 주판매시기")
sheet_title(ws10,"⏱️ 품번별 주 판매시기 & 판매집중도 분석",9)
hdr_row(ws10,2,["품번","품명","대분류","활용","피크월","SS비율%","초기3M비율%","집중도HHI","판매패턴"])
ws10.merge_cells("A3:I3")
c=ws10["A3"]; c.value="피크월=최다판매월 | SS비율=2025-01~06 판매비중 | 집중도HHI: 1.0=1개월집중/0.07=15개월균등"
c.fill=fill(C_SUB); c.font=fnt(False,C_WHITE,8); c.alignment=ctr()

sku_timing = sku_sorted.merge(monthly_sku[["품번","피크월","피크월_레이블","SS비율pct","초기비율pct","집중도HHI","판매패턴"]],
                               on="품번",how="left",suffixes=("","_m"))
for ri,(_, r) in enumerate(sku_timing.iterrows(),4):
    패턴 = str(r.get("판매패턴_m", r.get("판매패턴","N/A")))
    bg_ = {"초반집중":"E3F2FD","SS집중":C_MUST,"SS중심":"E8F5E9","연중분산":C_WATCH}.get(패턴, None)
    data_row(ws10,ri,[r["품번"],r["품명"],r["대분류"],r["활용"],
                      r.get("피크월_레이블_m", r.get("피크월_레이블","N/A")),
                      f"{r.get('SS비율pct_m', r.get('SS비율pct',0)):.1f}%",
                      f"{r.get('초기비율pct_m', r.get('초기비율pct',0)):.1f}%",
                      f"{r.get('집중도HHI_m', r.get('집중도HHI',0)):.3f}",패턴], bg=bg_)
set_widths(ws10,[14,26,10,8,10,10,12,12,12])
ws10.freeze_panes="A4"

wb.save(OUT_XLSX)
print(f"   ✓ Excel 저장: {os.path.basename(OUT_XLSX)}")

# ════════════════════════════════════════════════════════════════════════════
# 6. HTML 리포트
# ════════════════════════════════════════════════════════════════════════════
print("[6/9] HTML 리포트 생성...")

# ── Chart.js용 데이터 직렬화
top15 = sku_sorted.head(15)
chart_sku_labels = json.dumps([str(r) for r in top15["품번"]], ensure_ascii=False)
chart_sku_sales  = json.dumps([int(v) for v in top15["판매금액"]])
chart_sku_sell   = json.dumps([float(v) for v in top15["소화율_pct"]])
chart_sku_colors = json.dumps([
    ("#4CAF50" if "MUST" in str(r) else ("#FF9800" if "WATCH" in str(r) else "#F44336"))
    for r in top15["26FW판별"]
])

# 월별 추이 — TOP5 SKU
top5_codes = list(sku_sorted.head(5)["품번"])
monthly_chart = {}
for code in top5_codes:
    mrow = monthly_sku[monthly_sku["품번"]==code]
    if len(mrow):
        monthly_chart[code] = [int(mrow[m].values[0]) for m in MONTHS]
    else:
        monthly_chart[code] = [0]*len(MONTHS)
chart_monthly_labels = json.dumps(MONTHS)
chart_monthly_datasets = []
palette = ["#E94560","#0F3460","#533483","#E94560","#16213E"]
for i,(code,vals) in enumerate(monthly_chart.items()):
    short_name = str(sku_sorted[sku_sorted["품번"]==code]["품명"].values[0])[:16] if len(sku_sorted[sku_sorted["품번"]==code]) else code
    chart_monthly_datasets.append({
        "label": short_name,
        "data": vals,
        "borderColor": palette[i % len(palette)],
        "backgroundColor": palette[i % len(palette)] + "33",
        "tension": 0.4,
        "fill": False,
    })
chart_monthly_data = json.dumps(chart_monthly_datasets, ensure_ascii=False)

# IP 도넛
ip_clean = ip_df[ip_df["활용"].notna()].copy()
ip_clean["활용_str"] = ip_clean["활용"].astype(str)
chart_ip_labels = json.dumps(list(ip_clean["활용_str"]), ensure_ascii=False)
chart_ip_data   = json.dumps([int(v) for v in ip_clean["판매금액"]])

# 대분류 도넛
cat_clean = cat_main[cat_main["대분류"].notna()].copy()
chart_cat_labels = json.dumps([str(v) for v in cat_clean["대분류"]], ensure_ascii=False)
chart_cat_data   = json.dumps([int(v) for v in cat_clean["판매금액"]])

# 색상계열 바
chart_color_labels = json.dumps([str(v) for v in color_grp_df["색상계열"]], ensure_ascii=False)
chart_color_sell   = json.dumps([float(v) for v in color_grp_df["소화율avg"]])
chart_color_sales  = json.dumps([int(v) for v in color_grp_df["판매금액"]])

# 버블 차트 (소화율 × 마크업 × 판매금액)
bubble_data = []
for _, r in sku_sorted.iterrows():
    판별 = r["26FW판별"]
    color = "#4CAF50" if "MUST" in str(판별) else ("#FF9800" if "WATCH" in str(판별) else "#F44336")
    bubble_data.append({
        "x": round(float(r["소화율_pct"]),1) if pd.notna(r["소화율_pct"]) else 0,
        "y": round(float(r["마크업율_avg"]),0) if pd.notna(r["마크업율_avg"]) else 0,
        "r": max(4, min(30, int(r["판매금액"]) // 5000000)),
        "label": str(r["품번"]),
        "bg": color,
        "판별": str(판별),
    })
chart_bubble = json.dumps(bubble_data, ensure_ascii=False)

# 피크월 분포 바
peak_months_sorted = sorted(peak_dist.index.tolist())
chart_peak_labels = json.dumps(peak_months_sorted)
chart_peak_data   = json.dumps([int(peak_dist.get(m,0)) for m in peak_months_sorted])

# MUST/WATCH/SKIP 파이
chart_판별_labels = json.dumps(["MUST","WATCH","SKIP"], ensure_ascii=False)
chart_판별_data   = json.dumps([int(must_cnt),int(watch_cnt),int(skip_cnt)])

# 히트맵 HTML 생성 (품번 × 월)
heatmap_rows_html = []
mx_heat = monthly_sku[MONTHS].values.max() if len(monthly_sku) else 1

for _, sr in display_skus.iterrows():
    품번 = sr["품번"]
    품명 = str(sr["품명"])[:20]
    mrow = monthly_sku[monthly_sku["품번"]==품번]
    패턴 = mrow["판매패턴"].values[0] if len(mrow) else "N/A"
    판별_val = sku_sorted[sku_sorted["품번"]==품번]["26FW판별"].values
    판별_str = str(판별_val[0]) if len(판별_val) else ""
    판별_cls = "must" if "MUST" in 판별_str else ("skip" if "SKIP" in 판별_str else "watch")

    cells = f'<td class="sku-code {판별_cls}">{품번}</td>'
    cells += f'<td class="sku-name">{품명}</td>'
    cells += f'<td class="pattern">{패턴}</td>'
    total = 0
    for m in MONTHS:
        val = int(mrow[m].values[0]) if len(mrow) and m in mrow.columns else 0
        total += val
        if val == 0:
            cells += '<td class="heat-cell heat-zero">-</td>'
        else:
            ratio = val / mx_heat
            r_c = int(250 + (233-250)*ratio)
            g_c = int(251 + (69-251)*ratio)
            b_c = int(255 + (96-255)*ratio)
            bg_hex = f"#{r_c:02x}{g_c:02x}{b_c:02x}"
            txt_color = "#fff" if ratio > 0.6 else "#1a1a2e"
            cells += f'<td class="heat-cell" style="background:{bg_hex};color:{txt_color}">{val:,}</td>'
    cells += f'<td class="heat-total">{total:,}</td>'
    heatmap_rows_html.append(f"<tr>{cells}</tr>")

heatmap_html = "\n".join(heatmap_rows_html)
heatmap_month_headers = "".join(f'<th>{fmt_month(m)}</th>' for m in MONTHS)

# SKU 테이블 TOP15
top_table_rows = []
for _, r in sku_sorted.head(20).iterrows():
    판별 = r["26FW판별"]
    cls = "must" if "MUST" in str(판별) else ("skip" if "SKIP" in str(판별) else "watch")
    피크 = r.get("피크월_레이블","N/A")
    패턴 = r.get("판매패턴","N/A")
    top_table_rows.append(
        f'<tr>'
        f'<td class="sku-code {cls}">{r["품번"]}</td>'
        f'<td class="sku-name">{str(r["품명"])[:22]}</td>'
        f'<td>{str(r["대분류"]) if pd.notna(r["대분류"]) else "-"}</td>'
        f'<td>{str(r["활용"]) if pd.notna(r["활용"]) else "-"}</td>'
        f'<td>{str(r["볼캡유형"]) if pd.notna(r["볼캡유형"]) else "-"}</td>'
        f'<td class="num">{int(r["판매금액"]):,}</td>'
        f'<td class="num">{r["소화율_pct"]:.1f}%</td>'
        f'<td class="num">{r["마크업율_avg"]:.0f}%</td>'
        f'<td>{피크}</td>'
        f'<td>{패턴}</td>'
        f'<td class="badge {cls}">{판별}</td>'
        f'</tr>'
    )
top_table_html = "\n".join(top_table_rows)

# 26FW 액션 테이블
action_rows = []
actions = [
    ("P1","🟢","MUST SKU " + str(must_cnt) + "개 26FW 라인업 확정","MD","2026-04"),
    ("P1","🐱","키키/릴리 IP 신규 스타일 추가 기획","크리에이티브","2026-04"),
    ("P2","🔧","WATCH(VE) SKU 완사입비 재협상 / 소재 대체","프로덕트랩","2026-04"),
    ("P2","📦","컬러풀계열 소량+QR 전환 구조 설계","MD","2026-04"),
    ("P3","🔴","SKIP SKU 재고 소진 프로모션 계획","마케팅","2026-05"),
    ("P3","📅","피크월 기준 26FW 발주 일정 역산","MD/프로덕트랩","2026-04"),
]
for p,ico,act,담당,기한 in actions:
    action_rows.append(f"<tr><td class='pri-{p[-1]}'>{p}</td><td>{ico}</td><td>{act}</td><td>{담당}</td><td>{기한}</td></tr>")
action_html = "\n".join(action_rows)

# ── HTML 조립
html = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>와키윌리 25SS 모자 판매 분석 리포트</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
  :root {
    --navy: #1A1A2E; --accent: #E94560; --sub: #16213E;
    --light: #F0F4FF; --must: #4CAF50; --skip: #F44336;
    --watch: #FF9800; --white: #fff; --grey: #f5f6fa;
    --border: #e0e4f0; --text: #2c2c3e;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: "맑은 고딕", "Malgun Gothic", sans-serif; background: var(--grey); color: var(--text); font-size: 13px; }

  /* ── 헤더 */
  .hero { background: linear-gradient(135deg, var(--navy) 0%, var(--sub) 60%, var(--accent) 100%);
          padding: 36px 40px 28px; color: var(--white); }
  .hero h1 { font-size: 26px; font-weight: 800; letter-spacing: -0.5px; margin-bottom: 6px; }
  .hero .sub { font-size: 13px; opacity: 0.75; }
  .hero .tags { margin-top: 12px; display: flex; gap: 8px; flex-wrap: wrap; }
  .tag { background: rgba(255,255,255,0.15); border-radius: 20px; padding: 3px 12px; font-size: 11px; }

  /* ── 레이아웃 */
  .container { max-width: 1400px; margin: 0 auto; padding: 24px; }
  .section { background: var(--white); border-radius: 12px; padding: 24px; margin-bottom: 24px;
             box-shadow: 0 2px 12px rgba(26,26,46,0.06); }
  .section-title { font-size: 16px; font-weight: 700; color: var(--navy); margin-bottom: 16px;
                   padding-bottom: 10px; border-bottom: 2px solid var(--accent); display: flex; align-items: center; gap: 8px; }

  /* ── KPI 카드 */
  .kpi-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }
  .kpi-card { background: linear-gradient(135deg, var(--navy), var(--sub));
              border-radius: 10px; padding: 20px; color: var(--white); text-align: center; }
  .kpi-card .val { font-size: 24px; font-weight: 800; margin: 8px 0 4px; color: #fff; }
  .kpi-card .lbl { font-size: 11px; opacity: 0.7; }
  .kpi-card.accent { background: linear-gradient(135deg, var(--accent), #c0392b); }
  .kpi-card.green  { background: linear-gradient(135deg, #27ae60, #1e8449); }

  /* ── 판별 뱃지 */
  .badge       { font-size: 11px; border-radius: 4px; padding: 2px 6px; font-weight: 600; white-space: nowrap; }
  .badge.must  { background: #e8f5e9; color: #2e7d32; }
  .badge.watch { background: #fff3e0; color: #e65100; }
  .badge.skip  { background: #ffebee; color: #c62828; }

  /* ── 테이블 */
  .data-table { width: 100%; border-collapse: collapse; font-size: 12px; }
  .data-table th { background: var(--navy); color: var(--white); padding: 9px 10px;
                   text-align: center; font-weight: 600; white-space: nowrap; }
  .data-table td { padding: 7px 10px; text-align: center; border-bottom: 1px solid var(--border); }
  .data-table tr:hover { background: var(--light); }
  .data-table .sku-code { font-family: monospace; font-size: 11px; font-weight: 600; }
  .data-table .sku-name { text-align: left; }
  .data-table .num { text-align: right; font-variant-numeric: tabular-nums; }
  .must { color: var(--must); }
  .skip { color: var(--skip); }
  .watch { color: var(--watch); }

  /* ── 히트맵 */
  .heatmap-wrap { overflow-x: auto; }
  .heatmap-table { border-collapse: collapse; font-size: 11px; min-width: 900px; }
  .heatmap-table th { background: var(--navy); color: var(--white); padding: 7px 6px;
                      text-align: center; white-space: nowrap; font-size: 10px; }
  .heatmap-table th.sku-h { min-width: 110px; }
  .heatmap-table th.name-h { min-width: 130px; }
  .heatmap-table .sku-code { font-family: monospace; font-weight: 700; font-size: 10px; text-align: left; padding: 5px 8px; }
  .heatmap-table .sku-name { text-align: left; padding: 5px 6px; white-space: nowrap; }
  .heatmap-table .pattern { font-size: 10px; white-space: nowrap; }
  .heatmap-table .heat-cell { text-align: center; padding: 4px 3px; width: 48px; font-size: 10px; border: 1px solid #eee; }
  .heatmap-table .heat-zero { background: #fafbff; color: #ccc; text-align: center; padding: 4px 3px; border: 1px solid #eee; }
  .heatmap-table .heat-total { font-weight: 700; background: var(--light); text-align: right; padding: 4px 8px; }
  .heatmap-table tr:hover td { filter: brightness(0.96); }

  /* ── 차트 그리드 */
  .chart-grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }
  .chart-grid-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; }
  .chart-box { position: relative; height: 260px; }
  .chart-box-tall { position: relative; height: 340px; }
  .chart-box-wide { position: relative; height: 300px; }

  /* ── 판별 요약 카드 */
  .판별-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin-bottom: 20px; }
  .판별-card { border-radius: 10px; padding: 16px; text-align: center; }
  .판별-card.must  { background: #e8f5e9; border: 2px solid var(--must); }
  .판별-card.watch { background: #fff3e0; border: 2px solid var(--watch); }
  .판별-card.skip  { background: #ffebee; border: 2px solid var(--skip); }
  .판별-card .big { font-size: 32px; font-weight: 800; }
  .판별-card .lbl { font-size: 12px; font-weight: 600; margin-top: 4px; }
  .판별-card .desc { font-size: 10px; color: #666; margin-top: 4px; }

  /* ── 액션 테이블 */
  .action-table td { padding: 10px 12px; }
  .pri-1 { background: var(--accent); color: var(--white); border-radius: 4px; padding: 2px 8px; font-weight: 700; }
  .pri-2 { background: var(--watch); color: var(--white); border-radius: 4px; padding: 2px 8px; font-weight: 700; }
  .pri-3 { background: #78909c; color: var(--white); border-radius: 4px; padding: 2px 8px; font-weight: 700; }

  /* ── 푸터 */
  .footer { text-align: center; padding: 24px; color: #999; font-size: 11px; }

  @media (max-width: 900px) {
    .kpi-grid, .chart-grid-2, .chart-grid-3, .판별-grid { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>

<!-- ──────────────── HERO ──────────────── -->
<div class="hero">
  <h1>🧢 와키윌리 25SS 모자 판매 분석 리포트</h1>
  <div class="sub">담당: 데이터 인텔리전스 — 트렌드 애널리스트 + 인사이트 아키텍트</div>
  <div class="tags">
    <span class="tag">📅 분석기준: 2026-03-19</span>
    <span class="tag">📊 판매자료 162행</span>
    <span class="tag">🏗 구조화 59SKU</span>
    <span class="tag">📈 월별추이 15개월</span>
    <span class="tag">🎯 시즌: 25SS</span>
  </div>
</div>

<div class="container">

<!-- ──────────────── KPI 카드 ──────────────── -->
<div class="section">
  <div class="section-title">📌 핵심 KPI</div>
  <div class="kpi-grid">
    <div class="kpi-card accent"><div class="lbl">총 누계판매금액</div><div class="val">KPI_SALES_A억</div><div class="lbl">원</div></div>
    <div class="kpi-card"><div class="lbl">평균 소화율(수량)</div><div class="val">KPI_SELL_THRU%</div><div class="lbl">발주 대비 판매</div></div>
    <div class="kpi-card"><div class="lbl">평균 마크업율(사전)</div><div class="val">KPI_MARKUP%</div><div class="lbl">사전 V</div></div>
    <div class="kpi-card green"><div class="lbl">키키 IP 판매 비중</div><div class="val">KPI_KIKI%</div><div class="lbl">전체 대비</div></div>
  </div>
</div>

<!-- ──────────────── 26FW 판별 ──────────────── -->
<div class="section">
  <div class="section-title">🎯 26FW 기획 판별 — MUST / WATCH / SKIP</div>
  <div class="판별-grid">
    <div class="판별-card must">
      <div class="big" style="color:var(--must)">MUST_CNT</div>
      <div class="lbl">🟢 MUST — 재편성 확정</div>
      <div class="desc">소화율 ≥70% × 마크업 ≥250%</div>
    </div>
    <div class="판별-card watch">
      <div class="big" style="color:var(--watch)">WATCH_CNT</div>
      <div class="lbl">🟡 WATCH — 조건부 유지</div>
      <div class="desc">VE 검토 or QR 전환 필요</div>
    </div>
    <div class="판별-card skip">
      <div class="big" style="color:var(--skip)">SKIP_CNT</div>
      <div class="lbl">🔴 SKIP — 26FW 제외</div>
      <div class="desc">재고 소진 후 단종</div>
    </div>
  </div>
  <div class="chart-grid-3">
    <div class="chart-box"><canvas id="chart판별"></canvas></div>
    <div class="chart-box"><canvas id="chartIP"></canvas></div>
    <div class="chart-box"><canvas id="chartCat"></canvas></div>
  </div>
</div>

<!-- ──────────────── SKU 랭킹 바차트 ──────────────── -->
<div class="section">
  <div class="section-title">🏆 SKU 판매 랭킹 (Top 15)</div>
  <div class="chart-box-tall"><canvas id="chartSKU"></canvas></div>
  <div style="overflow-x:auto; margin-top:20px;">
    <table class="data-table">
      <thead><tr>
        <th>품번</th><th>품명</th><th>대분류</th><th>활용</th><th>유형</th>
        <th>판매금액(원)</th><th>소화율%</th><th>마크업%</th><th>피크월</th><th>판매패턴</th><th>26FW</th>
      </tr></thead>
      <tbody>TOP_TABLE_ROWS</tbody>
    </table>
  </div>
</div>

<!-- ──────────────── 월별 판매 추이 ──────────────── -->
<div class="section">
  <div class="section-title">📈 월별 판매수량 추이 — Top5 SKU (2025-01 ~ 2026-03)</div>
  <div class="chart-box-wide"><canvas id="chartMonthly"></canvas></div>
</div>

<!-- ──────────────── 버블 차트 ──────────────── -->
<div class="section">
  <div class="section-title">🔵 소화율 × 마크업 매트릭스 (버블크기=판매금액)</div>
  <div style="font-size:11px;color:#888;margin-bottom:12px;">우상단(소화율↑·마크업↑) = MUST | 버블 크기 = 판매금액 규모 | <span style="color:var(--must)">●MUST</span> <span style="color:var(--watch)">●WATCH</span> <span style="color:var(--skip)">●SKIP</span></div>
  <div class="chart-box-wide"><canvas id="chartBubble"></canvas></div>
</div>

<!-- ──────────────── 색상 & 피크월 분석 ──────────────── -->
<div class="section">
  <div class="section-title">🎨 색상계열 소화율 & ⏱ 피크 판매월 분포</div>
  <div class="chart-grid-2">
    <div class="chart-box"><canvas id="chartColor"></canvas></div>
    <div class="chart-box"><canvas id="chartPeak"></canvas></div>
  </div>
</div>

<!-- ──────────────── 월별 히트맵 ──────────────── -->
<div class="section">
  <div class="section-title">🌡️ 품번 × 월별 판매수량 히트맵</div>
  <div style="font-size:11px;color:#888;margin-bottom:10px;">색상 진할수록 해당 월 판매수량 많음 | 품번 컬러: <span class="must">●MUST</span> <span class="watch">●WATCH</span> <span class="skip">●SKIP</span></div>
  <div class="heatmap-wrap">
    <table class="heatmap-table">
      <thead>
        <tr>
          <th class="sku-h">품번</th>
          <th class="name-h">품명</th>
          <th>패턴</th>
          HEATMAP_MONTH_HEADERS
          <th>합계</th>
        </tr>
      </thead>
      <tbody>
HEATMAP_ROWS
      </tbody>
    </table>
  </div>
</div>

<!-- ──────────────── 26FW 액션 플랜 ──────────────── -->
<div class="section">
  <div class="section-title">🚀 26FW 기획 액션 플랜</div>
  <table class="data-table action-table">
    <thead><tr><th>우선순위</th><th>아이콘</th><th>액션</th><th>담당</th><th>기한</th></tr></thead>
    <tbody>ACTION_ROWS</tbody>
  </table>
</div>

</div><!-- /container -->

<div class="footer">Generated by FPOF 데이터 인텔리전스 | 와키윌리 패션 하우스 오케스트레이션 | 2026-03-19</div>

<!-- ──────────────── Chart.js Scripts ──────────────── -->
<script>
const COLORS_PALETTE = ['#E94560','#0F3460','#533483','#16213E','#1a6b4a','#c0392b','#8e44ad','#2471a3'];
const MONTHS = CHART_MONTHLY_LABELS;

// ── SKU 랭킹 바차트
new Chart(document.getElementById('chartSKU'), {
  type: 'bar',
  data: {
    labels: CHART_SKU_LABELS,
    datasets: [{
      label: '판매금액(원)',
      data: CHART_SKU_SALES,
      backgroundColor: CHART_SKU_COLORS,
      borderRadius: 6,
    }]
  },
  options: {
    responsive: true, maintainAspectRatio: false, indexAxis: 'y',
    plugins: { legend: { display: false }, tooltip: { callbacks: {
      label: ctx => `  ${Number(ctx.raw).toLocaleString()}원`
    }}},
    scales: {
      x: { ticks: { callback: v => (v/1e8).toFixed(1)+'억' }, grid: { color: '#f0f0f0' }},
      y: { ticks: { font: { size: 10 } }}
    }
  }
});

// ── 월별 추이 라인차트
const monthlyDatasets = CHART_MONTHLY_DATA;
new Chart(document.getElementById('chartMonthly'), {
  type: 'line',
  data: { labels: MONTHS, datasets: monthlyDatasets },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: { legend: { position: 'bottom', labels: { font: { size: 11 } }}},
    scales: {
      y: { beginAtZero: true, grid: { color: '#f0f0f0' },
           ticks: { callback: v => v.toLocaleString() }},
      x: { ticks: { font: { size: 10 } }}
    }
  }
});

// ── 26FW 판별 도넛
new Chart(document.getElementById('chart판별'), {
  type: 'doughnut',
  data: {
    labels: CHART_JUMEOK_LABELS,
    datasets: [{ data: CHART_JUMEOK_DATA,
      backgroundColor: ['#4CAF50','#FF9800','#F44336'],
      borderWidth: 2, borderColor: '#fff' }]
  },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: { legend: { position: 'bottom', labels: { font: { size: 10 } }},
               title: { display: true, text: '26FW 판별 분포', font: { size: 12 } }}
  }
});

// ── IP 도넛
new Chart(document.getElementById('chartIP'), {
  type: 'doughnut',
  data: {
    labels: CHART_IP_LABELS,
    datasets: [{ data: CHART_IP_DATA,
      backgroundColor: COLORS_PALETTE,
      borderWidth: 2, borderColor: '#fff' }]
  },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: { legend: { position: 'bottom', labels: { font: { size: 10 } }},
               title: { display: true, text: 'IP 활용별 판매금액', font: { size: 12 } }}
  }
});

// ── 대분류 도넛
new Chart(document.getElementById('chartCat'), {
  type: 'doughnut',
  data: {
    labels: CHART_CAT_LABELS,
    datasets: [{ data: CHART_CAT_DATA,
      backgroundColor: ['#1A1A2E','#E94560','#533483','#16213E'],
      borderWidth: 2, borderColor: '#fff' }]
  },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: { legend: { position: 'bottom', labels: { font: { size: 10 } }},
               title: { display: true, text: '대분류별 판매금액', font: { size: 12 } }}
  }
});

// ── 색상계열 바차트
new Chart(document.getElementById('chartColor'), {
  type: 'bar',
  data: {
    labels: CHART_COLOR_LABELS,
    datasets: [
      { label: '소화율(%)', data: CHART_COLOR_SELL, backgroundColor: '#E94560aa', yAxisID: 'y', borderRadius: 4 },
      { label: '판매금액', data: CHART_COLOR_SALES, backgroundColor: '#1A1A2Eaa', yAxisID: 'y1', borderRadius: 4 },
    ]
  },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: { legend: { position: 'bottom', labels: { font: { size: 10 } }},
               title: { display: true, text: '색상계열별 소화율 & 판매금액', font: { size: 12 } }},
    scales: {
      y:  { position: 'left',  max: 100, ticks: { callback: v => v+'%' }, grid: { color: '#f0f0f0' }},
      y1: { position: 'right', ticks: { callback: v => (v/1e8).toFixed(1)+'억' }, grid: { drawOnChartArea: false }},
    }
  }
});

// ── 피크월 분포 바차트
new Chart(document.getElementById('chartPeak'), {
  type: 'bar',
  data: {
    labels: CHART_PEAK_LABELS,
    datasets: [{ label: 'SKU 수', data: CHART_PEAK_DATA,
      backgroundColor: '#0F3460cc', borderRadius: 6 }]
  },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: { legend: { display: false },
               title: { display: true, text: '품번별 피크 판매월 분포', font: { size: 12 } }},
    scales: {
      y: { beginAtZero: true, ticks: { stepSize: 1 }, grid: { color: '#f0f0f0' }},
      x: { ticks: { font: { size: 9 } }}
    }
  }
});

// ── 버블 차트
const bubbleRaw = CHART_BUBBLE;
const bubbleDatasets = [
  { label:'MUST',  data: bubbleRaw.filter(d=>d.판별.includes('MUST')).map(d=>({x:d.x,y:d.y,r:d.r,label:d.label})),  backgroundColor:'rgba(76,175,80,0.6)',  borderColor:'rgba(76,175,80,0.9)' },
  { label:'WATCH', data: bubbleRaw.filter(d=>d.판별.includes('WATCH')).map(d=>({x:d.x,y:d.y,r:d.r,label:d.label})), backgroundColor:'rgba(255,152,0,0.6)', borderColor:'rgba(255,152,0,0.9)' },
  { label:'SKIP',  data: bubbleRaw.filter(d=>d.판별.includes('SKIP')).map(d=>({x:d.x,y:d.y,r:d.r,label:d.label})),  backgroundColor:'rgba(244,67,54,0.6)',  borderColor:'rgba(244,67,54,0.9)' },
];
new Chart(document.getElementById('chartBubble'), {
  type: 'bubble',
  data: { datasets: bubbleDatasets },
  options: {
    responsive: true, maintainAspectRatio: false,
    plugins: {
      legend: { position: 'bottom', labels: { font: { size: 11 } }},
      tooltip: { callbacks: {
        label: ctx => `${ctx.raw.label}  소화율:${ctx.raw.x}%  마크업:${ctx.raw.y}%`
      }}
    },
    scales: {
      x: { title: { display: true, text: '소화율 (%)' }, min: 0, max: 110,
           ticks: { callback: v => v+'%' }, grid: { color: '#f0f0f0' }},
      y: { title: { display: true, text: '마크업율 (%)' },
           ticks: { callback: v => v+'%' }, grid: { color: '#f0f0f0' }}
    }
  }
});
</script>
</body>
</html>
"""

# ── 데이터 주입
kiki_pct_val = ip_bin[ip_bin["IP구분"]=="키키IP"]["비중pct"].values
kiki_pct_str = f"{kiki_pct_val[0]:.1f}" if len(kiki_pct_val) else "0"

html = (html
    .replace("KPI_SALES_A억",   f"{total_sales_a/1e8:.1f}")
    .replace("KPI_SELL_THRU",   f"{avg_sell_thru:.1f}")
    .replace("KPI_MARKUP",      f"{avg_markup:.0f}")
    .replace("KPI_KIKI",        kiki_pct_str)
    .replace("MUST_CNT",        str(must_cnt))
    .replace("WATCH_CNT",       str(watch_cnt))
    .replace("SKIP_CNT",        str(skip_cnt))
    .replace("CHART_SKU_LABELS",         chart_sku_labels)
    .replace("CHART_SKU_SALES",          chart_sku_sales)
    .replace("CHART_SKU_COLORS",         chart_sku_colors)
    .replace("CHART_MONTHLY_LABELS",     chart_monthly_labels)
    .replace("CHART_MONTHLY_DATA",       chart_monthly_data)
    .replace("CHART_JUMEOK_LABELS",      chart_판별_labels)
    .replace("CHART_JUMEOK_DATA",        chart_판별_data)
    .replace("CHART_IP_LABELS",          chart_ip_labels)
    .replace("CHART_IP_DATA",            chart_ip_data)
    .replace("CHART_CAT_LABELS",         chart_cat_labels)
    .replace("CHART_CAT_DATA",           chart_cat_data)
    .replace("CHART_COLOR_LABELS",       chart_color_labels)
    .replace("CHART_COLOR_SELL",         chart_color_sell)
    .replace("CHART_COLOR_SALES",        chart_color_sales)
    .replace("CHART_PEAK_LABELS",        chart_peak_labels)
    .replace("CHART_PEAK_DATA",          chart_peak_data)
    .replace("CHART_BUBBLE",             chart_bubble)
    .replace("HEATMAP_MONTH_HEADERS",    heatmap_month_headers)
    .replace("HEATMAP_ROWS",             heatmap_html)
    .replace("TOP_TABLE_ROWS",           top_table_html)
    .replace("ACTION_ROWS",              action_html)
)

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(html)
print(f"   ✓ HTML 저장: {os.path.basename(OUT_HTML)}")

# ════════════════════════════════════════════════════════════════════════════
# 7. .fpof-state.json 업데이트
# ════════════════════════════════════════════════════════════════════════════
print("[7/9] 상태 업데이트...")
import json as _json

STATE_PATH = os.path.join(BASE, ".fpof-state.json")
with open(STATE_PATH, "r", encoding="utf-8") as f:
    state = _json.load(f)

if "operational" not in state: state["operational"] = {}
if "recent_outputs" not in state["operational"]: state["operational"]["recent_outputs"] = []

# 기존 항목 제거 후 갱신
state["operational"]["recent_outputs"] = [
    e for e in state["operational"]["recent_outputs"]
    if "check_sales-analysis_2026-03-19" not in str(e)
]
state["operational"]["recent_outputs"].append({
    "date": "2026-03-19",
    "type": "check",
    "version": "v2",
    "files": [
        "output/26FW/season-strategy/check_sales-analysis_2026-03-19.xlsx",
        "output/26FW/season-strategy/check_sales-analysis_2026-03-19.html",
    ],
    "summary": f"25SS 모자 판매분석v2 — SKU {total_sku}개, 소화율 {avg_sell_thru:.1f}%, MUST {must_cnt}/WATCH {watch_cnt}/SKIP {skip_cnt}, 월별추이 15개월"
})

with open(STATE_PATH, "w", encoding="utf-8") as f:
    _json.dump(state, f, ensure_ascii=False, indent=2)

print("\n" + "="*60)
print("✅ 분석 v2 완료")
print("="*60)
print(f"  📊 Excel (10시트): {os.path.basename(OUT_XLSX)}")
print(f"  🌐 HTML 리포트:    {os.path.basename(OUT_HTML)}")
print(f"\n  KPI:")
print(f"  - 총 판매금액:  {total_sales_a/1e8:.2f}억원")
print(f"  - 평균 소화율:  {avg_sell_thru:.1f}%")
print(f"  - 평균 마크업:  {avg_markup:.0f}%")
print(f"  - MUST/WATCH/SKIP: {must_cnt}/{watch_cnt}/{skip_cnt}")
print(f"\n  월별 데이터 (신규):")
print(f"  - 분석 품번: {len(monthly_sku)}개")
print(f"  - 피크 집중 SS비율 평균: {monthly_sku['SS비율pct'].mean():.1f}%")
print(f"  - 판매패턴 분포: {dict(pattern_dist)}")
print("="*60)
